from io import BytesIO
import re
import unittest

from openpyxl import Workbook, load_workbook

from app import create_app, db
from app.enrollment_service import create_enrollment
from app.import_wizard import normalize_student_phone, process_result_import, process_student_import, result_entry_import_template, student_template
from app.models import (
    AcademicClass,
    AcademicLevel,
    AcademicYear,
    AcademicYearClass,
    AcademicYearLevel,
    AcademicYearSubject,
    Exam,
    ExamMarkingConfiguration,
    Result,
    Student,
    Subject,
)


class ExcelImportContractTests(unittest.TestCase):
    class TestConfig:
        TESTING = True
        SECRET_KEY = "excel-import-contract"
        SQLALCHEMY_DATABASE_URI = "sqlite:///:memory:"
        SQLALCHEMY_TRACK_MODIFICATIONS = False
        WTF_CSRF_ENABLED = False

    def setUp(self):
        self.app = create_app(self.TestConfig)
        self.ctx = self.app.app_context()
        self.ctx.push()
        db.drop_all()
        db.create_all()

        self.year = AcademicYear(name="2031-2032", is_current=True)
        self.level = AcademicLevel(name="Secondary", sort_order=1)
        self.legacy_class = AcademicClass(name="Form Four", academic_level=self.level, sort_order=1)
        self.old_subject = Subject(name="Old English", academic_level=self.level, max_score=100)
        db.session.add_all([self.year, self.level, self.legacy_class, self.old_subject])
        db.session.flush()
        self.year_level = AcademicYearLevel(
            academic_year_id=self.year.id,
            legacy_level_id=self.level.id,
            name="Secondary",
            sort_order=1,
        )
        self.year_class = AcademicYearClass(
            academic_year_level=self.year_level,
            legacy_class_id=self.legacy_class.id,
            name="Form Four",
            sort_order=1,
        )
        self.year_subject = AcademicYearSubject(
            academic_year_id=self.year.id,
            academic_year_level=self.year_level,
            legacy_subject_id=self.old_subject.id,
            name="Current English",
            max_score=10,
            sort_order=1,
        )
        self.exam = Exam(name="1st Monthly", academic_year_id=self.year.id, is_active=True)
        db.session.add_all([self.year_level, self.year_class, self.year_subject, self.exam])
        db.session.flush()
        db.session.add(ExamMarkingConfiguration(
            academic_year_id=self.year.id,
            academic_year_level_id=self.year_level.id,
            exam_id=self.exam.id,
            default_full_marks=10,
        ))
        self.student = Student(
            student_code="TIS-001",
            full_name="Amina Ali",
            mother_name="Sahra Jama",
            gender="Female",
            academic_year_id=self.year.id,
            academic_level_id=self.level.id,
            academic_class_id=self.legacy_class.id,
        )
        db.session.add(self.student)
        db.session.flush()
        create_enrollment(
            self.student.id,
            self.year.id,
            self.year_level.id,
            self.year_class.id,
            enrollment_source="import",
        )
        db.session.commit()

    def tearDown(self):
        db.session.remove()
        db.drop_all()
        self.ctx.pop()

    def _workbook(self, score, year_name=None, display_headers=False):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Result Entry"
        sheet.append([
            "#", "Student ID", "Full Name", "Mother Name", "Class",
            "Exam Type", "Academic Year", "Current English",
        ] if display_headers else [
            "#", "student_id", "full_name", "mother_name", "class",
            "exam_type", "academic_year", "Current English",
        ])
        sheet.append([
            1, self.student.student_code, self.student.full_name,
            self.student.mother_name, self.year_class.name, self.exam.name,
            year_name or self.year.name, score,
        ])
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return stream

    def test_result_template_uses_renamed_year_subject_and_no_fake_row(self):
        workbook = result_entry_import_template(
            year_id=self.year.id,
            exam_id=self.exam.id,
            level_id=self.level.id,
            class_id=self.legacy_class.id,
        )
        sheet = workbook["Result Entry"]
        self.assertEqual(sheet.cell(row=1, column=8).value, "Current English")
        self.assertEqual(sheet.cell(row=2, column=2).value, self.student.student_code)
        self.assertEqual(sheet.cell(row=2, column=8).value, "")

    def test_result_import_uses_exact_scope_and_exam_maximum(self):
        summary = process_result_import(
            self._workbook(9),
            year_id=self.year.id,
            exam_id=self.exam.id,
            level_id=self.level.id,
            class_id=self.legacy_class.id,
        )
        self.assertEqual(summary["success_count"], 1, summary)
        self.assertEqual(summary["failed_count"], 0, summary)
        result = Result.query.filter_by(
            student_id=self.student.id,
            exam_id=self.exam.id,
            subject_id=self.old_subject.id,
        ).one()
        self.assertEqual(float(result.score), 9.0)

    def test_result_import_accepts_display_headers_for_exam_and_year(self):
        summary = process_result_import(
            self._workbook(9, display_headers=True),
            year_id=self.year.id,
            exam_id=self.exam.id,
            level_id=self.level.id,
            class_id=self.legacy_class.id,
        )
        self.assertEqual(summary["success_count"], 1, summary)
        self.assertEqual(summary["failed_count"], 0, summary)

    def test_student_phone_formats_normalize_to_one_value(self):
        expected = "+252611234567"
        for value in ("+252611234567", "252611234567", "611234567", "0611234567", "00252611234567"):
            self.assertEqual(normalize_student_phone(value), expected)

    def test_student_phone_import_accepts_other_common_prefixes_and_formats(self):
        for value in ("+252 63 123 4567", "252-65-1234567", "0671234567", "+1 (555) 123-4567"):
            normalized = normalize_student_phone(value)
            self.assertGreaterEqual(len(re.sub(r"\D", "", normalized)), 5)

    def test_student_import_accepts_supplied_phone_formats(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Students"
        sheet.append(["ID", "Name", "Mother", "Mobile", "Academic Level", "Class", "Section", "Academic Year", "Gender", "Photo Source"])
        for index, phone in enumerate(("+252 63 123 4567", "252-65-1234567", "0671234567", "+1 (555) 123-4567"), start=1):
            sheet.append([
                f"PHONE-{index}", f"Student {index}", "Parent Name", phone,
                "Secondary", "Form Four", "", self.year.name, "Female", "",
            ])
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)

        summary = process_student_import(stream)
        self.assertEqual(summary["success_count"], 4, summary)
        self.assertFalse(any("phone number invalid" in error for error in summary["errors"]), summary)

    def test_result_api_finds_student_id_without_case_matching(self):
        response = self.app.test_client().get("/api/results/tis-001")
        self.assertEqual(response.status_code, 404)
        self.assertEqual(response.get_json()["message"], "No published result.")

    def test_result_import_rejects_score_above_selected_exam_maximum(self):
        summary = process_result_import(
            self._workbook(11),
            year_id=self.year.id,
            exam_id=self.exam.id,
            level_id=self.level.id,
            class_id=self.legacy_class.id,
        )
        self.assertEqual(summary["success_count"], 0, summary)
        self.assertEqual(summary["failed_count"], 1, summary)
        self.assertIn("between 0 and 10", " ".join(summary["errors"]))
        self.assertEqual(Result.query.count(), 0)

    def test_result_import_rejects_row_from_another_academic_year(self):
        summary = process_result_import(
            self._workbook(9, year_name="2032-2033"),
            year_id=self.year.id,
            exam_id=self.exam.id,
            level_id=self.level.id,
            class_id=self.legacy_class.id,
        )
        self.assertEqual(summary["success_count"], 0, summary)
        self.assertEqual(summary["failed_count"], 1, summary)
        self.assertIn("does not exist in system", " ".join(summary["errors"]))

    def test_result_template_does_not_prefill_when_exam_belongs_to_another_year(self):
        other_year = AcademicYear(name="2032-2033", is_current=False)
        other_exam = Exam(name="Final", academic_year=other_year, is_active=True)
        db.session.add_all([other_year, other_exam])
        db.session.commit()

        workbook = result_entry_import_template(
            year_id=self.year.id,
            exam_id=other_exam.id,
            level_id=self.level.id,
            class_id=self.legacy_class.id,
        )
        sheet = workbook["Result Entry"]
        self.assertEqual(sheet.cell(row=2, column=2).value, None)
        self.assertIn("matching Academic Year", sheet.cell(row=2, column=1).value)

    def test_student_template_lists_year_aware_setup_values(self):
        workbook = student_template()
        sheet = workbook["Setup Values"]
        self.assertEqual(
            [cell.value for cell in sheet[1]],
            ["Academic Year", "Academic Level", "Class", "Section"],
        )
        self.assertIn(
            (self.year.name, self.year_level.name, self.year_class.name, ""),
            list(sheet.iter_rows(min_row=2, values_only=True)),
        )


if __name__ == "__main__":
    unittest.main()
