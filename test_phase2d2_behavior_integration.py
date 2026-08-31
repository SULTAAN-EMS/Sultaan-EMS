"""Phase 2D.2 coverage for session grading and Behavior integration."""

import unittest
from io import BytesIO
from decimal import Decimal

from app import db
from app.behavior_grading import (
    BehaviorGradeValidationError,
    behavior_grade_for_score,
    validate_behavior_grade_overlap,
    validate_behavior_grade_values,
)
from app.behavior_reporting import get_behavior_report_data
from app.behavior_service import record_event, restore_event, void_event
from app.models import (
    BehaviorAction,
    BehaviorCategory,
    BehaviorEvent,
    BehaviorGradeScale,
)
from openpyxl import load_workbook
from test_phase2e_behavior_reporting import TestPhase2EBehaviorReporting


class TestPhase2D2BehaviorIntegration(TestPhase2EBehaviorReporting):
    def setUp(self):
        super().setUp()
        self.positive = BehaviorCategory(
            configuration=self.configuration,
            name="Phase 2D2 Positive",
            polarity="positive",
            is_active=True,
        )
        self.positive_action = BehaviorAction(
            category=self.positive,
            name="Phase 2D2 Helpful",
            level_number=1,
            points=2,
            frequency="ad_hoc",
            is_active=True,
        )
        self.high_positive_action = BehaviorAction(
            category=self.positive,
            name="Phase 2D2 Outstanding",
            level_number=2,
            points=20,
            frequency="ad_hoc",
            is_active=True,
        )
        db.session.add_all([self.positive, self.positive_action, self.high_positive_action])
        db.session.commit()

    def _scale(self, session, grade, minimum, maximum, point=1, is_pass=True):
        return BehaviorGradeScale(
            configuration=self.configuration,
            session=session,
            grade=grade,
            min_score=minimum,
            max_score=maximum,
            grade_point=point,
            description=f"Behavior {grade}",
            sort_order=int(minimum),
            is_active=True,
            is_pass=is_pass,
        )

    def test_grade_ranges_are_raw_score_and_independent_per_session(self):
        first = self._scale(self.session_one, "A", 0, 25, 4)
        db.session.add(first)
        db.session.flush()

        # Integer-adjacent ranges are valid; genuine overlap is rejected.
        validate_behavior_grade_overlap(self.session_one, Decimal("26"), Decimal("50"))
        with self.assertRaises(BehaviorGradeValidationError):
            validate_behavior_grade_overlap(self.session_one, Decimal("25"), Decimal("30"))

        db.session.add_all([
            self._scale(self.session_one, "B", 26, 50, 3),
            self._scale(self.session_two, "A", 0, 50, 4),
        ])
        db.session.commit()

        self.assertEqual(
            BehaviorGradeScale.query.filter_by(
                behavior_configuration_id=self.configuration.id,
                grade="A",
            ).count(),
            2,
        )
        self.assertEqual(behavior_grade_for_score(self.session_one, 25)["grade"], "A")
        self.assertEqual(behavior_grade_for_score(self.session_one, 50)["grade"], "B")
        self.assertEqual(behavior_grade_for_score(self.session_two, 25)["grade"], "A")

        with self.assertRaises(BehaviorGradeValidationError):
            validate_behavior_grade_values("A", 0, 51, 4, session_maximum=50)

    def test_copying_a_previous_session_scale_creates_independent_rows(self):
        db.session.add_all([
            self._scale(self.session_one, "A", 0, 25, 4),
            self._scale(self.session_one, "B", 26, 50, 3),
        ])
        db.session.commit()
        self.admin.role = "super_admin"
        db.session.commit()

        client = self._client_as_admin()
        response = client.post(
            "/admin/behavior/grade-management",
            data={
                "config_id": self.configuration.id,
                "session_id": self.session_two.id,
                "action": "generate_previous",
                "source_session_id": self.session_one.id,
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)
        copied = BehaviorGradeScale.query.filter_by(
            behavior_session_id=self.session_two.id,
        ).order_by(BehaviorGradeScale.min_score).all()
        self.assertEqual([(row.grade, float(row.min_score), float(row.max_score)) for row in copied], [
            ("A", 0.0, 25.0),
            ("B", 26.0, 50.0),
        ])

        source = BehaviorGradeScale.query.filter_by(
            behavior_session_id=self.session_one.id,
            grade="A",
        ).one()
        source.max_score = 10
        db.session.commit()
        db.session.expire_all()
        target = BehaviorGradeScale.query.filter_by(
            behavior_session_id=self.session_two.id,
            grade="A",
        ).one()
        self.assertEqual(target.max_score, Decimal("25.000"))

    def test_behavior_event_void_restore_updates_the_same_report_projection(self):
        event = record_event(
            self.configuration,
            self.enrollment,
            self.session_one,
            self.positive,
            self.positive_action,
            idempotency_key="phase2d2-restore",
        )
        db.session.commit()

        active_report = get_behavior_report_data(self.student, self.exam_one)[0]
        self.assertEqual(active_report["session_score"], Decimal("27.000"))
        void_event(event, self.admin.id, "Correction")
        db.session.commit()
        void_report = get_behavior_report_data(self.student, self.exam_one)[0]
        self.assertEqual(void_report["session_score"], Decimal("25.000"))

        restore_event(event)
        db.session.commit()
        restored_report = get_behavior_report_data(self.student, self.exam_one)[0]
        self.assertEqual(restored_report["session_score"], Decimal("27.000"))
        self.assertEqual(BehaviorEvent.query.filter_by(status="active").count(), 1)

    def test_action_above_session_maximum_is_visible_but_not_selectable(self):
        self.session_one.maximum_score = 10
        db.session.commit()
        self.admin.role = "super_admin"
        db.session.commit()
        body = self._client_as_admin().get(
            f"/admin/behavior/students?config_id={self.configuration.id}"
            f"&session_id={self.session_one.id}"
        ).get_data(as_text=True)
        self.assertIn("Actions above the selected session maximum are disabled", body)
        self.assertIn("option.disabled = exceedsMaximum", body)
        with self.assertRaises(ValueError):
            record_event(
                self.configuration,
                self.enrollment,
                self.session_one,
                self.positive,
                self.high_positive_action,
                idempotency_key="phase2d2-too-large",
            )

    def test_whole_class_excel_contains_behavior_column_and_combined_total(self):
        self.admin.role = "super_admin"
        db.session.commit()
        response = self._client_as_admin().get(
            "/admin/advanced-results/export-class-excel"
            f"?year_id={self.year_one.id}&exam_id={self.exam_one.id}"
            f"&level_id={self.level_one.id}&class_id={self.class_one.id}"
        )

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.data), read_only=True)
        sheet = workbook["Class Results"]
        rows = sheet.iter_rows(values_only=True)
        headers = list(next(rows))
        row = list(next(rows))
        behavior_index = headers.index("Behavior: Dabeecad")
        self.assertEqual(row[behavior_index], 25.0)
        self.assertEqual(row[headers.index("Total")], 105.0)
        workbook.close()


if __name__ == "__main__":
    unittest.main()
