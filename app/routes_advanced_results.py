from collections import defaultdict
from decimal import Decimal
import math
import json
import re
from tempfile import NamedTemporaryFile
from datetime import date

from flask import Blueprint, abort, current_app, flash, jsonify, redirect, render_template, request, send_file, session, url_for
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from sqlalchemy import and_ as db_and, func, or_ as db_or
from sqlalchemy.orm import selectinload

from . import db
from .audit import audit
from .cloudinary_service import upload_image
from .import_wizard import process_result_import, process_student_import, result_entry_import_template, student_template
from .models import AcademicYear, AcademicClass, AcademicLevel, AcademicSection, AcademicYearClass, AcademicYearLevel, AcademicYearSubject, AttendanceRecord, Exam, ExamType, GradeScale, IncidentReport, Result, SchoolClass, Setting, Student, StudentEnrollment, StudentEnrollmentMovement, Subject, LabelTranslation
from .academic_hierarchy import students_for_year_scope_query, year_classes, year_levels, year_subjects
from .enrollment_service import (
    EnrollmentValidationError,
    apply_legacy_placement,
    create_enrollment,
    enrollment_placement_for_student,
    execute_bulk_transition,
    get_enrollment_for_student_year,
    ensure_legacy_enrollment_for_student,
    plan_bulk_transition,
    resolve_student_academic_context,
    student_enrollment_scope_query,
    transition_student_enrollment,
    validate_enrollment_scope,
)
from .permissions import can, enforce_endpoint_permission
from .security import ALLOWED_PHOTOS, ALLOWED_SHEETS, allowed_file
from .services import DEFAULT_GRADE_SCALES, academic_decimal_precision, academic_round, attendance_uf_subject_keys, competition_rank_lookup, get_label, get_settings, grade_for, grade_for_from_cache, load_grade_scale_cache, performance_tier_for, result_payload, subject_display_name
from .attendance_rules import counts_as_exam_sitting

advanced_results_bp = Blueprint("admin_advanced_results", __name__)


def ordinal(value):
    suffix = "th"
    if value % 100 not in (11, 12, 13):
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(value % 10, "th")
    return f"{value}{suffix}"


def stored_asset_url(path):
    """Return a browser-safe URL for Cloudinary, data, /static, or uploads paths."""
    if not path:
        return None
    value = str(path)
    if value.startswith(("http://", "https://", "data:", "/static/")):
        return value
    if value.startswith("uploads/"):
        return url_for("static", filename=value)
    return url_for("static", filename=f"uploads/{value}")


def _safe_download_name_part(value, fallback):
    """Return one portable filename segment without leaking UI-only characters."""
    invalid_characters = '<>:"/\\\\|?*'
    cleaned = "".join(
        " " if character in invalid_characters or ord(character) < 32 else character
        for character in str(value or "")
    )
    cleaned = " ".join(cleaned.split()).strip(" .")
    return cleaned or fallback


def _school_initials_for_download():
    settings = get_settings()
    school_name = settings.get("school_name") or settings.get("dashboard_title") or "EMS"
    words = re.findall(r"[A-Za-z0-9]+", school_name)
    return "".join(word[0].upper() for word in words[:6]) or "EMS"


def result_workbook_filename(year, exam, academic_class, *, results=False):
    """Build a human-readable workbook name from the currently selected scope."""
    class_name = _safe_download_name_part(getattr(academic_class, "name", None), "Class")
    exam_name = _safe_download_name_part(getattr(exam, "name", None), "Exam")
    year_name = _safe_download_name_part(getattr(year, "name", None), "Academic Year")
    result_prefix = "Natiijada " if results else ""
    return f"{_school_initials_for_download()} - {class_name} - {result_prefix}{exam_name} ({year_name}).xlsx"


@advanced_results_bp.before_request
@login_required
def require_login():
    enforce_endpoint_permission()
    ensure_results_label_seeds()


@advanced_results_bp.app_context_processor
def inject_results_hub_helpers():
    return {"rh_label": get_label, "subject_display_name": subject_display_name}


RESULTS_LABEL_SEEDS = [
    ("hub.brand.school", "so", "Dugsiga — Nidaamka Maamulka", "Shell"),
    ("hub.brand.subtitle", "so", "Results Hub", "Shell"),
    ("hub.user.role", "so", "Maamule · Super Admin", "Shell"),
    ("hub.tab.setup", "so", "Setup", "Shell"),
    ("hub.tab.dashboard", "so", "Dashboard", "Shell"),
    ("hub.tab.entry", "so", "Result Entry", "Shell"),
    ("hub.tab.roster", "so", "Liiska Fasalka", "Shell"),
    ("hub.tab.analytics", "so", "Analytics", "Shell"),
    ("hub.tab.grades", "so", "Grade Mgmt", "Shell"),
    ("hub.tab.settings", "so", "Settings", "Shell"),
    ("dashboard.eyebrow", "so", "Guudmarka", "Dashboard"),
    ("dashboard.title", "so", "Results Dashboard", "Dashboard"),
    ("dashboard.year", "so", "Sanad Dugsiyeedka", "Dashboard"),
    ("dashboard.exam_type", "so", "Exam Type", "Dashboard"),
    ("dashboard.total_students", "so", "Wadarta Ardayda", "Dashboard"),
    ("dashboard.subjects_entered", "so", "Maadooyinka la Xareeyay", "Dashboard"),
    ("dashboard.completion", "so", "Buuxinta Natiijooyinka", "Dashboard"),
    ("dashboard.active_classes", "so", "Fasallada Firfircoon", "Dashboard"),
    ("dashboard.step_level", "so", "Dooro Level", "Dashboard"),
    ("dashboard.step_class", "so", "Dooro Fasalka", "Dashboard"),
    ("dashboard.no_classes", "so", "Fasallo lama helin level-ka la doortay.", "Dashboard"),
    ("dashboard.select_exam_title", "so", "Dooro imtixaan si aad u aragto dashboard-ka.", "Dashboard"),
    ("entry.eyebrow", "so", "Form 4 · Bileedka 2aad · 2025-2026", "Result Entry"),
    ("entry.title", "so", "Whole-Class Result Entry", "Result Entry"),
    ("entry.subtitle", "so", "Geli dhibcaha ardayda oo dhan hal mar — validation-ka max_score ayaa toos u socda.", "Result Entry"),
    ("entry.unsaved", "so", "isbedel oo aan la kaydin", "Result Entry"),
    ("entry.draft", "so", "Draft", "Result Entry"),
    ("entry.published", "so", "Published", "Result Entry"),
    ("entry.save_all", "so", "Save All", "Result Entry"),
    ("entry.id", "so", "ID", "Result Entry"),
    ("entry.name", "so", "Magaca", "Result Entry"),
    ("entry.reset", "so", "Reset", "Result Entry"),
    ("entry.students", "so", "Students", "Result Entry"),
    ("entry.subjects", "so", "Subjects", "Result Entry"),
    ("roster.eyebrow", "so", "Form 4 · Bileedka 2aad · 2025-2026", "Roster"),
    ("roster.title", "so", "Liiska Natiijada Fasalka", "Roster"),
    ("roster.students", "so", "arday", "Roster"),
    ("roster.export_excel", "so", "Export Excel", "Roster"),
    ("roster.export_pdf", "so", "Export PDF", "Roster"),
    ("roster.search", "so", "Raadi arday ID ama magac...", "Roster"),
    ("roster.student_name", "so", "Magaca", "Roster"),
    ("roster.total", "so", "Total", "Roster"),
    ("roster.grade", "so", "Grade", "Roster"),
    ("analytics.eyebrow", "so", "Form 4 · Bileedka 2aad · 2025-2026", "Analytics"),
    ("analytics.title", "so", "Analytics — Infographic", "Analytics"),
    ("analytics.average", "so", "Celceliska Fasalka", "Analytics"),
    ("analytics.gpa", "so", "GPA Fasalka", "Analytics"),
    ("analytics.pass_rate", "so", "Pass Rate", "Analytics"),
    ("analytics.grade_distribution", "so", "Grade Distribution", "Analytics"),
    ("analytics.subject_performance", "so", "Subject-wise Performance", "Analytics"),
    ("analytics.exam_trend", "so", "Exam-Type Trend", "Analytics"),
    ("analytics.performers", "so", "Top & Bottom Performers", "Analytics"),
    ("analytics.top", "so", "Top 5", "Analytics"),
    ("analytics.bottom", "so", "Bottom 5", "Analytics"),
    ("grades.eyebrow", "so", "Grade Management", "Grade Management"),
    ("grades.title", "so", "Grade Management — Simplified", "Grade Management"),
    ("grades.subtitle", "so", "Academic Year → Exam Type → hal scale oo si toos ah loogu dabaqo dhammaan fasallada imtixaankan qaaday.", "Grade Management"),
    ("grades.applies_to", "so", "Applies To", "Grade Management"),
    ("grades.all_classes", "so", "All Classes", "Grade Management"),
    ("grades.generate", "so", "Generate Scale", "Grade Management"),
    ("grades.grade", "so", "Grade", "Grade Management"),
    ("grades.min", "so", "Min", "Grade Management"),
    ("grades.max", "so", "Max", "Grade Management"),
    ("grades.point", "so", "Point", "Grade Management"),
    ("grades.comment", "so", "Comment", "Grade Management"),
    ("grades.preview", "so", "Preview", "Grade Management"),
    ("grades.save", "so", "Save Grade Scales", "Grade Management"),
    ("settings.eyebrow", "so", "Full Customization", "Settings"),
    ("settings.title", "so", "Results Settings", "Settings"),
    ("settings.subtitle", "so", "Halkan waxaad ka bedeli kartaa habka module-ka Results u shaqeeyo — gudaha Results-ka, ma aha Settings-ka guud.", "Settings"),
    ("settings.labels_title", "so", "Label & Language Customization", "Settings"),
    ("settings.labels_desc", "so", "Halkan waxaad ka bedeli kartaa ereyga/label kasta oo systemka ka muuqda.", "Settings"),
    ("settings.default_language", "so", "Luuqadda Default-ka ah", "Settings"),
    ("settings.add_language", "so", "Ku dar Luuqad", "Settings"),
    ("settings.label_key", "so", "Label Key", "Settings"),
    ("settings.context", "so", "Meesha ka muuqato", "Settings"),
    ("settings.save_labels", "so", "Save Labels", "Settings"),
]


def ensure_results_label_seeds():
    if current_app.config.get("_results_label_seeds_checked"):
        return
    changed = False
    for label_key, language_code, text_value, context in RESULTS_LABEL_SEEDS:
        exists = LabelTranslation.query.filter_by(label_key=label_key, language_code=language_code).first()
        if exists:
            continue
        db.session.add(LabelTranslation(
            label_key=label_key,
            language_code=language_code,
            text_value=text_value,
            context=context,
        ))
        changed = True
    if changed:
        db.session.commit()
    current_app.config["_results_label_seeds_checked"] = True


def get_default_academic_year(year_id=None):
    """Return the requested year, otherwise the latest active academic year."""
    if year_id:
        return db.session.get(AcademicYear, year_id)
    return (
        AcademicYear.query.filter_by(is_current=True)
        .order_by(AcademicYear.name.desc(), AcademicYear.id.desc())
        .first()
        or AcademicYear.query.order_by(AcademicYear.name.desc(), AcademicYear.id.desc()).first()
    )


def get_latest_exam_for_year(academic_year):
    """Return the latest active exam for the selected academic year, falling back to any latest exam."""
    if not academic_year:
        return None
    return (
        Exam.query.filter_by(academic_year_id=academic_year.id, is_active=True)
        .order_by(Exam.id.desc())
        .first()
        or
        Exam.query.filter_by(academic_year_id=academic_year.id)
        .order_by(Exam.id.desc())
        .first()
    )


def subjects_for_scope(exam, level_id=None, class_id=None):
    """Return subjects assigned to the selected year-aware level."""
    # An explicitly requested level must win over an exam's optional default
    # level.  Analytics iterates real levels one by one; using the exam default
    # there would incorrectly reuse one level's subjects for every level.
    effective_level_id = level_id or (exam.academic_level_id if exam else None)
    if not effective_level_id and class_id:
        academic_class = db.session.get(AcademicClass, class_id)
        effective_level_id = academic_class.academic_level_id if academic_class else None

    if not effective_level_id:
        return []
    if exam and exam.academic_year_id:
        year_level, _year_class = _year_scope_ids_from_legacy(
            exam.academic_year_id,
            effective_level_id,
            class_id,
        )
        if not year_level:
            return []
        year_items = year_subjects(exam.academic_year_id, year_level.id)
        mapped_subjects = [
            db.session.get(Subject, item.legacy_subject_id)
            for item in year_items
            if item.legacy_subject_id
        ]
        mapped_subjects = [item for item in mapped_subjects if item]
        if mapped_subjects:
            return mapped_subjects
        # A year-scoped exam must never fall back to the global Subject table.
        # An empty year-level subject mapping means the selected context has
        # no configured subjects yet; showing a legacy subject from another
        # year would make the Results Hub look valid while saving the wrong
        # data.
        return []
    return (
        Subject.query.filter_by(academic_level_id=effective_level_id)
        .order_by(Subject.sort_order, Subject.name, Subject.id)
        .all()
    )


def analytics_subject_bridge(academic_year_id, year_level_id=None):
    """Return year-aware subjects and their legacy Result subjects."""
    year_items = year_subjects(academic_year_id, year_level_id)
    legacy_items = []
    seen = set()
    for item in year_items:
        if item.legacy_subject_id and item.legacy_subject_id not in seen:
            legacy_subject = db.session.get(Subject, item.legacy_subject_id)
            if legacy_subject:
                legacy_items.append(legacy_subject)
                seen.add(legacy_subject.id)
    return year_items, legacy_items


def subjects_for_year_level(exam, year_level_id):
    """Resolve the subject bridge for an enrollment's exact year-level."""
    if not exam or not year_level_id:
        return []
    year_level = db.session.get(AcademicYearLevel, year_level_id)
    if not year_level or year_level.academic_year_id != exam.academic_year_id:
        return []
    year_items = year_subjects(exam.academic_year_id, year_level.id)
    subjects = [
        db.session.get(Subject, item.legacy_subject_id)
        for item in year_items
        if item.legacy_subject_id
    ]
    subjects = [item for item in subjects if item]
    # A year-level without a subject bridge is incomplete setup.  Do not
    # widen the report/result view to the legacy global subject set.
    return subjects


def _year_scope_ids_from_legacy(academic_year_id, level_id=None, class_id=None):
    """Map legacy Results Hub selector IDs to the selected year's scopes."""
    year_level = None
    if level_id:
        year_level = AcademicYearLevel.query.filter_by(
            academic_year_id=academic_year_id,
            legacy_level_id=level_id,
        ).first()
    year_class = None
    if class_id:
        class_query = (
            AcademicYearClass.query
            .join(AcademicYearLevel, AcademicYearLevel.id == AcademicYearClass.academic_year_level_id)
            .filter(
                AcademicYearLevel.academic_year_id == academic_year_id,
                AcademicYearClass.legacy_class_id == class_id,
            )
        )
        if year_level:
            class_query = class_query.filter(AcademicYearClass.academic_year_level_id == year_level.id)
        year_class = class_query.first()
        year_level = year_level or (year_class.academic_year_level if year_class else None)
    return year_level, year_class


def students_for_scope_query(academic_year_id, level_id=None, class_id=None, section_id=None, exam=None):
    """Return enrollment-first students for a Results Hub historical scope."""
    effective_level_id = level_id or (exam.academic_level_id if exam else None)
    effective_class_id = class_id or (exam.academic_class_id if exam else None)
    effective_section_id = section_id or (exam.academic_section_id if exam else None)
    year_level, year_class = _year_scope_ids_from_legacy(
        academic_year_id,
        effective_level_id,
        effective_class_id,
    )
    if effective_level_id and not year_level:
        raise EnrollmentValidationError("Selected level is not configured for this academic year")
    if effective_class_id and not year_class:
        raise EnrollmentValidationError("Selected class is not configured for this academic year")
    return student_enrollment_scope_query(
        academic_year_id,
        academic_year_level_id=year_level.id if year_level else None,
        academic_year_class_id=year_class.id if year_class else None,
        academic_section_id=effective_section_id,
    )


def rank_student_in_scope(student, academic_year_id, exam, subjects, level_id=None, class_id=None, section_id=None):
    """Return the student's competition rank within their class/level scope."""
    if not student or not exam or not subjects:
        return "-"

    scoped_students = students_for_scope_query(
        academic_year_id,
        level_id=level_id,
        class_id=class_id,
        # A class rank must include every section of the same class.
        section_id=None,
        exam=exam,
    ).all()
    if not scoped_students:
        return "-"

    student_ids = [row.id for row in scoped_students]
    results = Result.query.filter(
        Result.exam_id == exam.id,
        Result.student_id.in_(student_ids),
    ).all()
    results_by_student = {}
    for result in results:
        results_by_student.setdefault(result.student_id, {})[result.subject_id] = result

    ranked = []
    for scoped_student in scoped_students:
        total_score = 0
        total_max = 0
        student_results = results_by_student.get(scoped_student.id, {})
        for subject in subjects:
            result = student_results.get(subject.id)
            total_score += float(result.score) if result else 0
            total_max += float(subject.max_score)
        percentage = round((total_score / total_max * 100), 2) if total_max > 0 else 0
        ranked.append((scoped_student.id, percentage, total_score))

    rank_lookup = competition_rank_lookup({student_id: percentage for student_id, percentage, _total_score in ranked})
    return rank_lookup.get(student.id, "-")


@advanced_results_bp.route("/")
def dashboard():
    """Results Hub landing page - kept as a compatibility route."""
    return redirect(url_for("admin_advanced_results.new_dashboard"))


@advanced_results_bp.route("/new-setup")
def new_setup():
    """Setup wizard - Master Configuration for the entire Results system - Read Only Dashboard"""
    level_id = int_or_none(request.args.get("level_id"))
    
    # Get selected year (current year by default)
    selected_year = AcademicYear.query.filter_by(is_current=True).first()
    selected_level = db.session.get(AcademicLevel, level_id) if level_id else None
    
    # Get all data for Setup wizard
    years = AcademicYear.query.order_by(AcademicYear.name.desc()).all()
    exams = Exam.query.filter_by(academic_year_id=selected_year.id).order_by(Exam.id.desc()).all() if selected_year else []
    levels = AcademicLevel.query.filter_by(is_active=True).order_by(AcademicLevel.sort_order).all()
    subjects = Subject.query.order_by(Subject.sort_order).all()
    classes = AcademicClass.query.order_by(AcademicClass.name).all()
    
    # Get last configuration update info
    from .models import AuditLog
    last_config_log = AuditLog.query.filter(
        AuditLog.action.like('%Setup%') | 
        AuditLog.action.like('%Configuration%') |
        AuditLog.action.like('%Academic Year%') |
        AuditLog.action.like('%Exam Type%')
    ).order_by(AuditLog.created_at.desc()).first()
    
    last_updated = last_config_log.created_at.strftime('%B %d, %Y at %H:%M') if last_config_log else None
    updated_by = last_config_log.username if last_config_log else 'Unknown'
    
    # Calculate step completion states
    step_states = {
        'academic_year': len(years) > 0,
        'exam_type': len(exams) > 0,
        'levels_classes': False,  # Will check below
        'subjects': False  # Will check below
    }
    
    # Check if any level has at least one class
    total_classes = 0
    for level in levels:
        class_count = level.classes.count()
        total_classes += class_count
        if class_count > 0:
            step_states['levels_classes'] = True
            break
    
    # Check if subjects are configured
    if subjects:
        step_states['subjects'] = True
    
    # Determine current active step
    if not step_states['academic_year']:
        current_step = 'academic_year'
    elif not step_states['exam_type']:
        current_step = 'exam_type'
    elif not step_states['levels_classes']:
        current_step = 'levels_classes'
    elif not step_states['subjects']:
        current_step = 'subjects'
    else:
        current_step = 'subjects'  # All complete
    
    return render_template(
        "admin/results_setup.html",
        years=years,
        exams=exams,
        levels=levels,
        subjects=subjects,
        classes=classes,
        total_classes=total_classes,
        selected_year=selected_year,
        selected_level=selected_level,
        selected_exam=None,
        step_states=step_states,
        current_step=current_step,
        last_updated=last_updated,
        updated_by=updated_by
    )


# Level CRUD Routes for Setup Page
@advanced_results_bp.route("/setup/levels/add", methods=["POST"])
def setup_add_level():
    """Add a new academic level from Setup page"""
    name = request.form.get("name", "").strip()
    
    if not name:
        flash("Level name is required.", "danger")
        return redirect(url_for("admin_advanced_results.new_setup"))
    
    # Get max sort_order
    from sqlalchemy import func
    max_order = db.session.query(func.max(AcademicLevel.sort_order)).scalar() or 0
    
    level = AcademicLevel(
        name=name,
        sort_order=max_order + 1,
        is_active=True
    )
    
    db.session.add(level)
    audit("System Setup", f"Added academic level: {level.name}")
    db.session.commit()
    flash("Level added successfully.", "success")
    return redirect(url_for("admin_advanced_results.new_setup"))


@advanced_results_bp.route("/setup/levels/<int:level_id>/edit", methods=["POST"])
def setup_edit_level(level_id):
    """Edit an academic level from Setup page"""
    level = db.session.get(AcademicLevel, level_id)
    if not level:
        flash("Level not found.", "danger")
        return redirect(url_for("admin_advanced_results.new_setup"))
    
    name = request.form.get("name", "").strip()
    if not name:
        flash("Level name is required.", "danger")
        return redirect(url_for("admin_advanced_results.new_setup"))
    
    level.name = name
    audit("System Setup", f"Edited academic level: {level.name}")
    db.session.commit()
    flash("Level updated successfully.", "success")
    return redirect(url_for("admin_advanced_results.new_setup"))


@advanced_results_bp.route("/setup/levels/<int:level_id>/delete", methods=["POST"])
def setup_delete_level(level_id):
    """Delete an academic level from Setup page"""
    level = db.session.get(AcademicLevel, level_id)
    if not level:
        flash("Level not found.", "danger")
        return redirect(url_for("admin_advanced_results.new_setup"))
    
    # Check if level has classes
    if level.classes.count() > 0:
        flash("Cannot delete level with existing classes.", "danger")
        return redirect(url_for("admin_advanced_results.new_setup"))
    
    db.session.delete(level)
    audit("System Setup", f"Deleted academic level: {level.name}")
    db.session.commit()
    flash("Level deleted successfully.", "success")
    return redirect(url_for("admin_advanced_results.new_setup"))


# Class CRUD Routes for Setup Page
@advanced_results_bp.route("/setup/classes/add", methods=["POST"])
def setup_add_class():
    """Add a new academic class from Setup page"""
    academic_level_id = int(request.form.get("academic_level_id"))
    name = request.form.get("name", "").strip()
    
    if not name:
        flash("Class name is required.", "danger")
        return redirect(url_for("admin_advanced_results.new_setup"))
    
    level = db.session.get(AcademicLevel, academic_level_id)
    if not level:
        flash("Academic level not found.", "danger")
        return redirect(url_for("admin_advanced_results.new_setup"))
    
    # Get max sort_order for this level
    from sqlalchemy import func
    max_order = db.session.query(func.max(AcademicClass.sort_order)).filter_by(
        academic_level_id=academic_level_id
    ).scalar() or 0
    
    cls = AcademicClass(
        academic_level_id=academic_level_id,
        name=name,
        sort_order=max_order + 1,
        is_active=True
    )
    
    db.session.add(cls)
    audit("System Setup", f"Added academic class: {cls.name}")
    db.session.commit()
    flash("Class added successfully.", "success")
    return redirect(url_for("admin_advanced_results.new_setup"))


@advanced_results_bp.route("/setup/classes/<int:class_id>/edit", methods=["POST"])
def setup_edit_class(class_id):
    """Edit an academic class from Setup page"""
    cls = db.session.get(AcademicClass, class_id)
    if not cls:
        flash("Class not found.", "danger")
        return redirect(url_for("admin_advanced_results.new_setup"))
    
    name = request.form.get("name", "").strip()
    if not name:
        flash("Class name is required.", "danger")
        return redirect(url_for("admin_advanced_results.new_setup"))
    
    cls.name = name
    audit("System Setup", f"Edited academic class: {cls.name}")
    db.session.commit()
    flash("Class updated successfully.", "success")
    return redirect(url_for("admin_advanced_results.new_setup"))


@advanced_results_bp.route("/setup/classes/<int:class_id>/delete", methods=["POST"])
def setup_delete_class(class_id):
    """Delete an academic class from Setup page"""
    cls = db.session.get(AcademicClass, class_id)
    if not cls:
        flash("Class not found.", "danger")
        return redirect(url_for("admin_advanced_results.new_setup"))
    
    # Check if class has sections
    if cls.sections.count() > 0:
        flash("Cannot delete class with existing sections.", "danger")
        return redirect(url_for("admin_advanced_results.new_setup"))
    
    db.session.delete(cls)
    audit("System Setup", f"Deleted academic class: {cls.name}")
    db.session.commit()
    flash("Class deleted successfully.", "success")
    return redirect(url_for("admin_advanced_results.new_setup"))


@advanced_results_bp.route("/new-dashboard")
def new_dashboard():
    """New results dashboard with auto-selection of active academic year"""
    year_id = int_or_none(request.args.get("year_id"))
    exam_id = int_or_none(request.args.get("exam_id"))
    level_id = int_or_none(request.args.get("level_id"))
    
    selected_year = get_default_academic_year(year_id)
    selected_exam = db.session.get(Exam, exam_id) if exam_id else get_latest_exam_for_year(selected_year)
    if selected_exam and selected_year and selected_exam.academic_year_id != selected_year.id:
        selected_exam = get_latest_exam_for_year(selected_year)
    # Get all years and exams for selectors
    years = AcademicYear.query.order_by(AcademicYear.name.desc()).all()
    exams = Exam.query.filter_by(academic_year_id=selected_year.id).order_by(Exam.id.desc()).all() if selected_year else []

    # Keep the dashboard selectors and every downstream query inside the
    # selected year's hierarchy. Legacy level IDs are still used by the
    # existing templates, but only mapped IDs are exposed here.
    year_level_scopes = year_levels(selected_year.id) if selected_year else []
    levels = [
        scope.legacy_level
        for scope in year_level_scopes
        if scope.legacy_level and scope.legacy_level.is_active
    ]
    valid_level_ids = {level.id for level in levels}
    selected_level = db.session.get(AcademicLevel, level_id) if level_id in valid_level_ids else None

    exam_scope_ready = True
    if selected_exam and selected_exam.academic_level_id:
        exam_year_level = next(
            (scope for scope in year_level_scopes if scope.legacy_level_id == selected_exam.academic_level_id),
            None,
        )
        exam_scope_ready = exam_year_level is not None
        if exam_scope_ready and selected_exam.academic_class_id:
            exam_year_class = next((
                item for item in year_classes(exam_year_level.id)
                if item.legacy_class_id == selected_exam.academic_class_id
            ), None)
            exam_scope_ready = exam_year_class is not None
            if exam_scope_ready and selected_exam.academic_section_id:
                exam_section = db.session.get(AcademicSection, selected_exam.academic_section_id)
                exam_scope_ready = bool(
                    exam_section
                    and exam_year_class.legacy_class_id == exam_section.academic_class_id
                )
        elif exam_scope_ready and selected_exam.academic_section_id:
            exam_section = db.session.get(AcademicSection, selected_exam.academic_section_id)
            exam_scope_ready = bool(exam_section)
    
    stats = {
        "total_students": 0,
        "total_subjects": 0,
        "completion_percentage": 0,
        "active_classes": 0,
    }
    class_cards = []
    
    if selected_exam and exam_scope_ready:
        stats = build_dashboard_stats(selected_exam)
        class_cards = build_class_cards(selected_exam, level_filter=selected_level)
    
    return render_template(
        "admin/results_dashboard.html",
        years=years,
        exams=exams,
        levels=levels,
        selected_year=selected_year,
        selected_exam=selected_exam,
        selected_level=selected_level,
        exam_scope_ready=exam_scope_ready,
        stats=stats,
        class_cards=class_cards,
        settings=get_settings(),
    )


@advanced_results_bp.route("/class-roster")
def class_roster():
    """Class roster view with all students and their results"""
    import traceback
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        year_id = int_or_none(request.args.get("year_id"))
        exam_id = int_or_none(request.args.get("exam_id"))
        level_id = int_or_none(request.args.get("level_id"))
        class_id = int_or_none(request.args.get("class_id"))
        section_id = int_or_none(request.args.get("section_id"))
        student_id = int_or_none(request.args.get("student_id"))
        search_query = request.args.get("q", "").strip()
    
        selected_year = get_default_academic_year(year_id)
        selected_exam = db.session.get(Exam, exam_id) if exam_id else get_latest_exam_for_year(selected_year)
        if selected_exam and selected_year and selected_exam.academic_year_id != selected_year.id:
            selected_exam = get_latest_exam_for_year(selected_year)
        
        if not selected_year:
            flash("Please select an academic year.", "warning")
            return redirect(url_for("admin_advanced_results.new_dashboard"))

        years = AcademicYear.query.order_by(AcademicYear.name.desc()).all()
        exams = Exam.query.filter_by(academic_year_id=selected_year.id).order_by(Exam.id.desc()).all() if selected_year else []
        year_level_scopes = year_levels(selected_year.id) if selected_year else []
        levels = [
            scope.legacy_level
            for scope in year_level_scopes
            if scope.legacy_level and scope.legacy_level.is_active
        ]
        valid_level_ids = {item.id for item in levels}
        requested_level_id = level_id or (selected_exam.academic_level_id if selected_exam else None)
        level_id = requested_level_id if requested_level_id in valid_level_ids else None
        selected_year_level_scope = next(
            (scope for scope in year_level_scopes if scope.legacy_level_id == level_id),
            None,
        )
        mapped_classes = [
            item.legacy_class
            for item in year_classes(selected_year_level_scope.id)
            if item.legacy_class and item.legacy_class.is_active
        ] if selected_year_level_scope else []
        valid_class_ids = {item.id for item in mapped_classes}
        requested_class_id = class_id or (selected_exam.academic_class_id if selected_exam else None)
        class_id = requested_class_id if requested_class_id in valid_class_ids else None
        requested_section_id = section_id or (selected_exam.academic_section_id if selected_exam else None)
        section_id = requested_section_id
        classes = mapped_classes
        if class_id:
            classes = [item for item in mapped_classes if item.id == class_id]
        sections = AcademicSection.query.filter_by(academic_class_id=class_id, is_active=True).order_by(AcademicSection.sort_order, AcademicSection.name).all() if class_id else []
        if section_id and not any(item.id == section_id for item in sections):
            section_id = None

        scope_info = {
            "level": db.session.get(AcademicLevel, level_id) if level_id else None,
            "class": db.session.get(AcademicClass, class_id) if class_id else None,
            "section": db.session.get(AcademicSection, section_id) if section_id else None,
        }
        
        # If no exam selected, show exam selection interface
        if not selected_exam:
            return render_template(
                "admin/class_roster.html",
                selected_year=selected_year,
                selected_exam=None,
                scope_info=scope_info,
                students=[],
                subjects=[],
                years=years,
                exams=exams,
                levels=levels,
                classes=classes,
                sections=sections,
                search_query=search_query,
                settings=get_settings(),
            )
        
        if not (level_id and class_id):
            return render_template(
                "admin/class_roster.html",
                selected_year=selected_year,
                selected_exam=selected_exam,
                scope_info=scope_info,
                students=[],
                subjects=[],
                years=years,
                exams=exams,
                levels=levels,
                classes=classes,
                sections=sections,
                search_query=search_query,
                settings=get_settings(),
            )
        
        # Use the shared Grade Management snapshot so legacy NULL-active rows,
        # exam-specific ranges, and global fallback ranges resolve identically
        # on the roster, portal, and exported reports.
        grade_cache = load_grade_scale_cache(selected_exam.id)
        report_tiers = get_report_tier_configs(
            year_id=selected_year.id,
            exam_id=selected_exam.id,
            level_id=level_id,
        )
        weak_config = report_tiers["weak"]
        fail_config = report_tiers["fail"]

        def cached_grade_for(score):
            return grade_for_from_cache(score, grade_cache)
        
        student_query = students_for_scope_query(
            selected_year.id,
            level_id=level_id,
            class_id=class_id,
            section_id=section_id,
        )
        
        # Apply search filter
        if search_query:
            search_pattern = f"%{search_query}%"
            student_query = student_query.filter(
                db_or(
                    Student.student_code.like(search_pattern),
                    Student.full_name.like(search_pattern)
                )
            )
        
        students = student_query.order_by(Student.full_name).all()
        
        subjects = subjects_for_scope(selected_exam, level_id=level_id, class_id=class_id)
        attendance_uf_keys = attendance_uf_subject_keys(
            selected_exam,
            [student.id for student in students],
            [subject.id for subject in subjects],
        )
        
        # Build results data for each student
        from .routes_public import feedback_access_token
        roster_data = []
        for student in students:
            # Get results for this student and exam (only published results)
            results = Result.query.filter_by(student_id=student.id, exam_id=selected_exam.id, is_published=True).all()
            results_dict = {r.subject_id: r for r in results}
            
            # Calculate totals and grades
            total_score = 0
            total_max = 0
            subject_data = []
            
            for subject in subjects:
                result = results_dict.get(subject.id)
                score = float(result.score) if result else 0
                max_score = float(subject.max_score)
                percentage = (score / max_score * 100) if max_score > 0 else 0
                
                total_score += score
                total_max += max_score
                
                # Get grade using exam-specific grade scale
                grade_info = cached_grade_for(percentage)
                tier = performance_tier_for(percentage, weak_config, fail_config)
                
                # Apply grade_override if present
                if result and result.grade_override:
                    grade_info = dict(grade_info)
                    grade_info["grade"] = result.grade_override
                
                subject_data.append({
                    "subject": subject,
                    "result": result,
                    "score": score,
                    "max_score": max_score,
                    "percentage": round(percentage, 2),
                    "grade": grade_info,
                    "is_fail": tier["is_fail"],
                    "is_weak": tier["is_weak"],
                    "is_uf": (student.id, subject.id) in attendance_uf_keys,
                })
            
            overall_percentage = round((total_score / total_max * 100), 2) if total_max > 0 else 0
            overall_grade = cached_grade_for(overall_percentage)
            overall_tier = performance_tier_for(overall_percentage, weak_config, fail_config)
            
            # Calculate GP (grade point average)
            total_points = sum(s["grade"]["grade_point"] for s in subject_data if s["grade"]["grade_point"])
            gp = academic_round(total_points / len(subject_data), get_settings()) if subject_data else 0
            
            roster_data.append({
                "student": student,
                "mg_token": feedback_access_token(student, selected_exam),
                "subject_data": subject_data,
                "total_score": total_score,
                "total_max": total_max,
                "percentage": overall_percentage,
                "grade": overall_grade,
                "is_fail": overall_tier["is_fail"],
                "is_weak": overall_tier["is_weak"],
                "gp": gp,
            })
        
        return render_template(
            "admin/class_roster.html",
            selected_year=selected_year,
            selected_exam=selected_exam,
            scope_info=scope_info,
            students=roster_data,
            subjects=subjects,
            years=years,
            exams=exams,
            levels=levels,
            classes=classes,
            sections=sections,
            search_query=search_query,
            settings=get_settings(),
            weak_config=weak_config,
            fail_config=fail_config,
        )
    except Exception as e:
        logger.error(f"Class roster error: {str(e)}")
        logger.error(traceback.format_exc())
        flash(f"An error occurred while loading class roster: {str(e)}", "danger")
        return redirect(url_for("admin_advanced_results.new_dashboard"))


@advanced_results_bp.route("/student-view")
def student_view():
    """Single student view with detailed results"""
    year_id = int_or_none(request.args.get("year_id"))
    exam_id = int_or_none(request.args.get("exam_id"))
    student_id = int_or_none(request.args.get("student_id"))
    
    # Get selected year, exam, and student
    selected_year = db.session.get(AcademicYear, year_id)
    selected_exam = db.session.get(Exam, exam_id)
    student = db.session.get(Student, student_id)
    
    if not selected_year or not student:
        flash("Invalid selection.", "warning")
        return redirect(url_for("admin_advanced_results.new_dashboard"))
    if not selected_exam or selected_exam.academic_year_id != selected_year.id:
        selected_exam = get_latest_exam_for_year(selected_year)
    if not selected_exam:
        flash("Please select an exam type.", "warning")
        return redirect(url_for("admin_advanced_results.result_entry", year_id=selected_year.id))

    selected_enrollment = get_enrollment_for_student_year(student.id, selected_year.id)
    selected_placement = enrollment_placement_for_student(student, selected_year.id) or {}
    selected_level_id = selected_placement.get("academic_level_id") or student.academic_level_id
    selected_class_id = selected_placement.get("academic_class_id") or student.academic_class_id
    if selected_enrollment:
        subjects = subjects_for_year_level(selected_exam, selected_enrollment.academic_year_level_id)
        selected_level_id = selected_enrollment.academic_year_level.legacy_level_id
        selected_class_id = selected_enrollment.academic_year_class.legacy_class_id
    else:
        subjects = subjects_for_scope(
            selected_exam,
            level_id=selected_level_id,
            class_id=selected_class_id,
        )
    
    # Get results for this student and exam
    results = Result.query.filter_by(student_id=student.id, exam_id=selected_exam.id).all()
    results_dict = {r.subject_id: r for r in results}
    
    # Resolve grades through the shared Grade Management cache.
    grade_cache = load_grade_scale_cache(selected_exam.id)

    def cached_grade_for(score):
        return grade_for_from_cache(score, grade_cache)
    
    # Build subject data
    subject_data = []
    total_score = 0
    total_max = 0
    grade_distribution = {}
    
    for subject in subjects:
        result = results_dict.get(subject.id)
        score = float(result.score) if result else 0
        max_score = float(subject.max_score)
        percentage = (score / max_score * 100) if max_score > 0 else 0
        
        total_score += score
        total_max += max_score
        
        # Get grade using exam-specific grade scale
        grade_info = cached_grade_for(percentage)
        
        # Apply grade_override if present
        if result and result.grade_override:
            grade_info = dict(grade_info)
            grade_info["grade"] = result.grade_override
        grade_distribution[grade_info["grade"]] = grade_distribution.get(grade_info["grade"], 0) + 1
        
        subject_data.append({
            "subject": subject,
            "result": result,
            "score": score,
            "max_score": max_score,
            "pass_mark": academic_round(max_score * 0.5, get_settings()),
            "percentage": round(percentage, 2),
            "grade": grade_info,
            "remark": grade_info.get("comment") or ("Pass" if grade_info.get("is_pass") else "Needs Improvement"),
        })
    
    overall_percentage = round((total_score / total_max * 100), 2) if total_max > 0 else 0
    overall_grade = cached_grade_for(overall_percentage)
    
    # Calculate GP
    total_points = sum(s["grade"]["grade_point"] for s in subject_data if s["grade"]["grade_point"])
    gp = academic_round(total_points / len(subject_data), get_settings()) if subject_data else 0
    status = "Passed" if overall_grade.get("is_pass") else "Failed"
    rank = rank_student_in_scope(
        student,
        selected_year.id,
        selected_exam,
        subjects,
        level_id=selected_level_id,
        class_id=selected_class_id,
        section_id=None,
    )
    
    student_view_settings = get_settings()
    student_view_precision = academic_decimal_precision(student_view_settings)
    student_view_step = {0: "1", 1: "0.1", 2: "0.01", 3: "0.001"}[student_view_precision]
    return render_template(
        "admin/student_view.html",
        years=AcademicYear.query.order_by(AcademicYear.name.desc()).all(),
        exams=Exam.query.filter_by(academic_year_id=selected_year.id).order_by(Exam.id.desc()).all(),
        selected_year=selected_year,
        selected_exam=selected_exam,
        student=student,
        selected_placement=selected_placement,
        subject_data=subject_data,
        total_score=total_score,
        total_max=total_max,
        percentage=overall_percentage,
        grade=overall_grade,
        gp=gp,
        rank=rank,
        status=status,
        grade_distribution=grade_distribution,
        settings=student_view_settings,
        decimal_step=student_view_step,
    )


@advanced_results_bp.route("/bulk", methods=["POST"])
def bulk():
    ids = [int(value) for value in request.form.getlist("result_ids") if value.isdigit()]
    action = request.form.get("action", "")
    if not ids:
        flash("Select result rows first.", "warning")
        return redirect(url_for("admin_advanced_results.dashboard"))
    rows = Result.query.filter(Result.id.in_(ids)).all()
    if action == "publish":
        for row in rows:
            row.is_published = True
        audit("Result Publishing", f"Bulk published {len(rows)} result rows")
    elif action == "unpublish":
        for row in rows:
            row.is_published = False
        audit("Result Publishing", f"Bulk unpublished {len(rows)} result rows")
    elif action == "lock":
        for row in rows:
            row.student.is_result_locked = True
            row.student.lock_reason = "Locked from advanced results."
        audit("Result Locking", f"Bulk locked {len(rows)} result rows")
    elif action == "unlock":
        for row in rows:
            row.student.is_result_locked = False
            row.student.lock_reason = ""
        audit("Result Locking", f"Bulk unlocked {len(rows)} result rows")
    elif action == "delete":
        for row in rows:
            db.session.delete(row)
        audit("Result Publishing", f"Bulk deleted {len(rows)} result rows")
    else:
        abort(400)
    db.session.commit()
    flash("Bulk result action completed.", "success")
    return redirect(url_for("admin_advanced_results.dashboard"))


@advanced_results_bp.route("/export.xlsx")
def export_excel():
    """Original advanced results Excel export - kept for backward compatibility"""
    filters = result_filters()
    wb = Workbook()
    ws = wb.active
    ws.title = "Advanced Results"
    ws.append(["Academic Year", "Exam", "Level", "Class", "Section", "Student ID", "Student Name", "Subject", "Score", "Published", "Locked"])
    for row in result_query(filters).order_by(Result.updated_at.desc()).all():
        placement = enrollment_placement_for_student(row.student, row.exam.academic_year_id) or {}
        ws.append([
            row.exam.academic_year.name,
            row.exam.name,
            placement.get("level_name") or row.student.level or "",
            placement.get("class_name") or (row.student.school_class.name if row.student.school_class else ""),
            placement.get("section_name") or row.student.section or "",
            row.student.student_code,
            row.student.full_name,
            row.subject.name,
            float(row.score),
            row.is_published,
            row.student.is_result_locked,
        ])
    audit("Result Export", "Exported advanced results")
    db.session.commit()
    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()
    return send_file(tmp.name, as_attachment=True, download_name="advanced_results.xlsx")


@advanced_results_bp.route("/export-student-pdf")
def export_student_pdf():
    """Export single student results as PDF"""
    year_id = int_or_none(request.args.get("year_id"))
    exam_id = int_or_none(request.args.get("exam_id"))
    student_id = int_or_none(request.args.get("student_id"))
    
    selected_year = db.session.get(AcademicYear, year_id)
    selected_exam = db.session.get(Exam, exam_id)
    student = db.session.get(Student, student_id)
    
    if not selected_year or not selected_exam or not student:
        abort(404)
    if selected_exam.academic_year_id != selected_year.id:
        abort(400)
    
    # Resolve the selected historical enrollment first. The permanent
    # Student identity may now have a newer placement in another year.
    selected_placement = enrollment_placement_for_student(student, selected_year.id)
    if not selected_placement:
        abort(404)
    student_level_id = selected_placement.get("academic_level_id")
    student_class_id = selected_placement.get("academic_class_id")
    subjects = (
        subjects_for_year_level(selected_exam, selected_placement.get("academic_year_level_id"))
        if selected_placement.get("academic_year_level_id")
        else (subjects_for_scope(selected_exam, level_id=student_level_id, class_id=student_class_id) if student_level_id else [])
    )
    subject_ids = [subject.id for subject in subjects]
    results = (
        Result.query.filter(
            Result.student_id == student.id,
            Result.exam_id == exam_id,
            Result.is_published.is_(True),
            Result.subject_id.in_(subject_ids),
        ).all()
        if subject_ids
        else []
    )
    results_dict = {r.subject_id: r for r in results}
    
    # Resolve grades through the shared Grade Management cache.
    grade_cache = load_grade_scale_cache(selected_exam.id)

    def cached_grade_for(score):
        return grade_for_from_cache(score, grade_cache)
    
    # Build data
    subject_data = []
    total_score = 0
    total_max = 0
    
    for subject in subjects:
        result = results_dict.get(subject.id)
        score = float(result.score) if result else 0
        max_score = float(subject.max_score)
        percentage = (score / max_score * 100) if max_score > 0 else 0
        
        total_score += score
        total_max += max_score
        
        grade_info = cached_grade_for(percentage)
        
        # Apply grade_override if present
        if result and result.grade_override:
            grade_info = dict(grade_info)
            grade_info["grade"] = result.grade_override
        
        subject_data.append({
            "subject": subject,
            "score": score,
            "max_score": max_score,
            "percentage": round(percentage, 2),
            "grade": grade_info,
        })
    
    overall_percentage = round((total_score / total_max * 100), 2) if total_max > 0 else 0
    overall_grade = cached_grade_for(overall_percentage)
    
    total_points = sum(s["grade"]["grade_point"] for s in subject_data if s["grade"]["grade_point"])
    gp = round(total_points / len(subject_data), 2) if subject_data else 0
    
    settings = get_settings()
    
    return render_template(
        "admin/pdf/student_result_pdf.html",
        selected_year=selected_year,
        selected_exam=selected_exam,
        student=student,
        selected_placement=selected_placement,
        subject_data=subject_data,
        total_score=total_score,
        total_max=total_max,
        percentage=overall_percentage,
        grade=overall_grade,
        gp=gp,
        settings=settings,
        date=date.today(),
    )


def get_report_tier_configs(year_id=None, exam_id=None, level_id=None):
    """Retrieve Weak and Fail Tier configurations for Whole-Class Report."""
    weak_key = f"weak_tier_{year_id or 0}_{exam_id or 0}_{level_id or 0}"
    fail_key = f"fail_tier_{year_id or 0}_{exam_id or 0}_{level_id or 0}"

    def _load_setting(key, default_dict):
        setting = db.session.get(Setting, key)
        if setting and setting.value:
            try:
                res = json.loads(setting.value)
                if "text_color" not in res:
                    res["text_color"] = default_dict.get("text_color", "#ffffff")
                if "bg_color" not in res:
                    res["bg_color"] = res.get("color", default_dict.get("bg_color", "#F5A400"))
                res["color"] = res["bg_color"]
                return res
            except Exception:
                pass
        return None

    default_weak = {"min": 50.0, "max": 59.99, "bg_color": "#F5A400", "color": "#F5A400", "text_color": "#ffffff"}
    default_fail = {"min": 0.0, "max": 49.99, "bg_color": "#DC2626", "color": "#DC2626", "text_color": "#ffffff"}

    weak_cfg = _load_setting(weak_key, default_weak)
    if not weak_cfg and level_id:
        weak_cfg = _load_setting(f"weak_tier_{year_id or 0}_{exam_id or 0}_0", default_weak)
    if not weak_cfg:
        weak_cfg = _load_setting("weak_tier_global", default_weak) or default_weak

    fail_cfg = _load_setting(fail_key, default_fail)
    if not fail_cfg and level_id:
        fail_cfg = _load_setting(f"fail_tier_{year_id or 0}_{exam_id or 0}_0", default_fail)
    if not fail_cfg:
        fail_cfg = _load_setting("fail_tier_global", default_fail) or default_fail

    return {
        "weak": weak_cfg,
        "fail": fail_cfg,
    }


def save_report_tier_configs(year_id, exam_id, level_id, weak_min, weak_max, weak_bg, weak_text, fail_min, fail_max, fail_bg, fail_text):
    """Save Weak and Fail tier configurations for a scope."""
    weak_key = f"weak_tier_{year_id or 0}_{exam_id or 0}_{level_id or 0}"
    fail_key = f"fail_tier_{year_id or 0}_{exam_id or 0}_{level_id or 0}"

    s_weak = db.session.get(Setting, weak_key) or Setting(key=weak_key)
    s_weak.value = json.dumps({
        "min": float(weak_min),
        "max": float(weak_max),
        "bg_color": str(weak_bg).strip(),
        "color": str(weak_bg).strip(),
        "text_color": str(weak_text).strip(),
    })
    db.session.add(s_weak)

    s_fail = db.session.get(Setting, fail_key) or Setting(key=fail_key)
    s_fail.value = json.dumps({
        "min": float(fail_min),
        "max": float(fail_max),
        "bg_color": str(fail_bg).strip(),
        "color": str(fail_bg).strip(),
        "text_color": str(fail_text).strip(),
    })
    db.session.add(s_fail)

    db.session.commit()


def get_weak_tier_config(year_id=None, exam_id=None, level_id=None):
    return get_report_tier_configs(year_id, exam_id, level_id)["weak"]


def save_weak_tier_config(year_id, exam_id, level_id, min_score, max_score, color):
    cfgs = get_report_tier_configs(year_id, exam_id, level_id)
    fail_c = cfgs["fail"]
    save_report_tier_configs(year_id, exam_id, level_id, min_score, max_score, color, "#ffffff", fail_c["min"], fail_c["max"], fail_c["bg_color"], fail_c["text_color"])


@advanced_results_bp.route("/export-class-pdf")
def export_class_pdf():
    """Export class results as PDF mark sheet (Whole-Class Result Report)"""
    year_id = int_or_none(request.args.get("year_id"))
    exam_id = int_or_none(request.args.get("exam_id"))
    level_id = int_or_none(request.args.get("level_id"))
    class_id = int_or_none(request.args.get("class_id"))
    section_id = int_or_none(request.args.get("section_id"))
    
    selected_year = db.session.get(AcademicYear, year_id)
    selected_exam = db.session.get(Exam, exam_id)
    
    if not selected_year or not selected_exam:
        abort(404)
    if not (level_id and class_id):
        abort(400)
    
    students = (
        students_for_scope_query(
            year_id,
            level_id=level_id,
            class_id=class_id,
            section_id=section_id,
        )
        .order_by(Student.full_name)
        .all()
    )
    subjects = subjects_for_scope(selected_exam, level_id=level_id, class_id=class_id)
    attendance_uf_keys = attendance_uf_subject_keys(
        selected_exam,
        [student.id for student in students],
        [subject.id for subject in subjects],
    )
    
    # Resolve grades through the shared Grade Management cache.
    grade_cache = load_grade_scale_cache(selected_exam.id)

    def cached_grade_for(score):
        return grade_for_from_cache(score, grade_cache)
    
    report_tiers = get_report_tier_configs(year_id=year_id, exam_id=exam_id, level_id=level_id)
    weak_config = report_tiers["weak"]
    fail_config = report_tiers["fail"]

    # Build roster data
    roster_data = []
    from .routes_public import feedback_access_token
    for student in students:
        results = Result.query.filter_by(student_id=student.id, exam_id=exam_id, is_published=True).all()
        results_dict = {r.subject_id: r for r in results}
        
        total_score = 0
        total_max = 0
        subject_data = []
        
        for subject in subjects:
            result = results_dict.get(subject.id)
            score = float(result.score) if result else 0
            max_score = float(subject.max_score)
            percentage = (score / max_score * 100) if max_score > 0 else 0
            
            total_score += score
            total_max += max_score
            
            grade_info = cached_grade_for(percentage)
            
            # Apply grade_override if present
            if result and result.grade_override:
                grade_info = dict(grade_info)
                grade_info["grade"] = result.grade_override

            tier = performance_tier_for(percentage, weak_config, fail_config)
            is_fail = tier["is_fail"]
            is_weak = tier["is_weak"]

            subject_data.append({
                "subject_id": subject.id,
                "score": score,
                "percentage": round(percentage, 2),
                "grade": grade_info,
                "is_fail": is_fail,
                "is_weak": is_weak,
                "is_uf": (student.id, subject.id) in attendance_uf_keys,
            })
        
        overall_percentage = round((total_score / total_max * 100), 2) if total_max > 0 else 0
        overall_grade = cached_grade_for(overall_percentage)
        overall_tier = performance_tier_for(overall_percentage, weak_config, fail_config)
        overall_fail = overall_tier["is_fail"]
        overall_weak = overall_tier["is_weak"]

        roster_data.append({
            "student": student,
            "mg_token": feedback_access_token(student, selected_exam),
            "subject_data": subject_data,
            "total_score": total_score,
            "total_max": total_max,
            "percentage": overall_percentage,
            "grade": overall_grade,
            "is_fail": overall_fail,
            "is_weak": overall_weak,
        })

    ranked_data = sorted(roster_data, key=lambda row: (row["percentage"], row["total_score"]), reverse=True)
    rank_lookup = competition_rank_lookup({row["student"].id: row["percentage"] for row in ranked_data})
    for row in ranked_data:
        row["rank"] = rank_lookup.get(row["student"].id, 0)
        row["rank_label"] = ordinal(row["rank"]) if row["rank"] else "-"
    
    roster_data = ranked_data
    
    # Calculate class stats
    class_average = round(sum(r["percentage"] for r in roster_data) / len(roster_data), 2) if roster_data else 0
    highest_total = round(max((row["total_score"] for row in roster_data), default=0), 2)
    lowest_total = round(min((row["total_score"] for row in roster_data), default=0), 2)
    average_total = round(sum(row["total_score"] for row in roster_data) / len(roster_data), 2) if roster_data else 0
    passed_count = sum(1 for row in roster_data if not row["is_fail"])
    failed_count = sum(1 for row in roster_data if row["is_fail"])
    pass_rate = round((passed_count / len(roster_data) * 100), 2) if roster_data else 0
    highest_score = round(max((row["percentage"] for row in roster_data), default=0), 2)
    lowest_score = round(min((row["percentage"] for row in roster_data), default=0), 2)
    subject_stats = []
    for index, subject in enumerate(subjects):
        scores = [float(row["subject_data"][index]["score"]) for row in roster_data if index < len(row["subject_data"])]
        subject_stats.append({
            "subject": subject,
            "highest": round(max(scores), 2) if scores else 0,
            "lowest": round(min(scores), 2) if scores else 0,
            "average": round(sum(scores) / len(scores), 2) if scores else 0,
        })
    students_per_page = 15
    total_pages = max(1, math.ceil(len(roster_data) / students_per_page))
    
    # Get scope info
    scope_info = {}
    if level_id:
        scope_info["level"] = db.session.get(AcademicLevel, level_id)
    if class_id:
        scope_info["class"] = db.session.get(AcademicClass, class_id)
    if section_id:
        scope_info["section"] = db.session.get(AcademicSection, section_id)
    
    settings = get_settings()
    
    return render_template(
        "admin/pdf/class_mark_sheet_pdf.html",
        selected_year=selected_year,
        selected_exam=selected_exam,
        scope_info=scope_info,
        students=ranked_data,
        subjects=subjects,
        subject_stats=subject_stats,
        class_average=class_average,
        highest_total=highest_total,
        lowest_total=lowest_total,
        average_total=average_total,
        passed_count=passed_count,
        failed_count=failed_count,
        pass_rate=pass_rate,
        highest_score=highest_score,
        lowest_score=lowest_score,
        completed_count=len(roster_data),
        students_per_page=students_per_page,
        total_pages=total_pages,
        settings=settings,
        weak_config=weak_config,
        fail_config=fail_config,
        report_tiers=report_tiers,
        generated_by=current_user.full_name or current_user.username,
        date=date.today(),
    )


@advanced_results_bp.route("/export-class-excel")
def export_class_excel():
    """Export class results as Excel with conditional formatting"""
    year_id = int_or_none(request.args.get("year_id"))
    exam_id = int_or_none(request.args.get("exam_id"))
    level_id = int_or_none(request.args.get("level_id"))
    class_id = int_or_none(request.args.get("class_id"))
    section_id = int_or_none(request.args.get("section_id"))
    
    selected_year = db.session.get(AcademicYear, year_id)
    selected_exam = db.session.get(Exam, exam_id)
    selected_class = db.session.get(AcademicClass, class_id)
    
    if not selected_year or not selected_exam:
        abort(404)
    if not (level_id and class_id):
        abort(400)
    
    students = (
        students_for_scope_query(
            year_id,
            level_id=level_id,
            class_id=class_id,
            section_id=section_id,
        )
        .order_by(Student.full_name)
        .all()
    )
    subjects = subjects_for_scope(selected_exam, level_id=level_id, class_id=class_id)
    attendance_uf_keys = attendance_uf_subject_keys(
        selected_exam,
        [student.id for student in students],
        [subject.id for subject in subjects],
    )
    
    # Resolve grades through the shared Grade Management cache.
    grade_cache = load_grade_scale_cache(selected_exam.id)
    report_tiers = get_report_tier_configs(
        year_id=year_id,
        exam_id=exam_id,
        level_id=level_id,
    )
    weak_config = report_tiers["weak"]
    fail_config = report_tiers["fail"]

    def cached_grade_for(score):
        return grade_for_from_cache(score, grade_cache)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Class Results"
    
    # Header row
    headers = ["ID", "Student Name", "Mother's Name", "Class"]
    for subject in subjects:
        headers.append(subject.name)
    headers.extend(["Total", "%", "Grade", "GP"])
    ws.append(headers)
    
    # Style header
    header_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Data rows
    row_num = 2
    for student in students:
        results = Result.query.filter_by(student_id=student.id, exam_id=exam_id, is_published=True).all()
        results_dict = {r.subject_id: r for r in results}
        
        total_score = 0
        total_max = 0
        row_data = [
            student.student_code,
            student.full_name,
            student.mother_name or "",
            f"{student.academic_class.name if student.academic_class else student.level or '-'}{' - ' + student.academic_section.name if student.academic_section else ''}"
        ]
        
        for subject in subjects:
            result = results_dict.get(subject.id)
            score = float(result.score) if result else 0
            max_score = float(subject.max_score)
            percentage = round((score / max_score * 100), 2) if max_score > 0 else 0
            
            total_score += score
            total_max += max_score
            
            row_data.append("⚠️ MG" if (student.id, subject.id) in attendance_uf_keys else score)
        
        overall_percentage = round((total_score / total_max * 100), 2) if total_max > 0 else 0
        overall_grade = cached_grade_for(overall_percentage)
        overall_tier = performance_tier_for(overall_percentage, weak_config, fail_config)
        
        total_points = 0
        for subject in subjects:
            result = results_dict.get(subject.id)
            if result:
                percentage = round((float(result.score) / float(subject.max_score) * 100), 2) if subject.max_score else 0
                grade_info = cached_grade_for(percentage)
                # Apply grade_override if present
                if result.grade_override:
                    grade_info = dict(grade_info)
                    grade_info["grade"] = result.grade_override
                total_points += grade_info["grade_point"]
        
        gp = academic_round(total_points / len(subjects), get_settings()) if subjects else 0
        
        row_data.extend([total_score, overall_percentage, overall_grade["grade"], gp])
        ws.append(row_data)
        
        # Apply conditional formatting to percentage column
        percentage_col = len(subjects) + 5  # ID, Name, Mother, Class + subjects + Total
        percentage_cell = ws.cell(row=row_num, column=percentage_col)
        
        if overall_tier["is_fail"]:
            fill_color = fail_config.get("bg_color", "#fee2e2")
        elif overall_tier["is_weak"]:
            fill_color = weak_config.get("bg_color", "#fef3c7")
        else:
            fill_color = overall_grade.get("background_color", "#ffffff")
        fill_color = str(fill_color).lstrip("#")
        percentage_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        row_num += 1
    
    audit("Result Export", f"Exported class results for exam {exam_id}")
    db.session.commit()
    
    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()
    return send_file(
        tmp.name,
        as_attachment=True,
        download_name=result_workbook_filename(
            selected_year,
            selected_exam,
            selected_class,
            results=True,
        ),
    )


@advanced_results_bp.route("/result-entry")
def result_entry():
    """Whole-Class Result Entry grid for bulk score entry"""
    import logging
    logger = logging.getLogger(__name__)
    
    year_id = int_or_none(request.args.get("year_id"))
    exam_id = int_or_none(request.args.get("exam_id"))
    level_id = int_or_none(request.args.get("level_id"))
    class_id = int_or_none(request.args.get("class_id"))
    section_id = int_or_none(request.args.get("section_id"))
    
    selected_year = get_default_academic_year(year_id)
    selected_exam = db.session.get(Exam, exam_id) if exam_id else get_latest_exam_for_year(selected_year)
    if selected_exam and selected_year and selected_exam.academic_year_id != selected_year.id:
        selected_exam = get_latest_exam_for_year(selected_year)
    
    # Debug logging - Result Entry exam selection
    if selected_exam:
        logger.info(f"RESULT ENTRY - Exam selected: ID={selected_exam.id}, Name={selected_exam.name}")
        logger.info(f"RESULT ENTRY - Exam academic_year_id: {selected_exam.academic_year_id}")
        logger.info(f"RESULT ENTRY - Exam academic_level_id: {selected_exam.academic_level_id}")
        logger.info(f"RESULT ENTRY - Exam academic_class_id: {selected_exam.academic_class_id}")
        logger.info(f"RESULT ENTRY - Exam academic_section_id: {selected_exam.academic_section_id}")
        logger.info(f"RESULT ENTRY - Exam is_active: {selected_exam.is_active}")
        logger.info(f"RESULT ENTRY - Exam is_published: {selected_exam.is_published}")
    else:
        logger.warning(f"RESULT ENTRY - No exam selected for year_id={year_id}, exam_id={exam_id}")
    
    if not selected_year:
        flash("Please select an academic year.", "warning")
        return redirect(url_for("admin_advanced_results.new_dashboard"))

    years = AcademicYear.query.order_by(AcademicYear.name.desc()).all()
    exams = Exam.query.filter_by(academic_year_id=selected_year.id).order_by(Exam.id.desc()).all() if selected_year else []
    # The Results Hub still submits legacy level/class IDs for compatibility,
    # but the option lists must be built from the selected year's mappings.
    # This prevents a level or class created in another academic year from
    # appearing in this form.
    year_level_scopes = year_levels(selected_year.id) if selected_year else []
    levels = [
        scope.legacy_level
        for scope in year_level_scopes
        if scope.legacy_level and scope.legacy_level.is_active
    ]
    
    # If no exam selected, show exam selection interface
    if not selected_exam:
        return render_template(
            "admin/result_entry.html",
            selected_year=selected_year,
            selected_exam=None,
            scope_info={},
            subjects=[],
            entry_grid=[],
            years=years,
            exams=exams,
            levels=levels,
            classes=[],
            sections=[],
            settings=get_settings(),
        )

    requested_level_id = level_id
    requested_class_id = class_id
    requested_section_id = section_id
    valid_level_ids = {scope.legacy_level_id for scope in year_level_scopes if scope.legacy_level_id}
    level_id = requested_level_id if requested_level_id in valid_level_ids else None
    if requested_level_id is None and selected_exam.academic_level_id in valid_level_ids:
        level_id = selected_exam.academic_level_id

    scope_for_level = next(
        (scope for scope in year_level_scopes if scope.legacy_level_id == level_id),
        None,
    )
    mapped_classes = [
        year_class.legacy_class
        for scope in ([scope_for_level] if scope_for_level else year_level_scopes)
        for year_class in year_classes(scope.id)
        if year_class.legacy_class and year_class.legacy_class.is_active
    ]
    valid_class_ids = {item.id for item in mapped_classes}
    class_id = requested_class_id if requested_class_id in valid_class_ids else None
    if requested_class_id is None and selected_exam.academic_class_id in valid_class_ids:
        class_id = selected_exam.academic_class_id
    classes = [item for item in mapped_classes if not level_id or item.academic_level_id == level_id]
    section_id = requested_section_id
    if section_id and not class_id:
        section_id = None
    sections = AcademicSection.query.filter_by(academic_class_id=class_id, is_active=True).order_by(AcademicSection.sort_order, AcademicSection.name).all() if class_id else []
    if section_id and not any(item.id == section_id for item in sections):
        section_id = None
    
    # Build scope info
    scope_info = {
        "level": db.session.get(AcademicLevel, level_id) if level_id else None,
        "class": db.session.get(AcademicClass, class_id) if class_id else None,
        "section": db.session.get(AcademicSection, section_id) if section_id else None,
    }
    
    subjects = subjects_for_scope(selected_exam, level_id=level_id, class_id=class_id)
    students = []
    if level_id and class_id:
        students = (
            students_for_scope_query(
                selected_year.id,
                level_id=level_id,
                class_id=class_id,
                section_id=section_id,
            )
            .order_by(Student.full_name)
            .all()
        )
    
    # Get existing results for these students and this exam
    student_ids = [s.id for s in students]
    existing_results = Result.query.filter(Result.student_id.in_(student_ids), Result.exam_id == selected_exam.id).all()
    results_dict = {(r.student_id, r.subject_id): r for r in existing_results}
    
    # Build entry grid data
    entry_grid = []
    for student in students:
        row_data = {
            "student": student,
            "results": {}
        }
        for subject in subjects:
            result = results_dict.get((student.id, subject.id))
            row_data["results"][subject.id] = {
                "result": result,
                "score": float(result.score) if result else None,
                "grade_override": result.grade_override if result else None,
                "is_published": result.is_published if result else True,
            }
        entry_grid.append(row_data)
    
    return render_template(
        "admin/result_entry.html",
        selected_year=selected_year,
        selected_exam=selected_exam,
        scope_info=scope_info,
        subjects=subjects,
        entry_grid=entry_grid,
        years=years,
        exams=exams,
        levels=levels,
        classes=classes,
        sections=sections,
        settings=get_settings(),
    )


@advanced_results_bp.route("/result-entry/autosave", methods=["POST"])
def autosave_result_entry():
    """Autosave a single score from the Results Hub entry grid."""
    year_id = int_or_none(request.form.get("year_id"))
    exam_id = int_or_none(request.form.get("exam_id"))
    level_id = int_or_none(request.form.get("level_id"))
    class_id = int_or_none(request.form.get("class_id"))
    section_id = int_or_none(request.form.get("section_id"))
    student_id = int_or_none(request.form.get("student_id"))
    subject_id = int_or_none(request.form.get("subject_id"))
    raw_score = request.form.get("score", "").strip()

    selected_year = db.session.get(AcademicYear, year_id)
    selected_exam = db.session.get(Exam, exam_id)
    student = db.session.get(Student, student_id)
    subject = db.session.get(Subject, subject_id)

    if not selected_year or not selected_exam or not student or not subject:
        return jsonify({"ok": False, "message": "Invalid result context."}), 400

    in_scope = students_for_scope_query(
        selected_year.id,
        level_id=level_id,
        class_id=class_id,
        section_id=section_id,
    ).filter(Student.id == student.id).first()
    if not in_scope:
        return jsonify({"ok": False, "message": "Student is outside the selected scope."}), 400

    if not raw_score:
        return jsonify({"ok": True, "status": "empty", "message": "No score entered."})

    try:
        score = float(raw_score)
    except ValueError:
        return jsonify({"ok": False, "message": "Invalid score."}), 400

    max_score = float(subject.max_score)
    if score < 0 or score > max_score:
        return jsonify({"ok": False, "message": f"Score must be between 0 and {max_score:g}."}), 400

    result = Result.query.filter_by(student_id=student.id, exam_id=selected_exam.id, subject_id=subject.id).first()
    if not result:
        result = Result(student=student, exam=selected_exam, subject=subject)
        db.session.add(result)

    result.score = score
    result.is_published = True
    audit("Result Entry", f"Autosaved {student.student_code} - {subject.name}: {score:g}")
    db.session.commit()

    return jsonify({
        "ok": True,
        "status": "saved",
        "score": score,
        "student_id": student.id,
        "subject_id": subject.id,
    })


@advanced_results_bp.route("/result-entry/save", methods=["POST"])
def save_result_entry():
    """Save bulk result entry from grid"""
    year_id = int_or_none(request.form.get("year_id"))
    exam_id = int_or_none(request.form.get("exam_id"))
    level_id = int_or_none(request.form.get("level_id"))
    class_id = int_or_none(request.form.get("class_id"))
    section_id = int_or_none(request.form.get("section_id"))
    
    selected_year = db.session.get(AcademicYear, year_id)
    selected_exam = db.session.get(Exam, exam_id)
    
    if not selected_year or not selected_exam:
        flash("Invalid selection.", "danger")
        return redirect(url_for("admin_advanced_results.new_dashboard"))
    
    subjects = subjects_for_scope(selected_exam, level_id=level_id, class_id=class_id)
    students = students_for_scope_query(
        selected_year.id,
        level_id=level_id,
        class_id=class_id,
        section_id=section_id,
    ).all()
    student_ids = [s.id for s in students]
    
    # Get existing results
    existing_results = Result.query.filter(Result.student_id.in_(student_ids), Result.exam_id == selected_exam.id).all()
    results_dict = {(r.student_id, r.subject_id): r for r in existing_results}
    
    # Process form data
    saved_count = 0
    validation_errors = []
    
    for student in students:
        for subject in subjects:
            score_key = f"score_{student.id}_{subject.id}"
            override_key = f"override_{student.id}_{subject.id}"
            published_key = f"published_{student.id}_{subject.id}"
            
            raw_score = request.form.get(score_key, "").strip()
            grade_override = request.form.get(override_key, "").strip()
            is_published = request.form.get(published_key) == "on"
            
            # Skip if no score entered
            if not raw_score:
                continue
            
            # Validate score against max_score
            try:
                score = float(raw_score)
                if score < 0 or score > float(subject.max_score):
                    validation_errors.append(f"{student.student_code} - {subject.name}: Score {score} exceeds max {subject.max_score}")
                    continue
            except ValueError:
                validation_errors.append(f"{student.student_code} - {subject.name}: Invalid score '{raw_score}'")
                continue
            
            # Get or create result
            result = results_dict.get((student.id, subject.id))
            if not result:
                result = Result(student=student, exam=selected_exam, subject=subject)
                db.session.add(result)
            
            result.score = score
            result.grade_override = grade_override if grade_override else None
            result.is_published = is_published
            saved_count += 1
    
    if validation_errors:
        flash(f"Saved {saved_count} results with {len(validation_errors)} validation errors.", "warning")
        for error in validation_errors[:5]:  # Show first 5 errors
            flash(error, "warning")
    else:
        flash(f"Successfully saved {saved_count} results.", "success")
    
    audit("Result Entry", f"Bulk saved {saved_count} results for exam {selected_exam.name}")
    db.session.commit()
    
    # Redirect back to entry grid with same scope
    return redirect(url_for("admin_advanced_results.result_entry", year_id=year_id, exam_id=exam_id, level_id=level_id, class_id=class_id, section_id=section_id))


@advanced_results_bp.route("/analytics")
def analytics():
    """Results analytics with charts matching design system"""
    year_id = int_or_none(request.args.get("year_id"))
    exam_id = int_or_none(request.args.get("exam_id"))
    level_id = int_or_none(request.args.get("level_id"))
    class_id = int_or_none(request.args.get("class_id"))
    section_id = int_or_none(request.args.get("section_id"))
    subject_id = int_or_none(request.args.get("subject_id"))
    top_limit = int_or_none(request.args.get("top_limit")) or 5
    bottom_limit = int_or_none(request.args.get("bottom_limit")) or 5
    
    # Initialize selected variables before any conditional logic
    selected_year = None
    selected_exam = None
    selected_level = None
    selected_class = None
    selected_section = None
    selected_subject = None
    
    # Get selected year
    selected_year = db.session.get(AcademicYear, year_id) if year_id else AcademicYear.query.filter_by(is_current=True).first()
    
    if not selected_year:
        flash("Please select an academic year.", "warning")
        return redirect(url_for("admin_advanced_results.new_dashboard"))
    
    # Get selected exam from parameter or auto-select most recent. A stale
    # exam ID from another year is never accepted into this scope.
    selected_exam = db.session.get(Exam, exam_id) if exam_id else None
    if selected_exam and selected_exam.academic_year_id != selected_year.id:
        selected_exam = None
    if not selected_exam:
        selected_exam = Exam.query.filter_by(academic_year_id=selected_year.id).order_by(Exam.id.desc()).first()
    
    # Get filter data for selectors
    years = AcademicYear.query.order_by(AcademicYear.name.desc()).all()
    exams = Exam.query.filter_by(academic_year_id=selected_year.id).order_by(Exam.id.desc()).all() if selected_year else []
    levels = year_levels(selected_year.id)
    selected_year_level = db.session.get(AcademicYearLevel, level_id) if level_id else None
    if selected_year_level and selected_year_level.academic_year_id != selected_year.id:
        selected_year_level = None
        level_id = None
    selected_year_class = db.session.get(AcademicYearClass, class_id) if class_id else None
    if selected_year_class and (
        not selected_year_class.academic_year_level
        or selected_year_class.academic_year_level.academic_year_id != selected_year.id
        or (selected_year_level and selected_year_class.academic_year_level_id != selected_year_level.id)
    ):
        selected_year_class = None
        class_id = None
    classes = year_classes(selected_year_level.id) if selected_year_level else [item for level in levels for item in year_classes(level.id)]
    sections = (
        AcademicSection.query.filter_by(academic_class_id=selected_year_class.legacy_class_id, is_active=True).order_by(AcademicSection.name).all()
        if selected_year_class and selected_year_class.legacy_class_id
        else []
    )
    year_scoped_subjects, legacy_scoped_subjects = analytics_subject_bridge(selected_year.id, level_id)
    selected_year_subject = db.session.get(AcademicYearSubject, subject_id) if subject_id else None
    if selected_year_subject and (
        selected_year_subject.academic_year_id != selected_year.id
        or (level_id and selected_year_subject.academic_year_level_id != level_id)
    ):
        selected_year_subject = None
        subject_id = None
    subjects = year_scoped_subjects
    
    # If still no exam available, show empty analytics
    if not selected_exam:
        # Provide empty analytics structure to prevent template errors
        empty_analytics = {
            "grade_distribution": {"labels": [], "counts": [], "colors": [], "total": 0},
            "subject_performance": {"labels": [], "scores": []},
            "exam_trend": {"labels": [], "scores": []},
            "pass_fail_ratio": {"pass": 0, "fail": 0, "total": 0},
            "completion_rate": {"percentage": 0.0, "actual": 0, "expected": 0},
            "student_pass_fail": {"pass_count": 0, "fail_count": 0, "total_students": 0, "pass_pct": 0.0, "fail_pct": 0.0},
            "overall_average": 0,
            "top_performers": [],
            "bottom_performers": [],
            "total_students": 0,
            "highest_score": 0,
            "lowest_score": 0,
        }
        return render_template(
            "admin/analytics.html",
            selected_year=selected_year,
            selected_exam=None,
            scope_info={},
            analytics=empty_analytics,
            years=years,
            exams=exams,
            levels=levels,
            classes=classes,
            sections=sections,
            subjects=subjects,
            selected_level_id=level_id,
            selected_class_id=class_id,
            selected_section_id=section_id,
            selected_subject_id=subject_id,
            top_limit=top_limit,
            bottom_limit=bottom_limit,
            settings=get_settings(),
        )
    
    # Build scope info
    selected_section = db.session.get(AcademicSection, section_id) if section_id else None
    if selected_section and (
        not selected_year_class
        or selected_section.academic_class_id != selected_year_class.legacy_class_id
    ):
        selected_section = None
        section_id = None

    scope_info = {
        "level": selected_year_level,
        "class": selected_year_class,
        "section": selected_section,
        "subject": selected_year_subject,
    }
    
    students = (
        students_for_year_scope_query(
            selected_year.id,
            year_level_id=level_id,
            year_class_id=class_id,
            section_id=selected_section.id if selected_section else None,
        )
        .options(
            selectinload(Student.academic_class),
            selectinload(Student.academic_section),
        )
        .order_by(Student.full_name)
        .all()
    )
    student_ids = [s.id for s in students]
    
    # Get results for this exam
    results_query = (
        Result.query.options(selectinload(Result.subject))
        .filter(
            Result.student_id.in_(student_ids),
            Result.exam_id == selected_exam.id,
            Result.is_published.is_(True),
        )
    )
    if scope_info["subject"]:
        results_query = results_query.filter_by(subject_id=scope_info["subject"].legacy_subject_id)
    results = results_query.all()
    
    # Scoped subjects for completion rate calculation
    scoped_subjects = (
        [db.session.get(Subject, scope_info["subject"].legacy_subject_id)]
        if scope_info["subject"] and scope_info["subject"].legacy_subject_id
        else ([] if scope_info["subject"] else legacy_scoped_subjects)
    )

    # Calculate analytics data with ranking limits
    analytics_data = build_analytics_data(results, students, selected_exam, top_limit, bottom_limit, scoped_subjects=scoped_subjects)
    
    return render_template(
        "admin/analytics.html",
        selected_year=selected_year,
        selected_exam=selected_exam,
        scope_info=scope_info,
        analytics=analytics_data,
        years=years,
        exams=exams,
        levels=levels,
        classes=classes,
        sections=sections,
        subjects=subjects,
        selected_level_id=level_id,
        selected_class_id=class_id,
        selected_section_id=section_id,
        selected_subject_id=subject_id,
        top_limit=top_limit,
        bottom_limit=bottom_limit,
        settings=get_settings(),
    )


def _analytics_report_student_level_id(student, classes_by_id, levels_by_name, legacy_class_matches):
    """Resolve a student's Results Hub level, retaining the legacy enrollment fallback."""
    if student.academic_level_id:
        return student.academic_level_id

    academic_class = classes_by_id.get(student.academic_class_id)
    if academic_class:
        return academic_class.academic_level_id

    if student.level:
        return levels_by_name.get(student.level.strip().lower())

    legacy_name = student.school_class.name if student.school_class else None
    matches = legacy_class_matches.get((legacy_name or "").strip().lower(), [])
    return matches[0].academic_level_id if len(matches) == 1 else None


def _analytics_report_student_class_id(student, classes_by_id, legacy_class_matches, level_id):
    """Resolve a student's Results Hub class, retaining the legacy enrollment fallback."""
    if student.academic_class_id:
        return student.academic_class_id

    legacy_name = student.school_class.name if student.school_class else None
    matches = legacy_class_matches.get((legacy_name or "").strip().lower(), [])
    for academic_class in matches:
        if academic_class.academic_level_id == level_id:
            return academic_class.id
    return matches[0].id if len(matches) == 1 else None


def _attendance_rows_for_exam_scope(academic_year, exam, student_ids):
    """Return the final attendance records for one Results Hub examination.

    Canonical attendance records point at ``Exam``.  The legacy ExamType arm
    keeps historical records readable during the ongoing Results Hub migration.
    It is deliberately restricted to the matching year/name, never all records
    in an academic year.
    """
    if not student_ids:
        return []
    scope_filters = [AttendanceRecord.exam_id == exam.id]
    legacy_exam_type = ExamType.query.filter_by(
        academic_year_id=academic_year.id,
        name=exam.name,
    ).first()
    if legacy_exam_type:
        scope_filters.append(
            db_and(
                AttendanceRecord.exam_id.is_(None),
                AttendanceRecord.exam_type_id == legacy_exam_type.id,
            )
        )
    return (
        AttendanceRecord.query.filter(
            AttendanceRecord.academic_year_id == academic_year.id,
            AttendanceRecord.student_id.in_(student_ids),
            db_or(*scope_filters),
        )
        .order_by(AttendanceRecord.recorded_at.desc(), AttendanceRecord.id.desc())
        .all()
    )


def build_analytics_results_report_data(academic_year, exam):
    """Build the reference report's LEVELS payload using published Results Hub data only."""
    year_level_scopes = year_levels(academic_year.id)
    legacy_report_mode = not year_level_scopes
    if legacy_report_mode:
        # A year with no Phase 1D bridge is an explicit legacy-only
        # compatibility case. Keep it readable without allowing this fallback
        # once any year-aware level exists.
        legacy_students = Student.query.filter_by(academic_year_id=academic_year.id).all()
        legacy_level_ids = {
            student.academic_level_id
            or (student.academic_class.academic_level_id if student.academic_class else None)
            for student in legacy_students
        }
        if exam.academic_level_id:
            legacy_level_ids.add(exam.academic_level_id)
        legacy_level_ids.discard(None)
        levels = (
            AcademicLevel.query
            .filter(AcademicLevel.id.in_(legacy_level_ids))
            .filter(AcademicLevel.is_active.is_(True))
            .order_by(AcademicLevel.sort_order, AcademicLevel.name, AcademicLevel.id)
            .all()
            if legacy_level_ids else []
        )
        classes = (
            AcademicClass.query
            .filter(AcademicClass.academic_level_id.in_(legacy_level_ids))
            .filter(AcademicClass.is_active.is_(True))
            .order_by(AcademicClass.sort_order, AcademicClass.name, AcademicClass.id)
            .all()
            if legacy_level_ids else []
        )
    else:
        levels = [scope.legacy_level for scope in year_level_scopes if scope.legacy_level and scope.legacy_level.is_active]
        classes = [
            year_class.legacy_class
            for scope in year_level_scopes
            for year_class in year_classes(scope.id)
            if year_class.legacy_class and year_class.legacy_class.is_active
        ]
    classes_by_id = {academic_class.id: academic_class for academic_class in classes}
    levels_by_name = {level.name.strip().lower(): level.id for level in levels if level.name}
    classes_by_level = defaultdict(list)
    legacy_class_matches = defaultdict(list)
    for academic_class in classes:
        classes_by_level[academic_class.academic_level_id].append(academic_class)
        legacy_class_matches[academic_class.name.strip().lower()].append(academic_class)

    # Results reports must start from the selected year's enrollment scope.
    # The mutable legacy placement on Student is only a compatibility source
    # for students that have no enrollment for this year.
    students = (
        students_for_year_scope_query(academic_year.id)
        .options(
            selectinload(Student.academic_level),
            selectinload(Student.academic_class),
            selectinload(Student.school_class),
        )
        .order_by(Student.full_name)
        .all()
    )
    students_by_level_class = defaultdict(list)
    scoped_students = []
    for student in students:
        placement = resolve_student_academic_context(student, academic_year.id)
        level_id = placement.get("academic_level_id") if placement else None
        class_id = placement.get("academic_class_id") if placement else None
        if (
            level_id in classes_by_level
            and class_id in classes_by_id
            and classes_by_id[class_id].academic_level_id == level_id
        ):
            students_by_level_class[(level_id, class_id)].append(student)
            scoped_students.append(student)

    included_levels = [
        level for level in levels
        if any(students_by_level_class.get((level.id, academic_class.id)) for academic_class in classes_by_level[level.id])
    ]
    student_ids = [student.id for student in scoped_students]
    if not student_ids:
        return []

    results = (
        Result.query.options(selectinload(Result.subject))
        .filter(
            Result.exam_id == exam.id,
            Result.is_published.is_(True),
            Result.student_id.in_(student_ids),
        )
        .all()
    )
    # Attendance is the source of truth for an examination sitting.  A student
    # counts only after Joogto/present or Daahid/late was recorded.  Older
    # records can contain localized labels, which ``counts_as_exam_sitting``
    # normalizes without rewriting historical data.
    latest_attendance = {}
    for record in _attendance_rows_for_exam_scope(academic_year, exam, student_ids):
        if not record.subject_id:
            continue
        latest_attendance.setdefault((record.student_id, record.subject_id), record)

    subject_sitting_student_ids = defaultdict(set)
    for (student_id, subject_id), record in latest_attendance.items():
        if counts_as_exam_sitting(record.status):
            subject_sitting_student_ids[subject_id].add(student_id)
    exam_sitting_student_ids = set().union(*subject_sitting_student_ids.values()) if subject_sitting_student_ids else set()

    results_by_student = defaultdict(list)
    results_by_level_subject = defaultdict(lambda: defaultdict(list))
    grade_cache = load_grade_scale_cache(exam.id)

    def is_pass(score):
        return bool(grade_for_from_cache(score, grade_cache).get("is_pass"))

    def gender_bucket(student):
        gender = str(student.gender or "").strip().lower()
        return "f" if gender == "female" else "m"

    student_level_lookup = {}
    for (level_id, _class_id), grouped_students in students_by_level_class.items():
        for student in grouped_students:
            student_level_lookup[student.id] = level_id
    subject_ids_by_level = defaultdict(set)
    if legacy_report_mode:
        for level in levels:
            subject_ids_by_level[level.id].update(
                subject.id
                for subject in Subject.query.filter_by(academic_level_id=level.id, is_active=True).all()
            )
    else:
        for scope in year_level_scopes:
            if not scope.legacy_level_id:
                continue
            subject_ids_by_level[scope.legacy_level_id].update(
                item.legacy_subject_id
                for item in year_subjects(academic_year.id, scope.id)
                if item.legacy_subject_id
            )
    for result in results:
        level_id = student_level_lookup.get(result.student_id)
        if (
            not level_id
            or not result.subject
            or not result.subject.max_score
            or result.subject_id not in subject_ids_by_level.get(level_id, set())
            or result.student_id not in subject_sitting_student_ids.get(result.subject_id, set())
        ):
            continue
        result_pct = round(float(result.score or 0) / float(result.subject.max_score) * 100, 4)
        results_by_student[result.student_id].append(result_pct)
        results_by_level_subject[level_id][result.subject_id].append((result.subject, result_pct))

    student_averages = {
        student_id: round(sum(scores) / len(scores), 4)
        for student_id, scores in results_by_student.items()
        if scores
    }

    level_styles = ("sec", "up", "low")
    level_titles = {
        "secondary": "Dugsiga Sare",
        "upper primary": "Dugsiga Dhexe",
        "lower primary": "Dugsiga Hoose",
    }
    levels_data = []
    for position, level in enumerate(included_levels, start=1):
        class_rows = []
        for academic_class in classes_by_level[level.id]:
            class_students = students_by_level_class.get((level.id, academic_class.id), [])
            if not class_students:
                continue

            row = {
                "name": academic_class.name,
                "m": 0,
                "mAbsent": 0,
                "mApp": 0,
                "mPassed": 0,
                "mFailed": 0,
                "mPass": 0,
                "f": 0,
                "fAbsent": 0,
                "fApp": 0,
                "fPassed": 0,
                "fFailed": 0,
                "fPass": 0,
                "avg": 0,
            }
            bucket_sitting_counts = {"m": 0, "f": 0}
            bucket_pass_counts = {"m": 0, "f": 0}
            class_scores = []
            for student in class_students:
                bucket = gender_bucket(student)
                row[bucket] += 1
                if student.id not in exam_sitting_student_ids:
                    row[f"{bucket}Absent"] += 1
                    continue
                bucket_sitting_counts[bucket] += 1
                student_average = student_averages.get(student.id)
                if student_average is not None:
                    class_scores.extend(results_by_student.get(student.id, []))
                    if is_pass(student_average):
                        bucket_pass_counts[bucket] += 1

            row["mApp"] = bucket_sitting_counts["m"]
            row["mPassed"] = bucket_pass_counts["m"]
            row["mFailed"] = row["mApp"] - row["mPassed"]
            row["mPass"] = round(row["mPassed"] / row["mApp"] * 100, 2) if row["mApp"] else 0
            row["fApp"] = bucket_sitting_counts["f"]
            row["fPassed"] = bucket_pass_counts["f"]
            row["fFailed"] = row["fApp"] - row["fPassed"]
            row["fPass"] = round(row["fPassed"] / row["fApp"] * 100, 2) if row["fApp"] else 0
            row["avg"] = round(sum(class_scores) / len(class_scores), 1) if class_scores else 0
            class_rows.append(row)

        subject_rows = []
        level_subjects = {
            subject.id: subject
            for subject in (
                Subject.query.filter_by(academic_level_id=level.id, is_active=True).all()
                if legacy_report_mode
                else subjects_for_scope(exam, level_id=level.id)
            )
        }
        for subject in sorted(level_subjects.values(), key=lambda item: (item.sort_order, item.name)):
            scores = [score for _subject, score in results_by_level_subject[level.id].get(subject.id, [])]
            sat_count = len(subject_sitting_student_ids.get(subject.id, set()))
            passed_count = sum(1 for score in scores if is_pass(score))
            subject_rows.append({
                "name": subject_display_name(subject),
                "avg": round(sum(scores) / len(scores), 1) if scores else 0,
                "appeared": sat_count,
                "passed": passed_count,
                "failed": max(sat_count - passed_count, 0),
                "pass": round(passed_count / sat_count * 100, 2) if sat_count else 0,
            })

        if not class_rows or not subject_rows:
            continue
        level_name = level.name or "Level"
        class_names = " - ".join(item.name for item in classes_by_level[level.id])
        level_students = [
            student
            for academic_class in classes_by_level[level.id]
            for student in students_by_level_class.get((level.id, academic_class.id), [])
        ]
        grade_counts = defaultdict(int)
        for student in level_students:
            if student.id not in exam_sitting_student_ids:
                continue
            student_average = student_averages.get(student.id)
            if student_average is not None:
                grade_counts[grade_for_from_cache(student_average, grade_cache).get("grade", "-")] += 1
        level_scores = [
            score
            for subject_records in results_by_level_subject[level.id].values()
            for _subject, score in subject_records
        ]
        levels_data.append({
            "key": f"level-{level.id}",
            "cls": level_styles[(position - 1) % len(level_styles)],
            "index": f"{position:02d} / {len(included_levels):02d}",
            "name": level_name,
            "title": level_titles.get(level_name.strip().lower(), level_name),
            "subtitle": f"{level_name} - {class_names}",
            "year": academic_year.name,
            "exam": exam.name,
            "classes": class_rows,
            "subjects": subject_rows,
            "overall_avg": round(sum(level_scores) / len(level_scores), 1) if level_scores else 0,
            "grade_counts": dict(grade_counts),
            "grade_total": sum(grade_counts.values()),
        })
    return levels_data


def analytics_report_grade_bands(exam):
    """Return the twelve report slots, populated from the live Grade Management scale."""
    active_rows = GradeScale.query.filter(
        db_or(GradeScale.is_active.is_(True), GradeScale.is_active.is_(None))
    )
    exam_scales = (
        active_rows.filter(GradeScale.exam_id == exam.id)
        .order_by(GradeScale.sort_order, GradeScale.min_score.desc())
        .all()
    )
    global_scales = (
        active_rows.filter(GradeScale.exam_id.is_(None))
        .order_by(GradeScale.sort_order, GradeScale.min_score.desc())
        .all()
    )

    # Grade resolution gives an exam-specific row priority, then falls back to the
    # global scale. Mirror that display behaviour here so a partially customised
    # exam still shows the complete, live Grade Management scale in the report.
    exam_grade_keys = {(scale.grade or "").strip().casefold() for scale in exam_scales}
    scales = exam_scales + [
        scale
        for scale in global_scales
        if (scale.grade or "").strip().casefold() not in exam_grade_keys
    ]
    scale_by_grade = {
        (scale.grade or "").strip().casefold(): scale
        for scale in scales
    }

    # The report has a fixed twelve-card visual grid.  Its labels and fallback
    # colours reuse the system's canonical Grade Management defaults; a slot
    # without a configured scale remains visibly empty and never affects the
    # grade engine or any calculation.
    bands = []
    for default in DEFAULT_GRADE_SCALES:
        scale = scale_by_grade.get(default["grade"].casefold())
        if scale:
            bands.append({
                "g": scale.grade,
                "lo": float(scale.min_score),
                "hi": float(scale.max_score),
                "color": scale.badge_color or default["badge_color"],
                "configured": True,
            })
        else:
            bands.append({
                "g": default["grade"],
                "lo": None,
                "hi": None,
                "color": default["badge_color"],
                "configured": False,
            })
    return bands


@advanced_results_bp.route("/analytics/results-report")
def analytics_results_report():
    """Render the printable, per-level exam analytics report from live Results Hub data."""
    year_id = int_or_none(request.args.get("year_id"))
    exam_id = int_or_none(request.args.get("exam_id"))
    selected_year = get_default_academic_year(year_id)
    if not selected_year:
        flash("Please select an academic year before opening the results report.", "warning")
        return redirect(url_for("admin_advanced_results.analytics"))

    selected_exam = db.session.get(Exam, exam_id) if exam_id else get_latest_exam_for_year(selected_year)
    if not selected_exam or selected_exam.academic_year_id != selected_year.id:
        flash("Please select a valid exam for the selected academic year.", "warning")
        return redirect(url_for("admin_advanced_results.analytics", year_id=selected_year.id))

    settings = get_settings()
    return render_template(
        "admin/analytics_results_report.html",
        levels_data=build_analytics_results_report_data(selected_year, selected_exam),
        grade_bands=analytics_report_grade_bands(selected_exam),
        report_school={
            "name": settings.get("school_name") or "",
            "logo_url": stored_asset_url(settings.get("logo_path")),
        },
    )


@advanced_results_bp.route("/analytics/grade-drill-down")
def analytics_grade_drill_down():
    """Return JSON list of students who achieved a given grade, matching current filter scope."""
    year_id    = int_or_none(request.args.get("year_id"))
    exam_id    = int_or_none(request.args.get("exam_id"))
    level_id   = int_or_none(request.args.get("level_id"))
    class_id   = int_or_none(request.args.get("class_id"))
    section_id = int_or_none(request.args.get("section_id"))
    subject_id = int_or_none(request.args.get("subject_id"))
    grade_letter = (request.args.get("grade") or "").strip()

    if not year_id or not exam_id or not grade_letter:
        return jsonify({"error": "Missing required params: year_id, exam_id, grade"}), 400

    selected_year = db.session.get(AcademicYear, year_id)
    selected_exam = db.session.get(Exam, exam_id)
    if not selected_year or not selected_exam:
        return jsonify({"error": "Invalid year or exam"}), 404

    grade_cache = load_grade_scale_cache(selected_exam.id)

    students = (
        students_for_scope_query(
            selected_year.id,
            level_id=level_id,
            class_id=class_id,
            section_id=section_id,
        )
        .options(
            selectinload(Student.academic_class),
            selectinload(Student.academic_section),
        )
        .order_by(Student.full_name)
        .all()
    )
    student_ids = [s.id for s in students]

    results_query = (
        Result.query.options(selectinload(Result.subject))
        .filter(
            Result.student_id.in_(student_ids),
            Result.exam_id == selected_exam.id,
            Result.is_published.is_(True),
        )
    )
    if subject_id:
        results_query = results_query.filter_by(subject_id=subject_id)
    results = results_query.all()

    by_student = {}
    for r in results:
        if not r.subject or not r.subject.max_score:
            continue
        by_student.setdefault(r.student_id, []).append(r)

    student_map = {s.id: s for s in students}

    rows = []
    for sid, res_list in by_student.items():
        student = student_map.get(sid)
        if not student:
            continue
        total_score = sum(float(r.score or 0) for r in res_list)
        total_max   = sum(float(r.subject.max_score) for r in res_list if r.subject and r.subject.max_score)
        avg_pct     = round(total_score / total_max * 100, 2) if total_max else 0
        grade_info  = grade_for_from_cache(avg_pct, grade_cache)

        if grade_info.get("grade") != grade_letter:
            continue

        placement = enrollment_placement_for_student(student, selected_year.id) or {}
        rows.append({
            "student_id":    student.student_code or str(student.id),
            "full_name":     student.full_name or "",
            "mother_name":   student.mother_name or "",
            "class_name":    placement.get("class_name") or (student.academic_class.name if student.academic_class else ""),
            "academic_year": selected_year.name,
            "exam_type":     selected_exam.name,
            "total":         round(total_score, 2),
            "percentage":    avg_pct,
            "grade":         grade_info.get("grade", "-"),
            "grade_point":   grade_info.get("grade_point", 0.0),
        })

    rows.sort(key=lambda x: x["percentage"], reverse=True)
    for i, row in enumerate(rows, 1):
        row["row_num"] = i

    return jsonify(rows)


def build_analytics_data(results, students, exam, top_limit=5, bottom_limit=5, scoped_subjects=None):
    """Build analytics data for charts using existing grade_for logic"""
    total_students_count = len(students) if students else 0
    num_scoped_subjects = len(scoped_subjects) if scoped_subjects else 0
    expected_results = total_students_count * num_scoped_subjects

    if not results:
        return {
            "grade_distribution": {"labels": [], "counts": [], "colors": [], "total": 0},
            "subject_performance": {"labels": [], "scores": []},
            "exam_trend": {"labels": [], "scores": []},
            "pass_fail_ratio": {"pass": 0, "fail": 0, "total": 0},
            "completion_rate": {"percentage": 0.0, "actual": 0, "expected": expected_results},
            "student_pass_fail": {"pass_count": 0, "fail_count": 0, "total_students": total_students_count, "pass_pct": 0.0, "fail_pct": 0.0},
            "overall_average": 0,
            "top_performers": [],
            "bottom_performers": [],
            "total_students": total_students_count,
            "highest_score": 0,
            "lowest_score": 0,
        }
    
    grade_cache = load_grade_scale_cache(exam.id)
    students_by_id = {student.id: student for student in students}

    def cached_grade_for(score):
        return grade_for_from_cache(score, grade_cache)
    
    # Calculate percentages for each result
    percentages = []
    for result in results:
        if result.subject.max_score:
            pct = round(float(result.score) / float(result.subject.max_score) * 100, 2)
            percentages.append(pct)
    
    # Calculate overall average
    overall_average = round(sum(percentages) / len(percentages), 2) if percentages else 0
    
    # Calculate highest and lowest scores
    highest_score = round(max(percentages), 2) if percentages else 0
    lowest_score = round(min(percentages), 2) if percentages else 0
    
    # Grade distribution using actual grade scales
    grade_counts = {}
    grade_colors = {}
    for pct in percentages:
        grade_info = cached_grade_for(pct)
        grade = grade_info["grade"]
        grade_counts[grade] = grade_counts.get(grade, 0) + 1
        if grade not in grade_colors:
            grade_colors[grade] = grade_info["badge_color"]
    
    grade_labels = sorted(grade_counts.keys())
    grade_values = [grade_counts[g] for g in grade_labels]
    grade_color_list = [grade_colors[g] for g in grade_labels]
    
    # Subject-wise performance
    subject_averages = {}
    for result in results:
        subject_name = result.subject.name
        pct = round(float(result.score) / float(result.subject.max_score) * 100, 2) if result.subject.max_score else 0
        if subject_name not in subject_averages:
            subject_averages[subject_name] = []
        subject_averages[subject_name].append(pct)
    
    subject_labels = sorted(subject_averages.keys())
    subject_values = [round(sum(subject_averages[s]) / len(subject_averages[s]), 2) for s in subject_labels]
    
    # Exam trend (compare with other exams in same year)
    year_exams = Exam.query.filter_by(academic_year_id=exam.academic_year_id).order_by(Exam.created_at).all()
    exam_trend_labels = [e.name for e in year_exams]
    year_exam_ids = [year_exam.id for year_exam in year_exams]
    trend_results = (
        Result.query.options(selectinload(Result.subject))
        .filter(Result.exam_id.in_(year_exam_ids), Result.is_published.is_(True))
        .all()
        if year_exam_ids else []
    )
    trend_percentages = {}
    for result in trend_results:
        if result.subject.max_score:
            trend_percentages.setdefault(result.exam_id, []).append(
                round(float(result.score) / float(result.subject.max_score) * 100, 2)
            )
    exam_trend_values = [
        round(sum(trend_percentages.get(year_exam.id, [])) / len(trend_percentages[year_exam.id]), 2)
        if trend_percentages.get(year_exam.id) else 0
        for year_exam in year_exams
    ]
    
    # Pass/fail ratio (entry-based)
    pass_count = sum(1 for pct in percentages if cached_grade_for(pct).get("is_pass"))
    fail_count = len(percentages) - pass_count
    
    # Completion Rate calculation
    actual_results = len(results)
    completion_rate_pct = round((actual_results / expected_results * 100), 1) if expected_results > 0 else 0.0

    # Student overall averages for top/bottom performers and student-based pass/fail
    student_averages = {}
    for result in results:
        student_id = result.student_id
        pct = round(float(result.score) / float(result.subject.max_score) * 100, 2) if result.subject.max_score else 0
        if student_id not in student_averages:
            student_averages[student_id] = []
        student_averages[student_id].append(pct)
    
    student_avg_list = [(sid, round(sum(pcts) / len(pcts), 2)) for sid, pcts in student_averages.items()]
    student_avg_list.sort(key=lambda x: x[1], reverse=True)

    # Student-based pass/fail ratio (overall average per student)
    student_pass_count = sum(1 for sid, avg in student_avg_list if cached_grade_for(avg).get("is_pass"))
    student_fail_count = total_students_count - student_pass_count
    student_pass_pct = round((student_pass_count / total_students_count * 100), 1) if total_students_count > 0 else 0.0
    student_fail_pct = round((student_fail_count / total_students_count * 100), 1) if total_students_count > 0 else 0.0

    top_performers = []
    bottom_performers = []
    
    for student_id, avg in student_avg_list[:top_limit]:
        student = students_by_id.get(student_id)
        if student:
            placement = enrollment_placement_for_student(student, exam.academic_year_id) or {}
            top_performers.append({
                "name": student.full_name,
                "mother_name": student.mother_name,
                "code": student.student_code,
                "average": avg,
                "grade": cached_grade_for(avg)["grade"],
                "class_name": placement.get("class_name") or (student.academic_class.name if student.academic_class else None),
                "section_name": placement.get("section_name") or (student.academic_section.name if student.academic_section else None),
            })
    
    for student_id, avg in student_avg_list[-bottom_limit:]:
        student = students_by_id.get(student_id)
        if student:
            placement = enrollment_placement_for_student(student, exam.academic_year_id) or {}
            bottom_performers.append({
                "name": student.full_name,
                "mother_name": student.mother_name,
                "code": student.student_code,
                "average": avg,
                "grade": cached_grade_for(avg)["grade"],
                "class_name": placement.get("class_name") or (student.academic_class.name if student.academic_class else None),
                "section_name": placement.get("section_name") or (student.academic_section.name if student.academic_section else None),
            })
    
    return {
        "grade_distribution": {
            "labels": grade_labels,
            "counts": grade_values,
            "colors": grade_color_list,
            "total": sum(grade_values) if grade_values else 0,
        },
        "subject_performance": {
            "labels": subject_labels,
            "scores": subject_values,
        },
        "exam_trend": {
            "labels": exam_trend_labels,
            "scores": exam_trend_values,
        },
        "pass_fail_ratio": {
            "pass": pass_count,
            "fail": fail_count,
            "total": len(percentages),
        },
        "completion_rate": {
            "percentage": completion_rate_pct,
            "actual": actual_results,
            "expected": expected_results,
        },
        "student_pass_fail": {
            "pass_count": student_pass_count,
            "fail_count": student_fail_count,
            "total_students": total_students_count,
            "pass_pct": student_pass_pct,
            "fail_pct": student_fail_pct,
        },
        "overall_average": overall_average,
        "top_performers": top_performers,
        "bottom_performers": bottom_performers,
        "total_students": total_students_count,
        "highest_score": highest_score,
        "lowest_score": lowest_score,
    }


@advanced_results_bp.route("/grade-management")
def grade_management():
    """Grade Management page scoped per exam"""
    year_id = int_or_none(request.args.get("year_id"))
    exam_id = int_or_none(request.args.get("exam_id"))
    level_id = int_or_none(request.args.get("level_id"))
    
    # Get selected year and exam
    selected_year = db.session.get(AcademicYear, year_id) if year_id else AcademicYear.query.filter_by(is_current=True).first()
    selected_exam = db.session.get(Exam, exam_id) if exam_id else None
    
    # Get all years, exams, and levels for selectors
    years = AcademicYear.query.order_by(AcademicYear.name.desc()).all()
    exams = Exam.query.filter_by(academic_year_id=selected_year.id).order_by(Exam.id.desc()).all() if selected_year else []
    levels = AcademicLevel.query.filter_by(is_active=True).order_by(AcademicLevel.sort_order).all()
    
    # Get grade scales for the selected exam (or global if no exam selected)
    if selected_exam:
        grade_scales = GradeScale.query.filter_by(exam_id=exam_id).order_by(GradeScale.sort_order.asc(), GradeScale.min_score.desc()).all()
        using_global = len(grade_scales) == 0
        if using_global:
            grade_scales = GradeScale.query.filter_by(exam_id=None).order_by(GradeScale.sort_order.asc(), GradeScale.min_score.desc()).all()
    else:
        grade_scales = GradeScale.query.filter_by(exam_id=None).order_by(GradeScale.sort_order.asc(), GradeScale.min_score.desc()).all()
        using_global = True
    
    # Get all exams with their configuration status
    all_exams = Exam.query.order_by(Exam.id.desc()).all()
    exam_status = []
    for exam in all_exams:
        has_custom_scales = GradeScale.query.filter_by(exam_id=exam.id).count() > 0
        exam_status.append({
            "exam": exam,
            "configured": has_custom_scales,
        })
    
    # Calculate total points for selected exam
    total_points = 0
    if selected_exam and selected_exam.academic_level_id:
        subjects = Subject.query.filter_by(academic_level_id=selected_exam.academic_level_id).all()
        total_points = sum(float(s.max_score) for s in subjects)

    report_tiers = get_report_tier_configs(
        year_id=selected_year.id if selected_year else None,
        exam_id=selected_exam.id if selected_exam else None,
        level_id=level_id,
    )
    
    decimal_precision = academic_decimal_precision(get_settings())
    decimal_step = {0: "1", 1: "0.1", 2: "0.01", 3: "0.001"}[decimal_precision]
    return render_template(
        "admin/grade_management.html",
        years=years,
        exams=exams,
        levels=levels,
        selected_year=selected_year,
        selected_exam=selected_exam,
        selected_level_id=level_id,
        grade_scales=grade_scales,
        using_global=using_global,
        exam_status=exam_status,
        total_points=total_points,
        decimal_precision=decimal_precision,
        decimal_step=decimal_step,
        weak_tier_config=report_tiers["weak"],
        fail_tier_config=report_tiers["fail"],
        report_tiers=report_tiers,
        settings=get_settings(),
    )


@advanced_results_bp.route("/grade-management/save-report-tiers", methods=["POST"])
def save_report_tiers():
    """Save Weak and Fail Tier configurations for Whole-Class Report."""
    year_id = int_or_none(request.form.get("year_id"))
    exam_id = int_or_none(request.form.get("exam_id"))
    level_id = int_or_none(request.form.get("level_id"))

    weak_min = request.form.get("weak_min", "50.0")
    weak_max = request.form.get("weak_max", "59.99")
    weak_bg = request.form.get("weak_bg_hex") or request.form.get("weak_bg_color") or "#F5A400"
    weak_text = request.form.get("weak_text_hex") or request.form.get("weak_text_color") or "#ffffff"

    fail_min = request.form.get("fail_min", "0.0")
    fail_max = request.form.get("fail_max", "49.99")
    fail_bg = request.form.get("fail_bg_hex") or request.form.get("fail_bg_color") or "#DC2626"
    fail_text = request.form.get("fail_text_hex") or request.form.get("fail_text_color") or "#ffffff"

    try:
        min_w = float(weak_min)
        max_w = float(weak_max)
        min_f = float(fail_min)
        max_f = float(fail_max)
        if min_w > max_w or min_f > max_f:
            flash("Minimum percentage cannot be greater than maximum percentage.", "danger")
            return redirect(url_for("admin_advanced_results.grade_management", year_id=year_id, exam_id=exam_id, level_id=level_id))

        save_report_tier_configs(
            year_id, exam_id, level_id,
            weak_min, weak_max, weak_bg, weak_text,
            fail_min, fail_max, fail_bg, fail_text
        )
        flash("Whole-Class Report Tier configurations (Weak & Fail) saved successfully.", "success")
    except Exception as exc:
        flash(f"Could not save Whole-Class Report tier settings: {exc}", "danger")

    return redirect(url_for("admin_advanced_results.grade_management", year_id=year_id, exam_id=exam_id, level_id=level_id))


@advanced_results_bp.route("/grade-management/generate-scale", methods=["POST", "GET"])
def generate_scale():
    """Generate Scale: automatically set all grade items to active together and save."""
    year_id = int_or_none(request.form.get("year_id") or request.args.get("year_id"))
    exam_id = int_or_none(request.form.get("exam_id") or request.args.get("exam_id"))
    level_id = int_or_none(request.form.get("level_id") or request.args.get("level_id"))

    try:
        if exam_id:
            existing = GradeScale.query.filter_by(exam_id=exam_id).all()
            if not existing:
                global_scales = GradeScale.query.filter_by(exam_id=None).order_by(GradeScale.sort_order.asc(), GradeScale.min_score.desc()).all()
                for source in global_scales:
                    clone = GradeScale(
                        exam_id=exam_id,
                        grade=source.grade,
                        min_score=source.min_score,
                        max_score=source.max_score,
                        comment=source.comment,
                        grade_point=source.grade_point,
                        is_pass=source.is_pass,
                        badge_color=source.badge_color,
                        text_color=source.text_color,
                        background_color=source.background_color,
                        border_color=source.border_color,
                        sort_order=source.sort_order,
                        is_active=True
                    )
                    db.session.add(clone)
            else:
                for scale in existing:
                    scale.is_active = True
        else:
            global_scales = GradeScale.query.filter_by(exam_id=None).all()
            for scale in global_scales:
                scale.is_active = True

        db.session.commit()
        flash("Generated and activated all grade scales successfully.", "success")
    except Exception as exc:
        db.session.rollback()
        flash(f"Could not generate scale: {exc}", "danger")

    return redirect(url_for("admin_advanced_results.grade_management", year_id=year_id, exam_id=exam_id, level_id=level_id))




@advanced_results_bp.route("/grade-management/save", methods=["POST"])
def save_grade_scales():
    """Save grade scales for an exam"""
    exam_id = int_or_none(request.form.get("exam_id"))
    selected_exam = db.session.get(Exam, exam_id) if exam_id else None
    try:
        decimal_precision = int(request.form.get("academic_decimal_precision", "2"))
    except (TypeError, ValueError):
        decimal_precision = 2
    if decimal_precision not in (0, 1, 2, 3):
        flash("Decimal precision must be between 0 and 3 places.", "danger")
        return redirect(url_for("admin_advanced_results.grade_management", year_id=request.form.get("year_id"), exam_id=exam_id) if exam_id else url_for("admin_advanced_results.grade_management"))
    posted_grade_ids = []
    for key in request.form:
        if key.startswith("grade_"):
            grade_id = int_or_none(key.split("_", 1)[1])
            if grade_id:
                posted_grade_ids.append(grade_id)

    existing_custom = GradeScale.query.filter_by(exam_id=exam_id).count() if exam_id else 0
    clone_global_for_exam = bool(exam_id and posted_grade_ids and existing_custom == 0)

    try:
        precision_setting = db.session.get(Setting, "academic_decimal_precision") or Setting(key="academic_decimal_precision")
        precision_setting.value = str(decimal_precision)
        db.session.add(precision_setting)
        target_rows = []
        if clone_global_for_exam:
            source_rows = GradeScale.query.filter(GradeScale.id.in_(posted_grade_ids)).order_by(GradeScale.sort_order.asc(), GradeScale.min_score.desc()).all()
            for source in source_rows:
                clone = GradeScale(exam_id=exam_id)
                db.session.add(clone)
                target_rows.append((clone, source.id))
        else:
            query_exam_id = exam_id if exam_id else None
            target_rows = [(grade, grade.id) for grade in GradeScale.query.filter_by(exam_id=query_exam_id).all()]

        for grade, form_id in target_rows:
            grade_value = request.form.get(f"grade_{form_id}", grade.grade).strip() or grade.grade
            min_score = _decimal_form_value(f"min_{form_id}", grade.min_score)
            max_score = _decimal_form_value(f"max_{form_id}", grade.max_score)
            grade_point = _decimal_form_value(f"point_{form_id}", grade.grade_point)
            if min_score > max_score:
                flash(f"Grade {grade_value}: minimum percentage cannot be greater than maximum percentage.", "danger")
                return redirect(url_for("admin_advanced_results.grade_management", year_id=selected_exam.academic_year_id if selected_exam else request.form.get("year_id"), exam_id=exam_id) if exam_id else url_for("admin_advanced_results.grade_management"))
            grade.grade = grade_value
            grade.min_score = min_score
            grade.max_score = max_score
            grade.grade_point = grade_point
            grade.comment = request.form.get(f"comment_{form_id}", grade.comment).strip() or grade.comment
            grade.is_pass = request.form.get(f"status_{form_id}", "fail") == "pass"
            grade.badge_color = request.form.get(f"badge_color_{form_id}", grade.badge_color)
            grade.text_color = request.form.get(f"text_color_{form_id}", grade.text_color)
            grade.background_color = request.form.get(f"background_color_{form_id}", grade.background_color)
            grade.border_color = request.form.get(f"border_color_{form_id}", grade.border_color)
            grade.sort_order = int_or_none(request.form.get(f"sort_order_{form_id}")) or grade.sort_order or 0
            grade.is_active = request.form.get(f"active_{form_id}") == "on"

        # Create new grade scale if provided
        new_grade = request.form.get("new_grade", "").strip()
        if new_grade and exam_id:
            min_score = _decimal_form_value("new_min", 0)
            max_score = _decimal_form_value("new_max", 0)
            if min_score > max_score:
                flash("New grade minimum percentage cannot be greater than maximum percentage.", "danger")
                return redirect(url_for("admin_advanced_results.grade_management", year_id=selected_exam.academic_year_id if selected_exam else None, exam_id=exam_id))
            db.session.add(GradeScale(
                grade=new_grade,
                exam_id=exam_id,
                min_score=min_score,
                max_score=max_score,
                grade_point=_decimal_form_value("new_point", 0),
                comment=request.form.get("new_comment", "").strip() or "Custom grade",
                is_pass=request.form.get("new_status", "pass") == "pass",
                badge_color=request.form.get("new_badge_color") or "#2563eb",
                text_color=request.form.get("new_text_color") or "#ffffff",
                background_color=request.form.get("new_background_color") or "#eff6ff",
                border_color=request.form.get("new_border_color") or "#3b82f6",
                sort_order=len(target_rows) + 1,
                is_active=True,
            ))

        audit("Grade Management", f"Saved grade scales for exam {exam_id if exam_id else 'global'}")
        db.session.commit()
        flash("Grade scales saved successfully.", "success")
    except Exception as exc:
        db.session.rollback()
        flash(f"Could not save grade scales: {exc}", "danger")

    year_id = selected_exam.academic_year_id if selected_exam else request.form.get("year_id")
    if exam_id:
        return redirect(url_for("admin_advanced_results.grade_management", year_id=year_id, exam_id=exam_id))
    return redirect(url_for("admin_advanced_results.grade_management", year_id=year_id))


def _decimal_form_value(field_name, default=0):
    value = request.form.get(field_name)
    if value in (None, ""):
        return Decimal(str(default or 0))
    try:
        return Decimal(str(value))
    except Exception:
        return Decimal(str(default or 0))


@advanced_results_bp.route("/settings")
def results_settings():
    """Results Settings page with label and language customization"""
    settings = get_settings()
    
    # Get all unique label keys
    label_keys = db.session.query(LabelTranslation.label_key).distinct().order_by(LabelTranslation.label_key).all()
    label_keys = [k[0] for k in label_keys]
    
    # Get all unique language codes
    language_codes = db.session.query(LabelTranslation.language_code).distinct().order_by(LabelTranslation.language_code).all()
    language_codes = [k[0] for k in language_codes]
    
    # Build label translations matrix
    label_matrix = {}
    for key in label_keys:
        label_matrix[key] = {}
        for lang in language_codes:
            translation = LabelTranslation.query.filter_by(label_key=key, language_code=lang).first()
            label_matrix[key][lang] = {
                'text_value': translation.text_value if translation else '',
                'context': translation.context if translation else '',
                'id': translation.id if translation else None
            }
    
    return render_template(
        "admin/results_settings.html",
        label_keys=label_keys,
        language_codes=language_codes,
        label_matrix=label_matrix,
        settings=settings,
    )


@advanced_results_bp.route("/students-management")
def students_management():
    """Enrollment-aware Student Management listing with legacy fallback."""
    year_id = int_or_none(request.args.get("year_id"))
    level_id = int_or_none(request.args.get("level_id"))
    class_id = int_or_none(request.args.get("class_id"))
    section_id = int_or_none(request.args.get("section_id"))
    search_query = request.args.get("q", "").strip()
    status_filter = request.args.get("status_filter", "")
    page = int_or_none(request.args.get("page", 1)) or 1
    per_page = 7  # Students per page for balanced layout
    
    # Get selected year
    selected_year = db.session.get(AcademicYear, year_id) if year_id else AcademicYear.query.filter_by(is_current=True).first()
    if year_id and not selected_year:
        abort(404)
    
    # Get all years for selector
    years = AcademicYear.query.order_by(AcademicYear.name.desc()).all()
    
    # Filters use year-aware records. The legacy fields are only a fallback
    # for students that have not yet been backfilled into StudentEnrollment.
    levels = year_levels(selected_year.id) if selected_year else []
    selected_year_level = db.session.get(AcademicYearLevel, level_id) if level_id else None
    if selected_year_level and (not selected_year or selected_year_level.academic_year_id != selected_year.id):
        flash("Selected level does not belong to the selected academic year.", "warning")
        return redirect(url_for("admin_advanced_results.students_management", year_id=selected_year.id if selected_year else None))
    classes = year_classes(level_id) if selected_year and level_id else (
        [item for level in levels for item in year_classes(level.id)] if selected_year else []
    )
    selected_year_class = db.session.get(AcademicYearClass, class_id) if class_id else None
    if selected_year_class and (not selected_year or selected_year_class.academic_year_level.academic_year_id != selected_year.id):
        flash("Selected class does not belong to the selected academic year.", "warning")
        return redirect(url_for("admin_advanced_results.students_management", year_id=selected_year.id if selected_year else None))
    if selected_year_class and level_id and selected_year_class.academic_year_level_id != level_id:
        flash("Selected class does not belong to the selected level.", "warning")
        return redirect(url_for("admin_advanced_results.students_management", year_id=selected_year.id if selected_year else None, level_id=level_id))
    sections = []
    if selected_year_class and selected_year_class.legacy_class_id:
        sections = AcademicSection.query.filter_by(
            academic_class_id=selected_year_class.legacy_class_id,
            is_active=True,
        ).order_by(AcademicSection.sort_order, AcademicSection.name).all()
    if section_id and (not selected_year_class or not any(section.id == section_id for section in sections)):
        flash("Selected section does not belong to the selected academic year class.", "warning")
        return redirect(url_for("admin_advanced_results.students_management", year_id=selected_year.id if selected_year else None, level_id=level_id, class_id=class_id))
    
    # Calculate statistics for summary cards
    stats = {
        "total_students": 0,
        "secondary": 0,
        "upper_primary": 0,
        "lower_primary": 0,
        "kindergarten": 0,
        "active_students": 0,
    }
    
    if selected_year:
        scope_query = student_enrollment_scope_query(selected_year.id)
        stats["total_students"] = scope_query.count()

        # Students by year-aware level
        for level in levels:
            level_count = student_enrollment_scope_query(
                selected_year.id,
                academic_year_level_id=level.id,
            ).count()
            level_name_lower = level.name.lower()
            if "kindergarten" in level_name_lower or "kg" in level_name_lower:
                stats["kindergarten"] = level_count
            elif "upper" in level_name_lower and "primary" in level_name_lower:
                stats["upper_primary"] = level_count
            elif "lower" in level_name_lower and "primary" in level_name_lower:
                stats["lower_primary"] = level_count
            elif "primary" in level_name_lower:
                # If just "primary" without upper/lower, count as upper primary
                stats["upper_primary"] += level_count
            elif "middle" in level_name_lower:
                stats["upper_primary"] += level_count  # Count middle as upper primary for this context
            elif "secondary" in level_name_lower:
                stats["secondary"] = level_count
        
        # Active students (not locked)
        stats["active_students"] = scope_query.filter(Student.is_result_locked.is_(False)).count()
    
    # Get students with filters
    students_query = (
        student_enrollment_scope_query(
            selected_year.id,
            academic_year_level_id=level_id,
            academic_year_class_id=class_id,
            academic_section_id=section_id,
        ) if selected_year else Student.query.filter(Student.id == -1)
    )
    
    # Apply search filter
    if search_query:
        search_pattern = f"%{search_query}%"
        students_query = students_query.filter(
            db.or_(
                Student.student_code.like(search_pattern),
                Student.full_name.like(search_pattern),
                Student.mother_name.like(search_pattern)
            )
        )
    
    # Apply status filter
    if status_filter == "locked":
        students_query = students_query.filter_by(is_result_locked=True)
    elif status_filter == "active":
        students_query = students_query.filter_by(is_result_locked=False)
    
    # Get total count for pagination
    total_students = students_query.count()
    
    # Apply pagination
    students = students_query.order_by(Student.student_code).offset((page - 1) * per_page).limit(per_page).all()
    student_ids = [student.id for student in students]
    enrollments = {}
    if student_ids and selected_year:
        enrollments = {
            enrollment.student_id: enrollment
            for enrollment in StudentEnrollment.query.filter(
                StudentEnrollment.student_id.in_(student_ids),
                StudentEnrollment.academic_year_id == selected_year.id,
            ).all()
        }
    incident_counts = {}
    incident_badges = {}
    if student_ids:
        incident_counts = dict(
            db.session.query(IncidentReport.student_id, func.count(IncidentReport.id))
            .filter(IncidentReport.student_id.in_(student_ids))
            .group_by(IncidentReport.student_id)
            .all()
        )
        recent_incidents = (
            IncidentReport.query.options(selectinload(IncidentReport.severity))
            .filter(IncidentReport.student_id.in_(student_ids))
            .order_by(IncidentReport.created_at.desc(), IncidentReport.id.desc())
            .all()
        )
        for incident in recent_incidents:
            if incident.student_id in incident_badges:
                continue
            incident_badges[incident.student_id] = {
                "severity": incident.severity.name if incident.severity else "Unknown",
                "color": incident.severity.color if incident.severity and incident.severity.color else "#e11d48",
            }
    
    # Calculate pagination info
    total_pages = (total_students + per_page - 1) // per_page if total_students > 0 else 1
    has_prev = page > 1
    has_next = page < total_pages
    
    return render_template(
        "admin/students_management.html",
        years=years,
        selected_year=selected_year,
        levels=levels,
        classes=classes,
        students=students,
        selected_level_id=level_id,
        selected_class_id=class_id,
        selected_section_id=section_id,
        sections=sections,
        enrollments=enrollments,
        q=search_query,
        status_filter=status_filter,
        stats=stats,
        page=page,
        per_page=per_page,
        total_students=total_students,
        total_pages=total_pages,
        has_prev=has_prev,
        has_next=has_next,
        incident_counts=incident_counts,
        incident_badges=incident_badges,
        settings=get_settings(),
    )


def _transition_hierarchy_payload(years):
    """Serialize the year-aware hierarchy for transfer form cascading selects."""
    payload = {}
    for year in years:
        year_levels_payload = []
        for level in year_levels(year.id):
            class_payload = []
            for year_class in year_classes(level.id):
                sections = []
                if year_class.legacy_class_id:
                    sections = AcademicSection.query.filter_by(
                        academic_class_id=year_class.legacy_class_id,
                        is_active=True,
                    ).order_by(AcademicSection.sort_order, AcademicSection.name).all()
                class_payload.append({
                    "id": year_class.id,
                    "name": year_class.name,
                    "sections": [{"id": item.id, "name": item.name} for item in sections],
                })
            year_levels_payload.append({
                "id": level.id,
                "name": level.name,
                "classes": class_payload,
            })
        payload[str(year.id)] = year_levels_payload
    return payload


def _transition_source_enrollments(student):
    return StudentEnrollment.query.filter_by(student_id=student.id).order_by(
        StudentEnrollment.academic_year_id.desc(),
        StudentEnrollment.id.desc(),
    ).all()


@advanced_results_bp.route("/students/<int:student_id>/transition", methods=["GET", "POST"])
def student_transition(student_id):
    """Move one permanent Student identity to a new year-aware placement."""
    student = db.session.get(Student, student_id)
    if not student:
        abort(404)
    years = AcademicYear.query.order_by(AcademicYear.name.desc()).all()
    source_enrollments = _transition_source_enrollments(student)
    source_id = int_or_none(request.values.get("source_enrollment_id"))
    source = db.session.get(StudentEnrollment, source_id) if source_id else (source_enrollments[0] if source_enrollments else None)
    action = request.values.get("action", "local_transfer")
    error = None

    if request.method == "POST":
        source_id = int_or_none(request.form.get("source_enrollment_id"))
        action = request.form.get("action", action)
        destination_year_id = int_or_none(request.form.get("destination_academic_year_id"))
        destination_level_id = int_or_none(request.form.get("destination_academic_year_level_id"))
        destination_class_id = int_or_none(request.form.get("destination_academic_year_class_id"))
        destination_section_id = int_or_none(request.form.get("destination_academic_section_id"))
        if request.form.get("confirm_transition") != "on":
            error = "Please confirm that the source enrollment will be preserved and a new placement will be created."
        else:
            try:
                # Older records may still have only the legacy Student
                # placement fields. Convert that one unambiguous source on
                # submit so the transition remains fully historical without
                # guessing when the mapping is ambiguous.
                if not source_id:
                    source = ensure_legacy_enrollment_for_student(student)
                    source_id = source.id
                source, destination = transition_student_enrollment(
                    student.id,
                    source_id,
                    destination_year_id,
                    destination_level_id,
                    destination_class_id,
                    destination_section_id,
                    action=action,
                    notes=request.form.get("notes") or None,
                    performed_by=current_user.id if current_user.is_authenticated else None,
                )
                db.session.commit()
                flash(
                    f"{student.full_name} moved successfully to {destination.academic_year.name} — {destination.academic_year_class.name}.",
                    "transition-success",
                )
                return redirect(url_for("admin_advanced_results.students_management", year_id=destination.academic_year_id))
            except (EnrollmentValidationError, ValueError) as exc:
                db.session.rollback()
                error = str(exc)
            except Exception:
                db.session.rollback()
                current_app.logger.exception("Student transition failed")
                error = "The transition could not be saved. No academic placement was changed."
        source = db.session.get(StudentEnrollment, source_id) if source_id else source

    selected_enrollment = source
    return render_template(
        "admin/student_transition.html",
        student=student,
        selected_enrollment=selected_enrollment,
        years=years,
        source_enrollments=source_enrollments,
        selected_source=source,
        hierarchy=_transition_hierarchy_payload(years),
        initial_action=action,
        movement_history=StudentEnrollmentMovement.query.filter_by(student_id=student.id)
        .order_by(StudentEnrollmentMovement.moved_at.desc(), StudentEnrollmentMovement.id.desc())
        .limit(50)
        .all(),
        error=error,
        settings=get_settings(),
    )


@advanced_results_bp.route("/student-transitions/class", methods=["GET", "POST"])
def class_transition():
    """Preview and execute a controlled whole-class transition."""
    years = AcademicYear.query.order_by(AcademicYear.name.desc()).all()
    values = request.form if request.method == "POST" else request.args
    source_year_id = int_or_none(values.get("source_academic_year_id"))
    source_level_id = int_or_none(values.get("source_academic_year_level_id"))
    source_class_id = int_or_none(values.get("source_academic_year_class_id"))
    source_section_id = int_or_none(values.get("source_academic_section_id"))
    destination_year_id = int_or_none(values.get("destination_academic_year_id"))
    destination_level_id = int_or_none(values.get("destination_academic_year_level_id"))
    destination_class_id = int_or_none(values.get("destination_academic_year_class_id"))
    destination_section_id = int_or_none(values.get("destination_academic_section_id"))
    action = values.get("action") or "promotion"
    excluded_student_ids = []
    error = None
    plan = None
    preview = None

    def build_plan():
        return plan_bulk_transition(
            source_year_id,
            source_level_id,
            source_class_id,
            destination_year_id,
            destination_level_id,
            destination_class_id,
            source_academic_section_id=source_section_id,
            destination_academic_section_id=destination_section_id,
            action=action,
            excluded_student_ids=excluded_student_ids,
        )

    def parse_exclusions():
        parsed = []
        for raw_id in values.getlist("excluded_student_ids"):
            student_id = int_or_none(raw_id)
            if student_id is None:
                raise EnrollmentValidationError("Excluded student ID must be valid")
            parsed.append(student_id)
        return sorted(set(parsed))

    if all(value is not None for value in (
        source_year_id, source_level_id, source_class_id,
        destination_year_id, destination_level_id, destination_class_id,
    )):
        try:
            excluded_student_ids = parse_exclusions()
            plan = build_plan()
            items = plan["items"]
            preview = {
                "total": len(items),
                "ready": sum(item["classification"] == "ELIGIBLE" for item in items),
                "eligible": sum(item["classification"] == "ELIGIBLE" for item in items),
                "skipped": sum(item["classification"] == "SKIPPED" for item in items),
                "excluded": sum(item["classification"] == "EXCLUDED" for item in items),
                "invalid": sum(item["classification"] == "INVALID" for item in items),
                "items": items,
            }
            if request.method == "POST" and request.form.get("mode") == "execute":
                if request.form.get("confirm_transition") != "on":
                    error = "Please confirm the reviewed student list before executing the transition."
                else:
                    created = execute_bulk_transition(
                        plan,
                        action=action,
                        notes=request.form.get("notes") or None,
                        performed_by=current_user.id if current_user.is_authenticated else None,
                    )
                    db.session.commit()
                    flash(
                        f"Class transition completed: {len(created)} transitioned, "
                        f"{preview['skipped']} skipped, {preview['excluded']} excluded, "
                        f"{preview['invalid']} invalid.",
                        "success",
                    )
                    return redirect(url_for("admin_advanced_results.students_management", year_id=destination_year_id))
        except (EnrollmentValidationError, ValueError) as exc:
            db.session.rollback()
            error = str(exc)
        except Exception:
            db.session.rollback()
            current_app.logger.exception("Whole-class transition failed")
            error = "The class transition could not be prepared. No records were changed."
    elif request.method == "POST" and request.form.get("mode") == "preview":
        error = "Select the complete source and destination hierarchy before previewing the transition."

    return render_template(
        "admin/class_transition.html",
        years=years,
        hierarchy=_transition_hierarchy_payload(years),
        values={
            "source_year_id": source_year_id,
            "source_level_id": source_level_id,
            "source_class_id": source_class_id,
            "source_section_id": source_section_id,
            "destination_year_id": destination_year_id,
            "destination_level_id": destination_level_id,
            "destination_class_id": destination_class_id,
            "destination_section_id": destination_section_id,
            "action": action,
            "excluded_student_ids": excluded_student_ids,
        },
        preview=preview,
        error=error,
        settings=get_settings(),
    )


@advanced_results_bp.route("/students/new", methods=["GET", "POST"])
@advanced_results_bp.route("/students/<int:student_id>/edit", methods=["GET", "POST"])
def student_form(student_id=None):
    """Create a permanent identity or edit identity-only student fields.

    Academic placement is created as a StudentEnrollment for new students and
    is intentionally read-only for existing students until the transfer/
    promotion workflow is introduced in Phase 2D.
    """
    student = db.session.get(Student, student_id) if student_id else Student()
    if student_id and not student:
        abort(404)
    requested_year_id = int_or_none(request.args.get("year_id"))
    selected_year = (
        db.session.get(AcademicYear, requested_year_id)
        if requested_year_id
        else (db.session.get(AcademicYear, student.academic_year_id) if student.id and student.academic_year_id else AcademicYear.query.filter_by(is_current=True).first())
    )
    if requested_year_id and not selected_year:
        abort(404)
    if request.method == "POST":
        try:
            saved_scope = save_student_from_form(student)
        except ValueError as exc:
            db.session.rollback()
            flash(str(exc), "danger")
            return redirect(request.url)
        except EnrollmentValidationError as exc:
            db.session.rollback()
            flash(str(exc), "danger")
            return redirect(request.url)
        audit("Student Updates", f"Saved student {student.student_code}")
        db.session.commit()
        flash("Student saved successfully.", "success")
        return redirect(url_for(
            "admin_advanced_results.students_management",
            year_id=saved_scope.get("academic_year_id") if saved_scope else student.academic_year_id,
            level_id=saved_scope.get("academic_year_level_id") if saved_scope else None,
            class_id=saved_scope.get("academic_year_class_id") if saved_scope else None,
        ))

    selected_enrollment = get_enrollment_for_student_year(student.id, selected_year.id) if student.id and selected_year else None
    year_levels_for_form = year_levels(selected_year.id) if selected_year else []
    selected_year_level_id = selected_enrollment.academic_year_level_id if selected_enrollment else None
    selected_year_class_id = selected_enrollment.academic_year_class_id if selected_enrollment else None
    selected_section_id = selected_enrollment.academic_section_id if selected_enrollment else None
    if (
        student.id
        and not selected_enrollment
        and selected_year
        and student.academic_year_id == selected_year.id
    ):
        legacy_level = next((item for item in year_levels_for_form if item.legacy_level_id == student.academic_level_id), None)
        selected_year_level_id = legacy_level.id if legacy_level else None
        if selected_year_level_id:
            legacy_class = db.session.get(AcademicClass, student.academic_class_id) if student.academic_class_id else None
            year_class = next((item for item in year_classes(selected_year_level_id) if legacy_class and item.legacy_class_id == legacy_class.id), None)
            selected_year_class_id = year_class.id if year_class else None
        selected_section_id = student.academic_section_id
    year_class_map = {
        level.id: year_classes(level.id)
        for level in year_levels_for_form
    }
    form_sections = []
    if selected_year_class_id:
        selected_year_class = db.session.get(AcademicYearClass, selected_year_class_id)
        if selected_year_class and selected_year_class.legacy_class_id:
            form_sections = AcademicSection.query.filter_by(
                academic_class_id=selected_year_class.legacy_class_id,
                is_active=True,
            ).order_by(AcademicSection.sort_order, AcademicSection.name).all()
    incident_reports = IncidentReport.query.filter_by(student_id=student.id).order_by(IncidentReport.created_at.desc()).limit(10).all() if student.id else []
    return render_template(
        "admin/student_form.html",
        student=student,
        years=AcademicYear.query.order_by(AcademicYear.name.desc()).all(),
        selected_year=selected_year,
        year_levels=year_levels_for_form,
        year_class_map=year_class_map,
        form_sections=form_sections,
        selected_year_level_id=selected_year_level_id,
        selected_year_class_id=selected_year_class_id,
        selected_section_id=selected_section_id,
        placement_locked=bool(student.id),
        incident_reports=incident_reports,
        student_form_action=url_for("admin_advanced_results.student_form", student_id=student.id) if student.id else url_for("admin_advanced_results.student_form"),
    )


@advanced_results_bp.route("/students/<int:student_id>/delete", methods=["POST"])
def delete_student(student_id):
    student = db.session.get(Student, student_id) or abort(404)
    if StudentEnrollment.query.filter_by(student_id=student.id).first():
        flash("Student has academic enrollment history and cannot be deleted from this screen.", "warning")
        return redirect(url_for("admin_advanced_results.students_management", year_id=student.academic_year_id))
    deleted_student = {
        "name": student.full_name,
        "code": student.student_code,
    }
    db.session.delete(student)
    audit("Student Updates", f"Deleted student {student.student_code}")
    db.session.commit()
    # Keep the successful result specific to this destructive action instead of
    # relying on a generic toast that can be missed after the list reloads.
    session["student_deleted_notice"] = deleted_student
    return redirect(url_for("admin_advanced_results.students_management"))


@advanced_results_bp.route("/students/<int:student_id>/data")
def student_data_json(student_id):
    """API endpoint to fetch student data as JSON for AJAX preview"""
    student = db.session.get(Student, student_id) or abort(404)
    selected_year_id = int_or_none(request.args.get("year_id")) or student.academic_year_id
    placement = resolve_student_academic_context(student, selected_year_id) if selected_year_id else None
    enrollment = placement.get("enrollment") if placement else None
    selected_year = db.session.get(AcademicYear, selected_year_id) if selected_year_id else None
    year_name = selected_year.name if placement and selected_year else None
    level_name = placement.get("level_name") if placement else None
    class_name = placement.get("class_name") if placement else ""
    section_name = placement.get("section_name") if placement else None
    
    return jsonify({
        "id": student.id,
        "student_code": student.student_code,
        "full_name": student.full_name,
        "mother_name": student.mother_name,
        "phone": student.phone,
        "gender": student.gender,
        "photo_path": student.photo_path,
        "photo_url": stored_asset_url(student.photo_path),
        "academic_year_level_id": placement.get("academic_year_level_id") if placement else None,
        "academic_year_level_name": level_name,
        "academic_year_class_id": placement.get("academic_year_class_id") if placement else None,
        "academic_year_class_name": class_name,
        "academic_section_id": placement.get("academic_section_id") if placement else None,
        "academic_section_name": section_name,
        "academic_year_id": placement.get("academic_year_id") if placement else None,
        "academic_year_name": year_name,
        "enrollment_status": enrollment.status if enrollment else ("legacy" if placement else None),
        "academic_outcome": enrollment.academic_outcome if enrollment else ("pending" if placement else None),
        "legacy_academic_level_id": student.academic_level_id,
        "legacy_academic_class_id": student.academic_class_id,
        "is_result_locked": student.is_result_locked,
        "is_active": student.is_active,
    })


@advanced_results_bp.route("/students/<int:student_id>/toggle-lock", methods=["POST"])
def toggle_student_lock(student_id):
    student = db.session.get(Student, student_id) or abort(404)
    if student.is_result_locked and not can("unlock_results"):
        abort(403)
    if not student.is_result_locked and not can("lock_results"):
        abort(403)
    student.is_result_locked = not student.is_result_locked
    if student.is_result_locked:
        student.lock_reason = request.form.get("lock_reason", "").strip() or "Outstanding clearance required."
        audit("Result Locking", f"Locked result for {student.student_code}")
    else:
        student.lock_reason = ""
        audit("Result Locking", f"Unlocked result for {student.student_code}")
    db.session.commit()
    flash("Result lock status updated.", "success")
    enrollment = get_enrollment_for_student_year(student.id, student.academic_year_id) if student.academic_year_id else None
    return redirect(url_for(
        "admin_advanced_results.students_management",
        year_id=enrollment.academic_year_id if enrollment else student.academic_year_id,
        level_id=enrollment.academic_year_level_id if enrollment else None,
        class_id=enrollment.academic_year_class_id if enrollment else None,
        section_id=enrollment.academic_section_id if enrollment and enrollment.academic_section_id else None,
    ))


@advanced_results_bp.route("/students/import", methods=["POST"])
def import_students():
    file = request.files.get("file")
    if not file or not allowed_file(file.filename, ALLOWED_SHEETS):
        flash("Upload a valid .xlsx file.", "danger")
        return redirect(url_for("admin_advanced_results.students_management"))
    try:
        summary = process_student_import(file)
        session["import_summary"] = summary
        audit("Import Operations", f"Student import: {summary['success_count']} saved, {summary['failed_count']} failed")
        db.session.commit()

        if summary["failed_count"] == 0 and summary["success_count"] > 0:
            flash(f"✅ {summary['success_count']} students imported successfully.", "success")
        elif summary["success_count"] > 0:
            flash(f"✅ {summary['success_count']} students imported successfully. ❌ {summary['failed_count']} rows failed validation.", "warning")
        else:
            flash(f"❌ Student import failed. {summary['failed_count']} rows failed validation.", "danger")
    except Exception as ex:
        db.session.rollback()
        flash(f"Error processing student import: {str(ex)}", "danger")

    return redirect(url_for("admin_advanced_results.students_management"))


@advanced_results_bp.route("/students/import/template")
def student_import_template():
    return workbook_response(student_template(), "student_import_template.xlsx")


@advanced_results_bp.route("/result-entry/import", methods=["POST"])
def import_results():
    file = request.files.get("file")
    year_id = request.form.get("year_id") or request.args.get("year_id") or ""
    exam_id = request.form.get("exam_id") or request.args.get("exam_id") or ""
    level_id = request.form.get("level_id") or request.args.get("level_id") or ""
    class_id = request.form.get("class_id") or request.args.get("class_id") or ""
    section_id = request.form.get("section_id") or request.args.get("section_id") or ""
    selected_year = db.session.get(AcademicYear, int_or_none(year_id))
    selected_exam = db.session.get(Exam, int_or_none(exam_id))
    selected_class = db.session.get(AcademicClass, int_or_none(class_id))

    redirect_url = url_for(
        "admin_advanced_results.result_entry",
        year_id=year_id,
        exam_id=exam_id,
        level_id=level_id,
        class_id=class_id,
        section_id=section_id
    )

    if not file or not allowed_file(file.filename, ALLOWED_SHEETS):
        return jsonify({"success": False, "error": "Upload a valid .xlsx file."}), 400

    try:
        summary = process_result_import(file)
        session["import_summary"] = summary
        audit("Import Operations", f"Result import: {summary['success_count']} rows saved, {summary['failed_count']} failed")
        db.session.commit()

        return jsonify({
            "success": True,
            "summary": summary,
            "result_filename": result_workbook_filename(
                selected_year,
                selected_exam,
                selected_class,
                results=True,
            ) if selected_year and selected_exam and selected_class else None,
        })
    except Exception as ex:
        db.session.rollback()
        current_app.logger.error("Result import failed with exception [%s]: %s", type(ex).__name__, str(ex), exc_info=True)
        return jsonify({
            "success": False,
            "error": f"Import Error ({type(ex).__name__}): {str(ex)}",
            "details": f"{type(ex).__name__}: {str(ex)}"
        }), 500


@advanced_results_bp.route("/result-entry/import/template")
def result_import_template():
    year_id = int_or_none(request.args.get("year_id"))
    exam_id = int_or_none(request.args.get("exam_id"))
    level_id = int_or_none(request.args.get("level_id"))
    class_id = int_or_none(request.args.get("class_id"))
    section_id = int_or_none(request.args.get("section_id"))
    wb = result_entry_import_template(
        year_id=year_id,
        exam_id=exam_id,
        level_id=level_id,
        class_id=class_id,
        section_id=section_id,
    )
    return workbook_response(
        wb,
        result_workbook_filename(
            db.session.get(AcademicYear, year_id),
            db.session.get(Exam, exam_id),
            db.session.get(AcademicClass, class_id),
        ),
    )



@advanced_results_bp.route("/students/export")
def export_students():
    year_id = int_or_none(request.args.get("year_id"))
    level_id = int_or_none(request.args.get("level_id"))
    class_id = int_or_none(request.args.get("class_id"))
    section_id = int_or_none(request.args.get("section_id"))
    selected_year = db.session.get(AcademicYear, year_id) if year_id else AcademicYear.query.filter_by(is_current=True).first()
    if year_id and not selected_year:
        abort(404)
    students_query = (
        student_enrollment_scope_query(selected_year.id, level_id, class_id, section_id)
        if selected_year else Student.query.filter(Student.id == -1)
    )
    students = students_query.order_by(Student.full_name).all()
    enrollment_by_student = {
        item.student_id: item
        for item in StudentEnrollment.query.filter(
            StudentEnrollment.student_id.in_([student.id for student in students]) if students else StudentEnrollment.id == -1,
            StudentEnrollment.academic_year_id == selected_year.id if selected_year else StudentEnrollment.id == -1,
        ).all()
    }
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(["student_id", "full_name", "mother_name", "phone", "gender", "academic_level", "class", "section", "academic_year", "enrollment_status", "academic_outcome", "active"])
    for student in students:
        enrollment = enrollment_by_student.get(student.id)
        class_name = enrollment.academic_year_class.name if enrollment else (student.academic_class.name if student.academic_class else (student.school_class.name if student.school_class else ""))
        level_name = enrollment.academic_year_level.name if enrollment else (student.academic_level.name if student.academic_level else student.level or "")
        section_name = enrollment.academic_section.name if enrollment and enrollment.academic_section else (student.academic_section.name if student.academic_section else student.section or "")
        year_name = enrollment.academic_year.name if enrollment else (student.academic_year.name if student.academic_year else "")
        ws.append([
            student.student_code, student.full_name, student.mother_name, student.phone,
            student.gender or "", level_name, class_name, section_name, year_name,
            enrollment.status if enrollment else "legacy", enrollment.academic_outcome if enrollment else "pending",
            student.is_active,
        ])
    filename = f"students-{_safe_download_name_part(selected_year.name if selected_year else 'all-years', 'students')}.xlsx"
    return workbook_response(wb, filename)


def save_student_from_form(student):
    is_new = student.id is None
    student.student_code = request.form["student_code"].strip()
    if not student.student_code:
        raise ValueError("Student ID is required.")
    duplicate_query = Student.query.filter(Student.student_code == student.student_code)
    if student.id:
        duplicate_query = duplicate_query.filter(Student.id != student.id)
    duplicate = duplicate_query.first()
    if duplicate:
        raise ValueError("Student ID already exists.")
    student.full_name = request.form["full_name"].strip()
    student.mother_name = request.form.get("mother_name", "").strip()
    student.phone = request.form.get("phone", "").strip()
    gender = request.form.get("gender", "").strip()
    if gender not in {"Male", "Female"}:
        raise ValueError("Please select Male or Female for the results report.")
    student.gender = gender
    student.note = request.form.get("note", "").strip()
    student.is_result_locked = bool(request.form.get("is_result_locked"))
    student.lock_reason = request.form.get("lock_reason", "").strip()
    student.is_active = bool(request.form.get("is_active"))
    photo = request.files.get("photo")
    if photo and photo.filename:
        if not allowed_file(photo.filename, ALLOWED_PHOTOS):
            raise ValueError("Photo must be JPG, PNG, or WEBP.")
        student.photo_path = upload_image(photo, "school/students")

    if not is_new:
        # Existing placement is historical data. Never overwrite it from the
        # identity edit form; Phase 2D owns transfer/promotion changes.
        return {
            "academic_year_id": student.academic_year_id,
            "academic_year_level_id": None,
            "academic_year_class_id": None,
        }

    year_id = int_or_none(request.form.get("academic_year_id"))
    year_level_id = int_or_none(request.form.get("academic_year_level_id") or request.form.get("academic_level_id"))
    year_class_id = int_or_none(request.form.get("academic_year_class_id") or request.form.get("academic_class_id"))
    section_id = int_or_none(request.form.get("academic_section_id"))
    scope = validate_enrollment_scope(year_id, year_level_id, year_class_id, section_id)
    student.academic_year_id = scope["academic_year"].id
    apply_legacy_placement(student, scope)
    db.session.add(student)
    db.session.flush()
    create_enrollment(
        student.id,
        year_id,
        year_level_id,
        year_class_id,
        section_id,
        enrollment_source="manual",
    )
    return {
        "academic_year_id": year_id,
        "academic_year_level_id": year_level_id,
        "academic_year_class_id": year_class_id,
    }


def sync_student_legacy_class(student):
    if not student.academic_class_id:
        return
    academic_class = db.session.get(AcademicClass, student.academic_class_id)
    if not academic_class:
        return
    school_class = SchoolClass.query.filter_by(name=academic_class.name).first()
    if not school_class:
        school_class = SchoolClass(name=academic_class.name)
        db.session.add(school_class)
        db.session.flush()
    student.school_class = school_class
    student.level = academic_class.academic_level.name if academic_class.academic_level else student.level
    if student.academic_section_id:
        section = db.session.get(AcademicSection, student.academic_section_id)
        student.section = section.name if section else student.section


def map_imported_student_class(student, class_name):
    school_class = SchoolClass.query.filter_by(name=class_name).first()
    if not school_class:
        school_class = SchoolClass(name=class_name)
        db.session.add(school_class)
        db.session.flush()
    student.school_class = school_class

    academic_class = AcademicClass.query.filter_by(name=class_name).first()
    if academic_class:
        student.academic_class = academic_class
        student.academic_level = academic_class.academic_level
        student.level = academic_class.academic_level.name if academic_class.academic_level else student.level


@advanced_results_bp.route("/settings/save-labels", methods=["POST"])
def save_label_translations():
    """Save label translations"""
    default_language = request.form.get("default_language", "").strip()
    if default_language:
        setting = db.session.get(Setting, "default_language") or Setting(key="default_language")
        setting.value = default_language
        db.session.add(setting)

    new_language = request.form.get("new_language", "").strip().lower()
    if new_language:
        for label_key, _, text_value, context in RESULTS_LABEL_SEEDS:
            exists = LabelTranslation.query.filter_by(label_key=label_key, language_code=new_language).first()
            if not exists:
                db.session.add(LabelTranslation(
                    label_key=label_key,
                    language_code=new_language,
                    text_value=text_value,
                    context=context,
                ))

    for key, value in request.form.items():
        if key.startswith('label_'):
            parts = key.split('_')
            if len(parts) >= 3:
                label_key = '_'.join(parts[1:-1])
                language_code = parts[-1]
                text_value = value.strip()
                
                # Find existing translation or create new
                translation = LabelTranslation.query.filter_by(label_key=label_key, language_code=language_code).first()
                if translation:
                    translation.text_value = text_value
                else:
                    db.session.add(LabelTranslation(
                        label_key=label_key,
                        language_code=language_code,
                        text_value=text_value,
                        context='Results Settings'
                    ))
    
    db.session.commit()
    flash("Label translations saved successfully.", "success")
    return redirect(url_for("admin_advanced_results.results_settings"))


def result_filters():
    return {
        "q": request.args.get("q", "").strip(),
        "year_id": int_or_none(request.args.get("year_id")),
        "exam_id": int_or_none(request.args.get("exam_id")),
        "class_id": int_or_none(request.args.get("class_id")),
        "level": request.args.get("level", "").strip(),
        "section": request.args.get("section", "").strip(),
        "status": request.args.get("status", "").strip(),
    }


def result_query(filters):
    query = Result.query.join(Result.student).join(Result.exam).join(Result.subject)
    if filters["q"]:
        q = f"%{filters['q']}%"
        query = query.filter(or_(Student.student_code.like(q), Student.full_name.like(q), Subject.name.like(q)))
    if filters["year_id"]:
        query = query.filter(Exam.academic_year_id == filters["year_id"])
    if filters["exam_id"]:
        query = query.filter(Result.exam_id == filters["exam_id"])
    if filters["class_id"]:
        query = query.filter(Student.class_id == filters["class_id"])
    if filters["level"]:
        query = query.filter(Student.level == filters["level"])
    if filters["section"]:
        query = query.filter(Student.section == filters["section"])
    if filters["status"] == "Published":
        query = query.filter(Result.is_published.is_(True))
    elif filters["status"] == "Locked":
        query = query.filter(Student.is_result_locked.is_(True))
    elif filters["status"] == "Unpublished":
        query = query.filter(Result.is_published.is_(False))
    return query


def group_results(rows):
    grouped = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(list)))))
    for row in rows:
        placement = enrollment_placement_for_student(row.student, row.exam.academic_year_id) or {}
        grouped[row.exam.academic_year.name][row.exam.name][placement.get("level_name") or row.student.level or "No Level"][placement.get("class_name") or (row.student.school_class.name if row.student.school_class else "No Class")][placement.get("section_name") or row.student.section or "No Section"].append(row)
    return grouped


def build_stats(payloads, rows):
    # Use the same active/global fallback rules as every result view.
    grade_cache = load_grade_scale_cache()

    def cached_grade_for(score):
        return grade_for_from_cache(score, grade_cache)
    
    averages = [Decimal(str(payload["average"])) for payload in payloads if payload.get("subjects")]
    pass_count = sum(1 for avg in averages if cached_grade_for(float(avg)).get("is_pass"))
    subject_totals = defaultdict(list)
    for row in rows:
        subject_totals[row.subject.name].append(round(float(Decimal(row.score) / Decimal(row.subject.max_score) * 100), 2) if row.subject.max_score else 0)
    subject_averages = {name: round(sum(values) / len(values), 2) for name, values in subject_totals.items() if values}
    ranked = sorted(payloads, key=lambda item: item.get("average", 0), reverse=True)
    return {
        "rows": len(rows),
        "students": len({row.student_id for row in rows}),
        "published": sum(1 for row in rows if row.is_published),
        "locked": len({row.student_id for row in rows if row.student.is_result_locked}),
        "pass_rate": round(pass_count / len(averages) * 100, 2) if averages else 0,
        "fail_rate": round((len(averages) - pass_count) / len(averages) * 100, 2) if averages else 0,
        "top_students": ranked[:5],
        "lowest_students": list(reversed(ranked[-5:])),
        "subject_averages": subject_averages,
        "exam_average": round(sum(averages) / len(averages), 2) if averages else 0,
    }


def distinct_values(column):
    return [value[0] for value in db.session.query(column).filter(column.isnot(None), column != "").distinct().order_by(column).all()]


def int_or_none(value):
    return int(value) if value and str(value).isdigit() else None


def workbook_response(workbook, filename):
    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    workbook.save(tmp.name)
    tmp.close()
    return send_file(
        tmp.name,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def build_dashboard_stats(exam):
    """Build statistics for the dashboard"""
    students_query = students_for_scope_query(exam.academic_year_id, exam=exam)
    student_ids = [student.id for student in students_query.all()]
    total_students = len(student_ids)

    subjects = subjects_for_scope(exam)
    total_subjects = len(subjects)
    
    expected_results = total_students * total_subjects if total_subjects > 0 else 0
    actual_results = 0
    if student_ids:
        actual_results = Result.query.filter(
            Result.exam_id == exam.id,
            Result.student_id.in_(student_ids),
            Result.is_published.is_(True),
        ).count()
    completion_percentage = round((actual_results / expected_results * 100), 2) if expected_results > 0 else 0
    
    # Get active classes
    if exam.academic_year_id:
        mapped_year_levels = year_levels(exam.academic_year_id)
        if exam.academic_level_id:
            mapped_level = next(
                (item for item in mapped_year_levels if item.legacy_level_id == exam.academic_level_id),
                None,
            )
            mapped_classes = year_classes(mapped_level.id) if mapped_level else []
            if exam.academic_class_id:
                active_classes = int(any(item.legacy_class_id == exam.academic_class_id for item in mapped_classes))
            else:
                active_classes = len(mapped_classes)
        else:
            active_classes = sum(len(year_classes(item.id)) for item in mapped_year_levels)
    elif exam.academic_class_id:
        active_classes = 1
    elif exam.academic_level_id:
        active_classes = AcademicClass.query.filter_by(academic_level_id=exam.academic_level_id, is_active=True).count()
    else:
        active_classes = AcademicClass.query.filter_by(is_active=True).count()
    
    return {
        "total_students": total_students,
        "total_subjects": total_subjects,
        "completion_percentage": completion_percentage,
        "active_classes": active_classes,
        "expected_results": expected_results,
        "actual_results": actual_results,
    }


def build_class_cards(exam, level_filter=None):
    """Build class cards for the dashboard, optionally filtered by level"""
    cards = []

    year_scope_by_level = {}
    year_scope_classes = {}
    if exam.academic_year_id:
        year_scope_by_level = {
            item.legacy_level_id: item
            for item in year_levels(exam.academic_year_id)
            if item.legacy_level_id
        }
        year_scope_classes = {
            item.legacy_class_id: item
            for scope in year_scope_by_level.values()
            for item in year_classes(scope.id)
            if item.legacy_class_id
        }
        if exam.academic_level_id and exam.academic_level_id not in year_scope_by_level:
            return []
        if exam.academic_class_id and exam.academic_class_id not in year_scope_classes:
            return []
        if exam.academic_section_id:
            exam_section = db.session.get(AcademicSection, exam.academic_section_id)
            mapped_class = year_scope_classes.get(exam.academic_class_id) if exam.academic_class_id else None
            if not exam_section or (mapped_class and mapped_class.legacy_class_id != exam_section.academic_class_id):
                return []
    
    # Determine scope based on exam configuration
    if exam.academic_section_id:
        # Single section - show section card
        section = db.session.get(AcademicSection, exam.academic_section_id)
        if section:
            cards.append(build_single_class_card(exam, section=section))
    elif exam.academic_class_id:
        # Single class - show section cards within it
        academic_class = db.session.get(AcademicClass, exam.academic_class_id)
        if academic_class:
            sections = AcademicSection.query.filter_by(academic_class_id=academic_class.id, is_active=True).all()
            for section in sections:
                cards.append(build_single_class_card(exam, section=section))
    elif exam.academic_level_id:
        # Single level - show class cards
        mapped_level = year_scope_by_level.get(exam.academic_level_id)
        classes = [
            item.legacy_class
            for item in year_classes(mapped_level.id)
            if mapped_level and item.legacy_class and item.legacy_class.is_active
        ] if mapped_level else AcademicClass.query.filter_by(academic_level_id=exam.academic_level_id, is_active=True).all()
        for cls in classes:
            cards.append(build_single_class_card(exam, academic_class=cls))
    else:
        # No scope - show level cards
        levels = [
            item.legacy_level
            for item in year_scope_by_level.values()
            if item.legacy_level and item.legacy_level.is_active
        ] if exam.academic_year_id else AcademicLevel.query.filter_by(is_active=True).order_by(AcademicLevel.sort_order).all()
        for level in levels:
            cards.append(build_single_class_card(exam, academic_level=level))
    
    # Apply level filter if provided
    if level_filter:
        cards = [card for card in cards if card.get('academic_level_id') == level_filter.id]
    
    return cards


def build_single_class_card(exam, academic_level=None, academic_class=None, section=None):
    """Build a single class/level/section card"""
    # Determine label based on what's provided
    if section:
        if not academic_class:
            academic_class = section.academic_class
        label = f"{academic_class.name} - {section.name}" if academic_class else section.name
        student_query = students_for_scope_query(
            academic_year_id=exam.academic_year_id,
            class_id=academic_class.id if academic_class else None,
            section_id=section.id,
        )
    elif academic_class:
        label = academic_class.name
        student_query = students_for_scope_query(
            academic_year_id=exam.academic_year_id,
            class_id=academic_class.id,
        )
    elif academic_level:
        label = academic_level.name
        student_query = students_for_scope_query(
            academic_year_id=exam.academic_year_id,
            level_id=academic_level.id,
        )
    else:
        label = "All Students"
        student_query = students_for_scope_query(exam.academic_year_id)
    
    student_count = student_query.count()
    
    # Calculate completion for this scope (only published results)
    student_ids = [s.id for s in student_query.all()]
    if student_ids:
        subjects = subjects_for_scope(
            exam,
            level_id=academic_level.id if academic_level else None,
            class_id=academic_class.id if academic_class else None,
        )
        expected = len(student_ids) * len(subjects)
        actual = Result.query.filter(Result.exam_id == exam.id, Result.student_id.in_(student_ids), Result.is_published.is_(True)).count()
        completion = round((actual / expected * 100), 2) if expected > 0 else 0
    else:
        completion = 0
    
    return {
        "label": label,
        "student_count": student_count,
        "completion_percentage": completion,
        "academic_level_id": academic_level.id if academic_level else None,
        "academic_class_id": academic_class.id if academic_class else None,
        "academic_section_id": section.id if section else None,
    }
