import io
import re
from io import BytesIO
from collections import Counter
from datetime import datetime
from urllib.parse import urlparse
from urllib.request import Request, urlopen
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation
from PIL import Image
from sqlalchemy.dialects.mysql import insert as mysql_insert
from . import db
from .models import (
    AcademicClass,
    AcademicLevel,
    AcademicSection,
    AcademicYear,
    AcademicYearClass,
    AcademicYearLevel,
    AcademicYearSubject,
    Exam,
    Result,
    SchoolClass,
    Student,
    Subject,
)
from .enrollment_service import (
    EnrollmentValidationError,
    apply_legacy_placement,
    create_enrollment,
    resolve_student_academic_context,
    student_enrollment_legacy_scope_query,
    validate_enrollment_scope,
)


# Keep the download readable for school staff.  The importer normalizes these
# display headers (and the earlier machine-style headers) to the same fields.
STUDENT_HEADERS = [
    "ID", "Name", "Mother", "Mobile", "Academic Level", "Class", "Section",
    "Academic Year", "Gender", "Photo Source",
]
STUDENT_REQUIRED_HEADERS = [
    "student_id", "full_name", "mother_name", "phone", "class",
    "academic_year", "gender",
]
YEAR_REGEX = re.compile(r"^\d{4}-\d{4}$")
PHOTO_SOURCE_MAX_BYTES = 8 * 1024 * 1024
PHOTO_SOURCE_TYPES = {"image/jpeg", "image/jpg", "image/png", "image/webp", "image/gif"}


def student_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(STUDENT_HEADERS)

    current_year = AcademicYear.query.filter_by(is_current=True).order_by(AcademicYear.id.desc()).first()
    active_year_name = current_year.name if current_year else ""
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{chr(64 + len(STUDENT_HEADERS))}1"
    ws.row_dimensions[1].height = 24

    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(name="Aptos", bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws["H1"].comment = Comment(
        f"Academic Year is optional per row. Blank values automatically use the current year: {active_year_name or 'none configured' }.",
        "SULTAAN EMS",
    )
    ws["J1"].comment = Comment(
        "Optional: direct HTTPS image URL. Invalid or unavailable photos do not reject the student row.",
        "SULTAAN EMS",
    )
    gender_validation = DataValidation(type="list", formula1='"Male,Female"', allow_blank=False)
    ws.add_data_validation(gender_validation)
    gender_validation.add("I2:I1000")
    for column, width in {"A": 18, "B": 28, "C": 25, "D": 20, "E": 22, "F": 20, "G": 16, "H": 18, "I": 14, "J": 48}.items():
        ws.column_dimensions[column].width = width
    if active_year_name:
        for row_number in range(2, 1001):
            ws.cell(row=row_number, column=8, value=active_year_name)
            ws.cell(row=row_number, column=8).font = Font(color="64748B", italic=True)

    guide = wb.create_sheet("Instructions")
    guide.append(["SULTAAN EMS - Student Import Guide"])
    guide.append(["Current Academic Year", active_year_name or "No current academic year configured"])
    guide.append([])
    guide.append(["Column", "Required", "What to enter", "Example"])
    guide_rows = [
        ("ID", "Yes", "Unique student ID; letters, numbers and safe symbols are accepted.", "TIS001"),
        ("Name", "Yes", "Student full name.", "Amina Ali Omar"),
        ("Mother", "Yes", "Mother/guardian name.", "Sahra Jama"),
        ("Mobile", "Yes", "Phone number. Spaces, +, 00, hyphens and parentheses are normalized; common Somali formats are accepted.", "2526177788474 / +2526177788474 / +252 61 777 8474"),
        ("Academic Level", "Yes when ambiguous", "Year-aware level from Setup.", "Secondary"),
        ("Class", "Yes", "Existing class name under the selected year and level.", "Form Four"),
        ("Section", "No", "Optional section under the selected class.", "A"),
        ("Academic Year", "No", f"Leave blank to use the current year ({active_year_name or 'configured active year'}).", active_year_name or "2026-2027"),
        ("Gender", "Yes", "Male or Female.", "Female"),
        ("Photo Source", "No", "Direct HTTPS image URL. A failed photo never rejects the student row.", "https://example.com/photo.jpg"),
    ]
    for row in guide_rows:
        guide.append(row)
    guide.freeze_panes = "A5"
    guide.column_dimensions["A"].width = 22
    guide.column_dimensions["B"].width = 12
    guide.column_dimensions["C"].width = 78
    guide.column_dimensions["D"].width = 34
    for cell in guide[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=13)
        cell.fill = header_fill
    for cell in guide[4]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    for row in guide.iter_rows(min_row=5):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    setup_values = wb.create_sheet("Setup Values")
    setup_values.append(["Academic Year", "Academic Level", "Class", "Section"])
    setup_value_rows = []
    for year in AcademicYear.query.order_by(AcademicYear.name.desc()).all():
        year_levels = AcademicYearLevel.query.filter_by(
            academic_year_id=year.id,
            is_active=True,
        ).order_by(AcademicYearLevel.sort_order, AcademicYearLevel.name).all()
        for year_level in year_levels:
            year_classes = AcademicYearClass.query.filter_by(
                academic_year_level_id=year_level.id,
                is_active=True,
            ).order_by(AcademicYearClass.sort_order, AcademicYearClass.name).all()
            for year_class in year_classes:
                sections = AcademicSection.query.filter_by(
                    academic_class_id=year_class.legacy_class_id,
                    is_active=True,
                ).order_by(AcademicSection.sort_order, AcademicSection.name).all()
                if sections:
                    setup_value_rows.extend(
                        [year.name, year_level.name, year_class.name, section.name]
                        for section in sections
                    )
                else:
                    setup_value_rows.append([year.name, year_level.name, year_class.name, ""])
    for values in setup_value_rows:
        setup_values.append(values)
    setup_values.freeze_panes = "A2"
    setup_values.auto_filter.ref = f"A1:D{max(1, setup_values.max_row)}"
    for cell in setup_values[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for column, width in {"A": 20, "B": 24, "C": 24, "D": 18}.items():
        setup_values.column_dimensions[column].width = width

    # Do not ship fake student rows that can be accidentally imported.
    return wb


def result_entry_import_template(year_id=None, exam_id=None, level_id=None, class_id=None, section_id=None):
    """Generate an Excel template for result import.

    When scope parameters are supplied the sheet is pre-filled with every
    enrolled student in the selected class and the subject columns that
    belong to that level.  The teacher only needs to enter marks.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Result Entry"

    # ── resolve scope objects ────────────────────────────────────────
    selected_year = db.session.get(AcademicYear, year_id) if year_id else None
    selected_exam = db.session.get(Exam, exam_id) if exam_id else None
    selected_level = db.session.get(AcademicLevel, level_id) if level_id else None
    selected_class = db.session.get(AcademicClass, class_id) if class_id else None
    selected_section = db.session.get(AcademicSection, section_id) if section_id else None
    exam_matches_year = not (
        selected_year
        and selected_exam
        and selected_exam.academic_year_id != selected_year.id
    )

    # Derive level from class if not explicitly given
    if selected_class and not selected_level:
        selected_level = selected_class.academic_level

    # ── resolve the exact year-aware scope ───────────────────────────
    # The legacy IDs are only selector bridges.  The downloaded workbook must
    # use the AcademicYearSubject rows that belong to this exact year/level.
    year_level = None
    year_class = None
    if selected_year:
        if selected_level:
            year_level = AcademicYearLevel.query.filter_by(
                academic_year_id=selected_year.id,
                legacy_level_id=selected_level.id,
                is_active=True,
            ).first()
        if selected_class:
            class_query = (
                AcademicYearClass.query
                .join(AcademicYearLevel, AcademicYearLevel.id == AcademicYearClass.academic_year_level_id)
                .filter(
                    AcademicYearLevel.academic_year_id == selected_year.id,
                    AcademicYearClass.legacy_class_id == selected_class.id,
                    AcademicYearClass.is_active.is_(True),
                )
            )
            if year_level:
                class_query = class_query.filter(AcademicYearClass.academic_year_level_id == year_level.id)
            class_matches = class_query.all()
            year_class = class_matches[0] if len(class_matches) == 1 else None
            year_level = year_level or (year_class.academic_year_level if year_class else None)

    # ── resolve subjects for scope ───────────────────────────────────
    effective_level_id = selected_level.id if selected_level else None
    if effective_level_id and selected_year and year_level:
        # Use the same valid year-aware subject bindings as the importer. This
        # keeps generated headers and validator resolution on one contract.
        scoped_subjects = [
            binding["year_subject"]
            for binding in _result_import_subject_bindings(selected_year, year_level)
        ]
    elif effective_level_id and selected_year:
        # An invalid year/level pairing must produce an empty safe template,
        # never a workbook widened to another year's global subjects.
        scoped_subjects = []
    elif effective_level_id:
        scoped_subjects = Subject.query.filter_by(academic_level_id=effective_level_id).order_by(Subject.sort_order, Subject.name).all()
    else:
        scoped_subjects = []
    subject_names = [s.name for s in scoped_subjects]

    # ── header row ───────────────────────────────────────────────────
    headers = ["#", "student_id", "full_name", "mother_name", "class",
               "exam_type", "academic_year"] + subject_names
    ws.append(headers)

    # Style the header row
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(name="Inter", bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        bottom=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # ── pre-fill student rows when scope is available ────────────────
    year_name = selected_year.name if selected_year else ""
    exam_type_name = selected_exam.name if selected_exam else ""
    class_name = selected_class.name if selected_class else ""

    if selected_year and selected_exam and exam_matches_year and selected_class and year_class:
        # Build student query identical to the result entry grid
        try:
            query = student_enrollment_legacy_scope_query(
                selected_year.id,
                legacy_class_id=selected_class.id,
                academic_section_id=selected_section.id if selected_section else None,
            )
        except EnrollmentValidationError:
            # Keep the generated workbook safe and empty when the selected
            # class is not configured for this year.  Never leak another
            # year's legacy students into a result template.
            query = Student.query.filter(Student.id == -1)
        students = query.order_by(Student.full_name).all()

        locked_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        locked_font = Font(name="Inter", size=10, color="334155")
        mark_font = Font(name="Inter", size=11)

        for idx, student in enumerate(students, start=1):
            row_data = [
                idx,
                student.student_code,
                student.full_name,
                student.mother_name or "",
                class_name,
                exam_type_name,
                year_name,
            ] + ["" for _ in subject_names]  # blank mark cells
            ws.append(row_data)

            # Style the pre-filled metadata cells (lock-grey)
            row_num = ws.max_row
            for col_i in range(1, 8):  # columns A-G  (# through academic_year)
                cell = ws.cell(row=row_num, column=col_i)
                cell.fill = locked_fill
                cell.font = locked_font
                cell.alignment = Alignment(horizontal="center")
            # Style the mark cells with a subtle prompt
            for col_i in range(8, 8 + len(subject_names)):
                cell = ws.cell(row=row_num, column=col_i)
                cell.font = mark_font
                cell.alignment = Alignment(horizontal="center")
    else:
        # Never ship fake students or marks.  A template without a selected
        # scope stays metadata-only and tells the user to choose a real scope.
        ws["A2"] = "Select a matching Academic Year, Exam Type, Academic Level and Class before downloading a result template."
        ws["A2"].font = Font(name="Inter", italic=True, color="64748B")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(7, len(headers)))
        ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[2].height = 32

    # Auto-size columns for readability
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)

    return wb


HEADER_ALIASES = {
    "student_id": {"student_id", "student_code", "student_number", "student_no", "id"},
    "class": {"class", "class_name", "school_class", "grade", "fasal"},
    "academic_level": {"academic_level", "level", "level_name", "heer"},
    "section": {"section", "section_name", "qayb"},
    "exam_type": {
        "exam_type", "examtype", "exam_name", "exam", "exam_title",
    },
    "academic_year": {
        "academic_year", "academicyear", "year_name", "academic_year_name",
        "year", "academic_session",
    },
    "full_name": {"full_name", "student_name", "name", "student_full_name"},
    "mother_name": {"mother_name", "mother", "mother_s_name", "mothers_name"},
    "phone": {"phone", "phone_number", "mobile", "mobile_number", "telephone"},
    "gender": {"gender", "sex", "student_gender"},
    "photo_source": {"photo_source", "photo_url", "photo_link", "image_url", "image_link", "photo"},
    "number": {"number", "no", "row_number"},
}


def clean_str(val):
    if val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    val_str = str(val)
    val_str = val_str.replace("\ufeff", "").replace("\u200b", "").replace("\xa0", " ").strip()
    return val_str


def normalize_student_phone(value):
    """Normalize common phone formats without requiring one operator prefix."""
    raw = clean_str(value)
    if not raw:
        return ""
    compact = re.sub(r"[\s()./-]+", "", raw)
    if compact.startswith("00"):
        compact = compact[2:]
    if compact.startswith("+"):
        compact = compact[1:]
    if not compact.isdigit():
        return compact
    if compact.startswith("252"):
        national = compact[3:]
    else:
        national = compact[1:] if compact.startswith("0") else compact
    # Keep the canonical +252 representation for Somali mobile prefixes while
    # allowing other valid phone/country formats to pass through unchanged.
    if re.fullmatch(r"6\d\d{7,8}", national) or re.fullmatch(r"7\d\d{7,8}", national):
        return f"+252{national}"
    return f"+{compact}" if raw.strip().startswith("+") else compact


def fetch_photo_source(photo_source):
    """Validate an optional remote image and reuse Cloudinary when configured.

    A photo is deliberately non-blocking for the student row: callers can
    import the student even when a remote image is unavailable.
    """
    source = clean_str(photo_source)
    if not source:
        return None, None

    parsed = urlparse(source)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        return None, "Photo Source must be a valid HTTP(S) image URL."

    try:
        with urlopen(Request(source, headers={"User-Agent": "SULTAAN-EMS student importer"}), timeout=3) as response:
            content_type = (response.headers.get("Content-Type") or "").split(";", 1)[0].lower()
            payload = response.read(PHOTO_SOURCE_MAX_BYTES + 1)
    except Exception:
        return None, "Photo Source could not be reached."

    if len(payload) > PHOTO_SOURCE_MAX_BYTES:
        return None, "Photo Source is larger than the 8 MB limit."
    if content_type and content_type not in PHOTO_SOURCE_TYPES and content_type not in {"application/octet-stream", "binary/octet-stream"}:
        return None, "Photo Source is not a supported image type."
    try:
        image = Image.open(BytesIO(payload))
        image.verify()
    except Exception:
        return None, "Photo Source is not a valid readable image."

    try:
        from flask import current_app
        if current_app.config.get("CLOUDINARY_CLOUD_NAME"):
            from .cloudinary_service import upload_image
            # Upload the validated bytes we already fetched. This avoids a
            # second remote fetch by Cloudinary and keeps broken/slow URLs
            # from holding the import request open unnecessarily.
            image_file = BytesIO(payload)
            image_file.content_type = content_type or "application/octet-stream"
            image_file.filename = "student-photo"
            return upload_image(image_file, "school/students"), None
    except Exception:
        return None, "Photo Source could not be saved to image storage."

    # Remote links are already browser-safe and remain compatible with the
    # existing Student.photo_path field when Cloudinary is not configured.
    return source, None


def normalize_student_gender(value):
    """Accept the template's standard values plus concise Somali/English input."""
    normalized = " ".join(clean_str(value).lower().split())
    aliases = {
        "male": "Male",
        "m": "Male",
        "lab": "Male",
        "female": "Female",
        "f": "Female",
        "dhedig": "Female",
    }
    return aliases.get(normalized, clean_str(value).strip().title())


def normalize_header_key(raw_header):
    cleaned = clean_str(raw_header).lower()
    if not cleaned:
        return ""
    slug = re.sub(r"[\s\-]+", "_", cleaned)
    slug = re.sub(r"[^\w]+", "", slug)
    slug = slug.strip("_")

    for canonical_key, aliases in HEADER_ALIASES.items():
        if slug in aliases:
            return canonical_key
    return slug


def get_import_worksheet(wb, target_name="Result Entry"):
    target_lower = target_name.lower().strip()
    for sheet in wb.worksheets:
        if sheet.title.lower().strip() == target_lower:
            return sheet

    for sheet in wb.worksheets:
        for row in sheet.iter_rows(max_row=10, values_only=True):
            if not row:
                continue
            row_keys = {normalize_header_key(c) for c in row if c is not None}
            if "student_id" in row_keys:
                return sheet

    return wb.active


def _is_result_entry_workbook(wb):
    """Identify a result workbook before it reaches Student Import validation."""
    for sheet in wb.worksheets:
        if sheet.title.strip().casefold() == "result entry":
            return True
        for row in sheet.iter_rows(max_row=10, values_only=True):
            if not row:
                continue
            row_keys = {normalize_header_key(value) for value in row if value is not None}
            if {"student_id", "full_name", "class", "exam_type", "academic_year"}.issubset(row_keys):
                return True
    return False


def detect_header_row(ws, required_canonical_keys, max_scan_rows=10):
    best_row_idx = 1
    max_r = getattr(ws, "max_row", 1) or 1
    best_raw_headers = [clean_str(cell.value) for cell in ws[1]] if max_r >= 1 else []
    best_norm_headers = [normalize_header_key(h) for h in best_raw_headers]
    best_match_count = -1

    scan_limit = min(max_r, max_scan_rows)
    if scan_limit < 1:
        scan_limit = 1

    for row_idx in range(1, scan_limit + 1):
        row_cells = ws[row_idx]
        raw = [clean_str(cell.value) for cell in row_cells]
        norm = [normalize_header_key(h) for h in raw]
        match_count = sum(1 for req in required_canonical_keys if req in norm)

        if match_count > best_match_count:
            best_match_count = match_count
            best_row_idx = row_idx
            best_raw_headers = raw
            best_norm_headers = norm

        if match_count == len(required_canonical_keys):
            break

    return best_row_idx, best_raw_headers, best_norm_headers


def process_student_import(file):
    # Uploaded FileStorage streams can behave differently under Gunicorn than
    # in Flask's development server. Reading a stable copy prevents header
    # detection from starting at an unexpected stream position.
    if hasattr(file, "read"):
        file_obj = io.BytesIO(file.read())
        if hasattr(file, "seek"):
            file.seek(0)
    else:
        file_obj = file
    wb = load_workbook(file_obj, data_only=True)
    has_students_sheet = any(
        sheet.title.strip().casefold() == "students" for sheet in wb.worksheets
    )
    if not has_students_sheet and _is_result_entry_workbook(wb):
        return {
            "success_count": 0,
            "failed_count": 0,
            "errors": [
                "This workbook is a Result Entry template. Upload it from Result Entry -> Import Results."
            ],
            "kind": "Students",
        }
    ws = get_import_worksheet(wb, target_name="Students")

    header_row_idx, raw_headers, headers_norm = detect_header_row(
        ws, STUDENT_REQUIRED_HEADERS, max_scan_rows=10
    )

    missing = [h for h in STUDENT_REQUIRED_HEADERS if h not in headers_norm]
    if missing:
        return {
            "success_count": 0,
            "failed_count": 0,
            "errors": [f"Missing required columns in header: {', '.join(missing)}"],
            "kind": "Students"
        }

    # Pre-fetch lookup data
    existing_students = {
        clean_str(s.student_code).casefold(): s
        for s in Student.query.all()
    }
    existing_years = {y.name: y for y in AcademicYear.query.all()}
    current_year = AcademicYear.query.filter_by(is_current=True).order_by(AcademicYear.id.desc()).first()
    default_year_name = current_year.name if current_year else ""

    seen_file_ids = set()
    valid_students_to_add = []
    failed_errors = []
    photo_warnings = []
    row_results = []
    success_count = 0
    failed_count = 0
    no_photo_count = 0
    photo_attached_count = 0
    photo_failed_count = 0
    academic_year_column = headers_norm.index("academic_year") if "academic_year" in headers_norm else None

    for row_idx, row_cells in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
        if not row_cells or all(c is None or str(c).strip() == "" for c in row_cells):
            continue
        # Downloaded templates pre-fill the active year down the usable rows.
        # A row containing only that default is still an intentionally blank row.
        if (
            default_year_name
            and academic_year_column is not None
            and clean_str(row_cells[academic_year_column] if academic_year_column < len(row_cells) else "") == default_year_name
            and all(clean_str(value) == "" for index, value in enumerate(row_cells) if index != academic_year_column)
        ):
            continue

        data = {}
        for col_i, val in enumerate(row_cells):
            if col_i < len(headers_norm) and headers_norm[col_i]:
                data[headers_norm[col_i]] = clean_str(val)

        student_id = data.get("student_id", "")
        full_name = data.get("full_name", "")
        mother_name = data.get("mother_name", "")
        phone = data.get("phone", "")
        gender = normalize_student_gender(data.get("gender", ""))
        academic_level_name = data.get("academic_level", "")
        class_name = data.get("class", "")
        section_name = data.get("section", "")
        academic_year = data.get("academic_year", "") or default_year_name
        photo_source = data.get("photo_source", "")
        phone = normalize_student_phone(phone)

        row_errors = []

        # 1. student_id validation
        if not student_id:
            row_errors.append(f"Row {row_idx}: student_id is required.")
        elif student_id.casefold() in seen_file_ids:
            row_errors.append(f"Row {row_idx}: student_id '{student_id}' is duplicated within this file.")
        elif student_id.casefold() in existing_students:
            row_errors.append(f"Row {row_idx}: student_id '{student_id}' already exists in database.")
        else:
            seen_file_ids.add(student_id.casefold())

        # 2. full_name validation
        if not full_name:
            row_errors.append(f"Row {row_idx}: full_name is required.")

        # 3. mother_name validation
        if not mother_name:
            row_errors.append(f"Row {row_idx}: mother_name is required.")

        # 4. phone validation: accept supplied formats, but reject empty or
        # clearly non-phone values instead of enforcing a single prefix.
        if not phone:
            row_errors.append(f"Row {row_idx}: phone is required.")
        elif len(re.sub(r"\D", "", phone)) < 5:
            row_errors.append(f"Row {row_idx}: phone number must contain at least 5 digits.")

        # 5. gender validation
        if gender not in {"Male", "Female"}:
            row_errors.append(f"Row {row_idx}: gender must be Male or Female.")

        # 6. class validation
        if not class_name:
            row_errors.append(f"Row {row_idx}: class is required.")
        # Class is resolved only inside the selected academic year and level.

        # 7. academic_year validation
        if not academic_year:
            row_errors.append(f"Row {row_idx}: academic_year is required.")
        elif not YEAR_REGEX.match(academic_year):
            row_errors.append(f"Row {row_idx}: academic_year format invalid (must be YYYY-YYYY).")
        elif academic_year not in existing_years:
            row_errors.append(f"Row {row_idx}: academic_year '{academic_year}' does not exist in database.")

        if row_errors:
            failed_count += 1
            failed_errors.extend(row_errors)
            row_results.append({"row": row_idx, "status": "invalid", "detail": "; ".join(row_errors)})
        else:
            year_obj = existing_years[academic_year]
            # Import only from the currently usable year-aware setup records.
            # Archived rows must not make an otherwise valid class ambiguous.
            level_query = AcademicYearLevel.query.filter(
                AcademicYearLevel.academic_year_id == year_obj.id,
                AcademicYearLevel.is_active.is_(True),
            )
            if academic_level_name:
                level_query = level_query.filter(AcademicYearLevel.name.ilike(academic_level_name))
            level_candidates = level_query.all()
            if not academic_level_name:
                level_candidates = AcademicYearLevel.query.filter_by(
                    academic_year_id=year_obj.id,
                    is_active=True,
                ).all()
            class_candidates = []
            if academic_level_name:
                class_candidates = AcademicYearClass.query.join(AcademicYearLevel).filter(
                    AcademicYearLevel.academic_year_id == year_obj.id,
                    AcademicYearLevel.is_active.is_(True),
                    AcademicYearLevel.name.ilike(academic_level_name),
                    AcademicYearClass.name.ilike(class_name),
                    AcademicYearClass.is_active.is_(True),
                ).all()
            else:
                class_candidates = AcademicYearClass.query.join(AcademicYearLevel).filter(
                    AcademicYearLevel.academic_year_id == year_obj.id,
                    AcademicYearLevel.is_active.is_(True),
                    AcademicYearClass.name.ilike(class_name),
                    AcademicYearClass.is_active.is_(True),
                ).all()
            legacy_only_class = None
            if len(class_candidates) != 1:
                # Older installations may have no year-aware hierarchy at all.
                # Preserve that compatibility path only when the selected year
                # has zero year-aware levels; never use it to bypass a partial
                # or conflicting modern setup.
                year_has_scoped_setup = AcademicYearLevel.query.filter_by(
                    academic_year_id=year_obj.id,
                ).first()
                legacy_candidates = AcademicClass.query.filter_by(
                    name=class_name,
                    is_active=True,
                ).all()
                legacy_level_matches = [
                    item for item in legacy_candidates
                    if not academic_level_name
                    or (item.academic_level and _import_label_key(item.academic_level.name) == _import_label_key(academic_level_name))
                ]
                if not year_has_scoped_setup and len(legacy_level_matches) == 1:
                    legacy_only_class = legacy_level_matches[0]
                else:
                    row_errors.append(
                        f"Row {row_idx}: class '{class_name}' is ambiguous or missing for academic year '{academic_year}'. Provide the year-aware Academic Level."
                    )
            elif academic_level_name and len(level_candidates) != 1:
                row_errors.append(f"Row {row_idx}: academic level '{academic_level_name}' is missing or ambiguous for academic year '{academic_year}'.")
            if row_errors:
                failed_count += 1
                failed_errors.extend(row_errors)
                row_results.append({"row": row_idx, "status": "invalid", "detail": "; ".join(row_errors)})
                continue
            year_class = class_candidates[0] if class_candidates else None
            year_level = year_class.academic_year_level if year_class else None
            section = None
            if section_name:
                section_class_id = year_class.legacy_class_id if year_class else legacy_only_class.id
                if not section_class_id:
                    row_errors.append(f"Row {row_idx}: section cannot be resolved for class '{class_name}'.")
                else:
                    section = AcademicSection.query.filter_by(
                        academic_class_id=section_class_id,
                        name=section_name,
                        is_active=True,
                    ).first()
                    if not section:
                        row_errors.append(f"Row {row_idx}: section '{section_name}' does not belong to class '{class_name}'.")
            if row_errors:
                failed_count += 1
                failed_errors.extend(row_errors)
                row_results.append({"row": row_idx, "status": "invalid", "detail": "; ".join(row_errors)})
                continue
            scope = None
            if not legacy_only_class:
                scope = validate_enrollment_scope(year_obj.id, year_level.id, year_class.id, section.id if section else None)
            photo_path = None
            if photo_source:
                photo_path, photo_error = fetch_photo_source(photo_source)
                if photo_path:
                    photo_attached_count += 1
                else:
                    photo_failed_count += 1
                    warning = f"Row {row_idx}: photo skipped - {photo_error}"
                    photo_warnings.append(warning)
            else:
                no_photo_count += 1

            new_student = Student(
                student_code=student_id,
                full_name=full_name,
                mother_name=mother_name,
                phone=phone,
                gender=gender,
                academic_year=year_obj,
                photo_path=photo_path,
                is_active=True
            )
            if legacy_only_class:
                new_student.academic_year_id = year_obj.id
                new_student.academic_class_id = legacy_only_class.id
                new_student.academic_level_id = legacy_only_class.academic_level_id
                new_student.level = legacy_only_class.academic_level.name if legacy_only_class.academic_level else academic_level_name
                new_student.section = section.name if section else None
                new_student.academic_section_id = section.id if section else None
                school_class = SchoolClass.query.filter_by(name=legacy_only_class.name).first()
                if not school_class:
                    school_class = SchoolClass(name=legacy_only_class.name)
                    db.session.add(school_class)
                    db.session.flush()
                new_student.class_id = school_class.id
            else:
                apply_legacy_placement(new_student, scope)

            valid_students_to_add.append((new_student, scope, section.id if section else None))
            success_count += 1
            row_results.append({
                "row": row_idx,
                "status": "imported",
                "photo": "attached" if photo_path else ("failed" if photo_source else "none"),
            })

    if valid_students_to_add:
        try:
            db.session.add_all([
                student for student, _scope, _section_id in valid_students_to_add
            ])
            db.session.flush()
            for student, scope, section_id in valid_students_to_add:
                if scope:
                    create_enrollment(
                        student.id,
                        scope["academic_year"].id,
                        scope["academic_year_level"].id,
                        scope["academic_year_class"].id,
                        section_id,
                        enrollment_source="import",
                    )
            db.session.commit()
        except Exception as ex:
            db.session.rollback()
            failed_errors.append(f"Database error while saving students: {str(ex)}")
            return {
                "success_count": 0,
                "failed_count": success_count + failed_count,
                "errors": failed_errors,
                "photo_warnings": photo_warnings,
                "row_results": row_results,
                "no_photo_count": no_photo_count,
                "photo_attached_count": photo_attached_count,
                "photo_failed_count": photo_failed_count,
                "kind": "Students"
            }

    return {
        "success_count": success_count,
        "failed_count": failed_count,
        "errors": failed_errors,
        "photo_warnings": photo_warnings,
        "row_results": row_results,
        "no_photo_count": no_photo_count,
        "photo_attached_count": photo_attached_count,
        "photo_failed_count": photo_failed_count,
        "kind": "Students"
    }

def _import_label_key(value):
    return " ".join(clean_str(value).casefold().split())


def _resolve_result_import_scope(year_obj, level_id=None, class_id=None, section_id=None):
    """Resolve legacy selector IDs into one exact year-aware import scope."""
    if not year_obj:
        return None, None, None

    year_level = None
    if level_id not in (None, ""):
        year_level = AcademicYearLevel.query.filter_by(
            academic_year_id=year_obj.id,
            legacy_level_id=int(level_id),
            is_active=True,
        ).one_or_none()
        if not year_level:
            raise EnrollmentValidationError("Academic level is not configured for this academic year")

    year_class = None
    if class_id not in (None, ""):
        class_query = (
            AcademicYearClass.query
            .join(AcademicYearLevel, AcademicYearLevel.id == AcademicYearClass.academic_year_level_id)
            .filter(
                AcademicYearLevel.academic_year_id == year_obj.id,
                AcademicYearClass.legacy_class_id == int(class_id),
                AcademicYearClass.is_active.is_(True),
            )
        )
        if year_level:
            class_query = class_query.filter(AcademicYearClass.academic_year_level_id == year_level.id)
        matches = class_query.all()
        if len(matches) != 1:
            raise EnrollmentValidationError("Academic class is ambiguous or not configured for this academic year")
        year_class = matches[0]
        year_level = year_level or year_class.academic_year_level

    section = None
    if section_id not in (None, ""):
        section = db.session.get(AcademicSection, int(section_id))
        if not section:
            raise EnrollmentValidationError("Academic section does not exist")
        if not year_class or year_class.legacy_class_id != section.academic_class_id:
            raise EnrollmentValidationError("Academic section does not belong to the selected academic class")

    return year_level, year_class, section


def _result_import_subject_bindings(year_obj=None, year_level=None):
    """Return only subjects usable by the selected year/level.

    A binding keeps the year-aware display name/max mark alongside the legacy
    subject identity used by the existing Result table.
    """
    bindings = []
    if year_obj:
        query = AcademicYearSubject.query.filter_by(
            academic_year_id=year_obj.id,
            subject_kind="exam",
            is_active=True,
        )
        if year_level:
            query = query.filter_by(academic_year_level_id=year_level.id)
        for year_subject in query.order_by(
            AcademicYearSubject.sort_order,
            AcademicYearSubject.name,
            AcademicYearSubject.id,
        ).all():
            legacy_subject = db.session.get(Subject, year_subject.legacy_subject_id) if year_subject.legacy_subject_id else None
            if legacy_subject:
                bindings.append({"year_subject": year_subject, "subject": legacy_subject})
        return bindings

    return [
        {"year_subject": None, "subject": subject}
        for subject in Subject.query.filter(Subject.is_active.is_(True)).all()
        if subject.name
    ]


def _result_import_max_score(binding, exam, year_level=None):
    """Resolve the same exam-aware maximum used by Result Entry."""
    year_subject = binding.get("year_subject")
    if exam:
        from .services import get_exam_marking_configuration

        configuration = get_exam_marking_configuration(
            exam,
            academic_year_level_id=year_level.id if year_level else None,
            academic_level_id=binding["subject"].academic_level_id,
        )
        if configuration:
            return float(configuration.default_full_marks)
    if year_subject:
        return float(year_subject.max_score or 0)
    return float(binding["subject"].max_score or 0)


def process_result_import(
    file,
    *,
    year_id=None,
    exam_id=None,
    level_id=None,
    class_id=None,
    section_id=None,
):
    """Import results against the exact selected academic context.

    The optional scope arguments are supplied by the Result Entry page.  The
    no-scope path remains for older integrations, but it still resolves each
    row's student/year placement before accepting a subject mark.
    """
    if hasattr(file, "read"):
        file_obj = io.BytesIO(file.read())
        if hasattr(file, "seek"):
            file.seek(0)
    else:
        file_obj = file

    wb = load_workbook(file_obj, data_only=True)
    ws = get_import_worksheet(wb, target_name="Result Entry")
    # Whole-class exports use one sheet containing students from multiple
    # classes/levels.  Their row class and enrolled year-aware level are the
    # scope for each row; a page-level class selector must not force every row
    # through one class and make valid subject columns appear ambiguous.
    is_all_results_workbook = _import_label_key(ws.title) == "all results"
    required_fields = ["student_id", "class", "exam_type", "academic_year"]
    header_row_idx, raw_headers, headers_norm = detect_header_row(ws, required_fields, max_scan_rows=10)
    missing = [field for field in required_fields if field not in headers_norm]
    if missing:
        return {
            "success_count": 0,
            "failed_count": 0,
            "errors": [f"Missing required headers in file: {', '.join(missing)}"],
            "kind": "Results",
        }

    from sqlalchemy.orm import joinedload

    selected_year = db.session.get(AcademicYear, int(year_id)) if year_id not in (None, "") else None
    selected_exam = db.session.get(Exam, int(exam_id)) if exam_id not in (None, "") else None
    if year_id not in (None, "") and not selected_year:
        return {"success_count": 0, "failed_count": 0, "errors": ["Selected academic year does not exist."], "kind": "Results"}
    if exam_id not in (None, "") and not selected_exam:
        return {"success_count": 0, "failed_count": 0, "errors": ["Selected exam type does not exist."], "kind": "Results"}
    if selected_exam and selected_year and selected_exam.academic_year_id != selected_year.id:
        return {"success_count": 0, "failed_count": 0, "errors": ["Selected exam type does not belong to the selected academic year."], "kind": "Results"}
    selected_year = selected_year or (selected_exam.academic_year if selected_exam else None)

    effective_level_id = level_id or (selected_exam.academic_level_id if selected_exam else None)
    effective_class_id = class_id or (selected_exam.academic_class_id if selected_exam else None)
    try:
        selected_year_level, selected_year_class, selected_section = _resolve_result_import_scope(
            selected_year,
            None if is_all_results_workbook else effective_level_id,
            None if is_all_results_workbook else effective_class_id,
            None if is_all_results_workbook else section_id,
        )
    except (EnrollmentValidationError, ValueError) as exc:
        return {"success_count": 0, "failed_count": 0, "errors": [str(exc)], "kind": "Results"}

    bindings = _result_import_subject_bindings(selected_year, selected_year_level)

    # The YearAcademicSubject name is the canonical Result Entry column name.
    # A legacy Subject name is only a compatibility alias.  Keeping them in
    # separate maps prevents an old legacy name from making a valid, renamed
    # year-aware subject column look ambiguous.
    year_subject_keys = {}
    legacy_subject_keys = {}

    def add_subject_key(mapping, label, binding):
        for candidate in {label, normalize_header_key(label)}:
            key = _import_label_key(candidate)
            if not key:
                continue
            entries = mapping.setdefault(key, [])
            if not any(item["year_subject"].id == binding["year_subject"].id for item in entries):
                entries.append(binding)

    for binding in bindings:
        year_subject = binding["year_subject"]
        subject = binding["subject"]
        if year_subject:
            add_subject_key(year_subject_keys, year_subject.name, binding)
        add_subject_key(legacy_subject_keys, subject.name, binding)

    fixed_header_set = {"#", "student_id", "full_name", "mother_name", "class", "exam_type", "academic_year"}
    subject_cols = {}
    unmapped_subject_columns = {}
    for col_idx, header in enumerate(raw_headers):
        normalized = headers_norm[col_idx] if col_idx < len(headers_norm) else ""
        if not header or normalized in fixed_header_set or _import_label_key(header) in fixed_header_set:
            continue
        header_keys = [_import_label_key(header), _import_label_key(normalize_header_key(header))]
        candidates = []
        # Exact year-aware names always win.  Only older workbooks that do not
        # contain that name use the legacy Subject alias as a fallback.
        for key in header_keys:
            if key and year_subject_keys.get(key):
                candidates = year_subject_keys[key]
                break
        if not candidates:
            for key in header_keys:
                if key and legacy_subject_keys.get(key):
                    candidates = legacy_subject_keys[key]
                    break
        if candidates:
            subject_cols[col_idx] = candidates
        else:
            unmapped_subject_columns[col_idx] = header

    existing_students = {
        clean_str(student.student_code).casefold(): student
        for student in Student.query.options(
            joinedload(Student.academic_class),
            joinedload(Student.school_class),
        ).all()
    }
    existing_years = {year.name: year for year in AcademicYear.query.all()}
    existing_exams = {
        (_import_label_key(exam.name), exam.academic_year_id): exam
        for exam in Exam.query.all()
    }

    success_count = 0
    failed_count = 0
    failed_errors = []
    valid_row_entries = []

    for row_idx, row_cells in enumerate(
        ws.iter_rows(min_row=header_row_idx + 1, values_only=True),
        start=header_row_idx + 1,
    ):
        try:
            if not row_cells or all(c is None or clean_str(c) == "" for c in row_cells):
                continue

            row_map = {
                headers_norm[col_idx]: clean_str(value)
                for col_idx, value in enumerate(row_cells)
                if col_idx < len(headers_norm) and headers_norm[col_idx]
            }
            student_id = row_map.get("student_id", "")
            provided_class = row_map.get("class", "")
            exam_name = row_map.get("exam_type", "")
            academic_year_name = row_map.get("academic_year", "")
            row_errors = []

            year_obj = existing_years.get(academic_year_name) if academic_year_name else None
            if not academic_year_name:
                row_errors.append(f"Row {row_idx}: academic_year is required.")
            elif not YEAR_REGEX.match(academic_year_name):
                row_errors.append(f"Row {row_idx}: academic_year format invalid (must be YYYY-YYYY).")
            elif not year_obj:
                row_errors.append(f"Row {row_idx}: academic_year '{academic_year_name}' does not exist in system.")
            elif selected_year and year_obj.id != selected_year.id:
                row_errors.append(
                    f"Row {row_idx}: academic_year '{academic_year_name}' does not match the selected academic year '{selected_year.name}'."
                )

            student_obj = existing_students.get(student_id.casefold()) if student_id else None
            if not student_id:
                row_errors.append(f"Row {row_idx}: student_id is required.")
            elif not student_obj:
                row_errors.append(f"Row {row_idx}: student_id '{student_id}' not found in system.")

            placement = (
                resolve_student_academic_context(student_obj, year_obj.id)
                if student_obj and year_obj else None
            )

            if not provided_class:
                row_errors.append(f"Row {row_idx}: class is required.")
            elif student_obj and year_obj and not placement:
                row_errors.append(
                    f"Row {row_idx}: student_id '{student_id}' has no enrollment or legacy placement for academic year '{academic_year_name}'."
                )
            elif student_obj and year_obj and placement:
                if selected_year_class and placement.get("academic_year_class_id") != selected_year_class.id:
                    row_errors.append(
                        f"Row {row_idx}: student_id '{student_id}' is not enrolled in the selected academic class."
                    )
                elif selected_section and placement.get("academic_section_id") != selected_section.id:
                    row_errors.append(
                        f"Row {row_idx}: student_id '{student_id}' is not enrolled in the selected academic section."
                    )
                elif not selected_year_class:
                    actual_class = _import_label_key(placement.get("class_name"))
                    if actual_class != _import_label_key(provided_class):
                        row_errors.append(
                            f"Row {row_idx}: class '{provided_class}' does not match student's class '{placement.get('class_name') or ''}'."
                        )

            row_exam = selected_exam
            if not exam_name:
                row_errors.append(f"Row {row_idx}: exam_type is required.")
            elif selected_exam and _import_label_key(exam_name) != _import_label_key(selected_exam.name):
                row_errors.append(
                    f"Row {row_idx}: exam_type '{exam_name}' does not match the selected exam '{selected_exam.name}'."
                )
            elif not row_exam and year_obj:
                row_exam = existing_exams.get((_import_label_key(exam_name), year_obj.id))

            if selected_year_level and placement and placement.get("academic_year_level_id") != selected_year_level.id:
                row_errors.append(
                    f"Row {row_idx}: student_id '{student_id}' is not enrolled in the selected academic level."
                )

            row_subject_marks = []
            resolved_subject_ids = set()
            student_year_level_id = placement.get("academic_year_level_id") if placement else None
            effective_year_level = selected_year_level
            if not effective_year_level and student_year_level_id:
                effective_year_level = db.session.get(AcademicYearLevel, student_year_level_id)
            for col_idx, header in unmapped_subject_columns.items():
                if col_idx < len(row_cells) and clean_str(row_cells[col_idx]) != "":
                    row_errors.append(
                        f"Row {row_idx}: subject column '{header}' is not configured for the selected academic year and level."
                    )

            for col_idx, candidates in subject_cols.items():
                if col_idx >= len(row_cells) or clean_str(row_cells[col_idx]) == "":
                    continue
                eligible = []
                for binding in candidates:
                    year_subject = binding["year_subject"]
                    subject = binding["subject"]
                    if year_subject:
                        if not year_obj or year_subject.academic_year_id != year_obj.id:
                            continue
                        # The Result Entry page already selected this year-aware
                        # level.  Enrollment mismatches are reported separately;
                        # they must not turn every otherwise-valid subject into a
                        # misleading "not uniquely configured" error.
                        target_level_id = (
                            selected_year_level.id
                            if selected_year_level
                            else student_year_level_id
                        )
                        if target_level_id and year_subject.academic_year_level_id != target_level_id:
                            continue
                    elif selected_year or student_year_level_id:
                        if not student_year_level_id or subject.academic_level_id != placement.get("academic_level_id"):
                            continue
                    eligible.append(binding)

                if not eligible:
                    header_name = raw_headers[col_idx] if col_idx < len(raw_headers) else f"column {col_idx + 1}"
                    row_errors.append(
                        f"Row {row_idx}: subject '{header_name}' is not configured for the student's academic year and level."
                    )
                    continue
                if len(eligible) > 1:
                    header_name = raw_headers[col_idx] if col_idx < len(raw_headers) else f"column {col_idx + 1}"
                    row_errors.append(
                        f"Row {row_idx}: subject '{header_name}' has multiple active configurations for the student's academic year and level."
                    )
                    continue

                binding = eligible[0]
                subject = binding["subject"]
                if subject.id in resolved_subject_ids:
                    row_errors.append(f"Row {row_idx}: subject '{subject.name}' appears more than once in the file.")
                    continue
                resolved_subject_ids.add(subject.id)
                cell_val = row_cells[col_idx]
                try:
                    score_num = float(cell_val)
                    max_score = _result_import_max_score(binding, row_exam, effective_year_level)
                    if score_num < 0 or score_num > max_score:
                        row_errors.append(
                            f"Row {row_idx}: mark for '{binding['year_subject'].name if binding['year_subject'] else subject.name}' ({score_num:g}) must be between 0 and {max_score:g}."
                        )
                    else:
                        row_subject_marks.append((subject, score_num))
                except (TypeError, ValueError):
                    row_errors.append(f"Row {row_idx}: mark for '{subject.name}' must be numeric (got '{cell_val}').")

            if row_errors:
                failed_count += 1
                failed_errors.extend(row_errors)
                continue

            if not row_exam and year_obj:
                row_exam = existing_exams.get((_import_label_key(exam_name), year_obj.id))
            valid_row_entries.append({
                "row_idx": row_idx,
                "student": student_obj,
                "exam_name": exam_name,
                "exam": row_exam,
                "year": year_obj,
                "marks": row_subject_marks,
            })
        except Exception as exc:
            failed_count += 1
            failed_errors.append(f"Row {row_idx}: Unhandled error parsing row ({str(exc)})")

    if valid_row_entries:
        for entry in valid_row_entries:
            if entry["exam"] or not entry["year"]:
                continue
            exam_key = (_import_label_key(entry["exam_name"]), entry["year"].id)
            exam_obj = existing_exams.get(exam_key)
            if not exam_obj:
                exam_obj = Exam(
                    name=entry["exam_name"].strip(),
                    academic_year_id=entry["year"].id,
                    is_active=True,
                    is_published=True,
                )
                db.session.add(exam_obj)
                db.session.flush()
                existing_exams[exam_key] = exam_obj
            entry["exam"] = exam_obj

        result_rows = []
        for entry in valid_row_entries:
            if not entry["year"] or not entry["student"] or not entry["exam"]:
                failed_count += 1
                failed_errors.append(f"Row {entry['row_idx']}: Missing required academic context.")
                continue
            now = datetime.utcnow()
            for subject, score_num in entry["marks"]:
                result_rows.append({
                    "student_id": entry["student"].id,
                    "exam_id": entry["exam"].id,
                    "subject_id": subject.id,
                    "score": score_num,
                    "is_published": True,
                    "created_at": now,
                    "updated_at": now,
                })
            success_count += 1

        if result_rows:
            if db.session.get_bind().dialect.name == "mysql":
                statement = mysql_insert(Result.__table__).values(result_rows)
                db.session.execute(statement.on_duplicate_key_update(
                    score=statement.inserted.score,
                    is_published=statement.inserted.is_published,
                    updated_at=statement.inserted.updated_at,
                ))
            else:
                student_ids = list({row["student_id"] for row in result_rows})
                exam_ids = list({row["exam_id"] for row in result_rows})
                existing_results = Result.query.filter(
                    Result.student_id.in_(student_ids),
                    Result.exam_id.in_(exam_ids),
                ).all()
                existing_results_map = {
                    (result.student_id, result.exam_id, result.subject_id): result
                    for result in existing_results
                }
                new_results = []
                for row in result_rows:
                    result_key = (row["student_id"], row["exam_id"], row["subject_id"])
                    result = existing_results_map.get(result_key)
                    if result:
                        result.score = row["score"]
                        result.is_published = True
                        result.updated_at = row["updated_at"]
                    else:
                        new_results.append(Result(**row))
                if new_results:
                    db.session.add_all(new_results)

        try:
            db.session.commit()
        except Exception as exc:
            db.session.rollback()
            return {
                "success_count": 0,
                "failed_count": success_count + failed_count,
                "errors": [f"Fatal database error during commit: {str(exc)}"],
                "kind": "Results",
            }

    return {
        "success_count": success_count,
        "failed_count": failed_count,
        "errors": failed_errors,
        "kind": "Results",
    }
