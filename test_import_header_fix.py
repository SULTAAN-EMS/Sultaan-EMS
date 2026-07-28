import io
import openpyxl
from app import create_app
from app.import_wizard import process_result_import, result_entry_import_template

def run_tests():
    app = create_app()
    with app.app_context():
        print("=== Test 1: Standard 'Get Template' Output for Form Two ===")
        # Generate Form Two template (year_id=1, exam_id=1, class_id=5)
        wb1 = result_entry_import_template(year_id=1, exam_id=1, class_id=5)
        ws1 = wb1["Result Entry"]
        
        # Fill in sample marks for students
        for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
            if row[1].value:  # student_id
                for cell in row[7:]:  # subject columns
                    cell.value = 85
                    
        buf1 = io.BytesIO()
        wb1.save(buf1)
        buf1.seek(0)
        
        res1 = process_result_import(buf1)
        print("Result 1:", res1)
        assert res1["success_count"] > 0, f"Expected >0 success, got {res1}"
        assert res1["failed_count"] == 0, f"Expected 0 failed, got {res1}"
        assert len(res1["errors"]) == 0, f"Expected no errors, got {res1['errors']}"
        print("[OK] Test 1 Passed!\n")

        print("=== Test 2: Active Sheet Misdirection (Active sheet is Summary tab) ===")
        wb2 = openpyxl.Workbook()
        ws_summary = wb2.active
        ws_summary.title = "Summary"
        ws_summary.append(["Class", "Notes", "Total Students"])
        ws_summary.append(["Form Two", "Import template", 10])
        
        # Create Result Entry sheet
        ws_data = wb2.create_sheet(title="Result Entry")
        headers = ["#", "student_id", "full_name", "mother_name", "class", "exam_type", "academic_year", "MATH", "PHYSICS"]
        ws_data.append(headers)
        ws_data.append([1, "3001", "Amina Ali Omar", "Sahra Jama", "Form One A", "Midterm", "2025-2026", 90, 88])
        
        buf2 = io.BytesIO()
        wb2.save(buf2)
        buf2.seek(0)
        
        res2 = process_result_import(buf2)
        print("Result 2:", res2)
        assert res2["success_count"] == 1, f"Expected 1 success, got {res2}"
        assert res2["failed_count"] == 0, f"Expected 0 failed, got {res2}"
        print("[OK] Test 2 Passed!\n")

        print("=== Test 3: Headers with Spaces / Capitalization ('Student ID', 'Exam Type', 'Academic Year') ===")
        wb3 = openpyxl.Workbook()
        ws3 = wb3.active
        ws3.title = "Result Entry"
        ws3.append(["#", "Student ID", "Full Name", "Mother Name", "Class", "Exam Type", "Academic Year", "MATH", "PHYSICS"])
        ws3.append([1, "3001", "Amina Ali Omar", "Sahra Jama", "Form One A", "Midterm", "2025-2026", 95, 92])
        
        buf3 = io.BytesIO()
        wb3.save(buf3)
        buf3.seek(0)
        
        res3 = process_result_import(buf3)
        print("Result 3:", res3)
        assert res3["success_count"] == 1, f"Expected 1 success, got {res3}"
        assert res3["failed_count"] == 0, f"Expected 0 failed, got {res3}"
        print("[OK] Test 3 Passed!\n")

        print("=== Test 4: Header Row Offset (Title in Row 1, Headers in Row 2) ===")
        wb4 = openpyxl.Workbook()
        ws4 = wb4.active
        ws4.title = "Result Entry"
        ws4.append(["SULTAN SCHOLASTIC EXAMINATION RESULT IMPORT FILE"])
        ws4.append(["#", "student_id", "full_name", "mother_name", "class", "exam_type", "academic_year", "MATH"])
        ws4.append([1, "3001", "Amina Ali Omar", "Sahra Jama", "Form One A", "Midterm", "2025-2026", 88])
        
        buf4 = io.BytesIO()
        wb4.save(buf4)
        buf4.seek(0)
        
        res4 = process_result_import(buf4)
        print("Result 4:", res4)
        assert res4["success_count"] == 1, f"Expected 1 success, got {res4}"
        assert res4["failed_count"] == 0, f"Expected 0 failed, got {res4}"
        print("[OK] Test 4 Passed!\n")

        print("=== Test 5: Headers with BOM / Unicode Whitespace ===")
        wb5 = openpyxl.Workbook()
        ws5 = wb5.active
        ws5.title = "Result Entry"
        ws5.append(["#", "\ufeffstudent_id", "full_name", "mother_name", "class\xa0", "exam_type", "academic_year\u200b", "MATH"])
        ws5.append([1, "3001", "Amina Ali Omar", "Sahra Jama", "Form One A", "Midterm", "2025-2026", 91])
        
        buf5 = io.BytesIO()
        wb5.save(buf5)
        buf5.seek(0)
        
        res5 = process_result_import(buf5)
        print("Result 5:", res5)
        assert res5["success_count"] == 1, f"Expected 1 success, got {res5}"
        assert res5["failed_count"] == 0, f"Expected 0 failed, got {res5}"
        print("[OK] Test 5 Passed!\n")

        print("ALL IMPORT HEADER FIX TESTS PASSED SUCCESSFULLY!")

if __name__ == "__main__":
    run_tests()
